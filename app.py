from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file, g, flash
from werkzeug.security import check_password_hash, generate_password_hash
import os, io, csv, re, sqlite3, logging
from datetime import datetime, timedelta, timezone

DUBAI_TZ = timezone(timedelta(hours=4))
def dubai_today():
    """Returns today's date in Dubai (UTC+4) regardless of server timezone."""
    return datetime.now(DUBAI_TZ).date()
from functools import wraps

try:
    import openpyxl
    HAS_XL = True
except: HAS_XL = False

try:
    import cloudinary, cloudinary.uploader
    cloudinary.config(cloud_name=os.getenv('CLOUDINARY_CLOUD_NAME',''),
                      api_key=os.getenv('CLOUDINARY_API_KEY',''),
                      api_secret=os.getenv('CLOUDINARY_API_SECRET',''))
    HAS_CLD = True
except: HAS_CLD = False

# ── RISK ASSESSMENT LOOKUPS ──────────────────────────────────
RISK_LOOKUPS = {
    'countries': {'AFGHANISTAN':3,'ALAND ISLANDS':2,'ALBANIA':2,'ALGERIA':3,'AMERICAN SAMOA':2,'ANDORRA':2,'ANGOLA':3,'ANGUILLA':2,'ANTIGUA & BARBUDA':2,'ARGENTINA':2,'ARMENIA':2,'ARUBA':2,'AUSTRALIA':1,'AUSTRIA':2,'AZERBAIJAN':2,'BAHAMAS':2,'BAHRAIN':2,'BANGLADESH':2,'BARBADOS':2,'BELARUS':3,'BELGIUM':2,'BELIZE':2,'BENIN':2,'BERMUDA':2,'BHUTAN':1,'BOLIVIA':3,'BOSNIA AND HERZEGOVINA':2,'BOTSWANA':2,'BRAZIL':2,'BRITISH VIRGIN ISLANDS':2,'BRUNEI':1,'BULGARIA':2,'BURKINA FASO':3,'BURUNDI':3,'CAMBODIA':3,'CAMEROON':2,'CANADA':1,'CAPE VERDE':2,'CAYMAN ISLANDS':2,'CENTRAL AFRICAN REPUBLIC':3,'CHAD':3,'CHILE':2,'CHINA':2,'CHRISTMAS ISLAND':2,'COCOS (KEELING) ISLANDS':2,'COLOMBIA':2,'COMOROS':3,'CONGO':3,'COOK ISLANDS':2,'COSTA RICA':2,'COTE D IVOIRE':3,'CROATIA':2,'CUBA':3,'CURACAO':2,'CYPRUS':2,'CZECH REPUBLIC':2,'DEMOCRATIC REPUBLIC OF CONGO':3,'DENMARK':1,'DJIBOUTI':2,'DOMINICA':2,'DOMINICAN REPUBLIC':2,'EAST TIMOR':2,'ECUADOR':2,'EGYPT':3,'EL SALVADOR':3,'EQUATORIAL GUINEA':3,'ERITREA':3,'ESTONIA':2,'ETHIOPIA':3,'FALKLAND ISLANDS':2,'FAROE ISLANDS':2,'FIJI':2,'FINLAND':1,'FRANCE':1,'FRENCH GUIANA':2,'FRENCH POLYNESIA':2,'GABON':3,'GAMBIA':3,'GEORGIA':2,'GERMANY':1,'GHANA':3,'GIBRALTAR':2,'GREECE':2,'GREENLAND':2,'GRENADA':2,'GUADELOUPE':2,'GUAM':2,'GUATEMALA':3,'GUERNSEY':2,'GUINEA':3,'GUINEA BISSAU':3,'GUYANA':2,'HAITI':3,'HONDURAS':3,'HONG KONG':1,'HUNGARY':2,'ICELAND':1,'INDIA':2,'INDONESIA':2,'IRAN':3,'IRAQ':3,'IRELAND':1,'ISLE OF MAN':2,'ISRAEL':2,'ITALY':2,'JAMAICA':2,'JAPAN':1,'JERSEY':2,'JORDAN':2,'KAZAKHSTAN':3,'KENYA':3,'KIRIBATI':2,'KOREA NORTH':3,'KOREA SOUTH':1,'KOSOVO':2,'KUWAIT':2,'KYRGYZSTAN':3,'LAOS':3,'LATVIA':2,'LEBANON':3,'LESOTHO':2,'LIBERIA':3,'LIBYA':3,'LIECHTENSTEIN':2,'LITHUANIA':2,'LUXEMBOURG':1,'MACAO':2,'MACEDONIA':2,'MADAGASCAR':2,'MALAWI':2,'MALAYSIA':2,'MALDIVES':2,'MALI':3,'MALTA':2,'MARSHALL ISLANDS':2,'MARTINIQUE':2,'MAURITANIA':3,'MAURITIUS':2,'MAYOTTE':2,'MEXICO':3,'MICRONESIA':2,'MOLDOVA':3,'MONACO':2,'MONGOLIA':2,'MONTENEGRO':2,'MONTSERRAT':2,'MOROCCO':2,'MOZAMBIQUE':3,'MYANMAR':3,'NAMIBIA':2,'NAURU':2,'NEPAL':2,'NETHERLANDS':1,'NEW CALEDONIA':2,'NEW ZEALAND':1,'NICARAGUA':3,'NIGER':3,'NIGERIA':3,'NIUE':2,'NORFOLK ISLAND':2,'NORTHERN MARIANA ISLANDS':2,'NORWAY':1,'OMAN':2,'PAKISTAN':3,'PALAU':2,'PALESTINE':3,'PANAMA':3,'PAPUA NEW GUINEA':3,'PARAGUAY':2,'PERU':3,'PHILIPPINES':2,'PITCAIRN':2,'POLAND':2,'PORTUGAL':2,'PUERTO RICO':2,'QATAR':2,'REUNION':2,'ROMANIA':2,'RUSSIAN FEDERATION':3,'RWANDA':3,'SAINT BARTHELEMY':2,'SAINT HELENA':2,'SAINT KITTS & NEVIS':2,'SAINT LUCIA':2,'SAINT MARTIN':2,'SAINT PIERRE & MIQUELON':2,'SAINT VINCENT & THE GRENADINES':2,'SAMOA':2,'SAN MARINO':2,'SAO TOME & PRINCIPE':2,'SAUDI ARABIA':2,'SENEGAL':2,'SERBIA':2,'SEYCHELLES':2,'SIERRA LEONE':3,'SINGAPORE':1,'SINT MAARTEN':2,'SLOVAKIA':2,'SLOVENIA':2,'SOLOMON ISLANDS':2,'SOMALIA':3,'SOUTH AFRICA':2,'SOUTH SUDAN':3,'SPAIN':2,'SRI LANKA':2,'SUDAN':3,'SURINAME':2,'SWAZILAND':2,'SWEDEN':1,'SWITZERLAND':1,'SYRIA':3,'TAIWAN':2,'TAJIKISTAN':3,'TANZANIA':3,'THAILAND':2,'TIMOR-LESTE':2,'TOGO':3,'TOKELAU':2,'TONGA':2,'TRINIDAD AND TOBAGO':2,'TUNISIA':2,'TURKEY':2,'TURKMENISTAN':3,'TURKS & CAICOS ISLANDS':2,'TUVALU':2,'UGANDA':3,'UKRAINE':3,'UNITED ARAB EMIRATES':2,'UNITED KINGDOM':1,'UNITED STATES':1,'URUGUAY':2,'UZBEKISTAN':3,'VANUATU':2,'VATICAN':2,'VENEZUELA':3,'VIETNAM':2,'VIRGIN ISLANDS':2,'WALLIS & FUTUNA':2,'WESTERN SAHARA':3,'YEMEN':3,'ZAMBIA':3,'ZIMBABWE':3,'NOT APPLICABLE':1},
    'nature_of_business': {'Accounting Firm':2,'Advertising':2,'Airlines':2,'Arms Dealer':3,'Auction':2,'Auditors / Audit Firm':2,'Automobile Service':1,'Banking':3,'Brokerage Business':3,'Building Contracting':2,'Building Maintenance':1,'Building Management':1,'Car Business (New & Used Sales)':2,'Cargo & Logistics':3,'Carpentry':1,'Carpets Trading':1,'Catering':1,'Charities/Non-Profit Organizations':3,'Chemicals Trading':3,'Cleaning and Maintenance':1,'Commission Agent':2,'Communication Department':2,'Computer/Laptop Sales':1,'Computer/Laptop Service & Repairs':1,'Construction Items Shop':1,'Consultancy Services':3,'Corporate Services':3,'Cosmetics Trading':1,'Credit provider':2,'Dealers in Precious Metals & Stones':3,'Engineering':2,'Exchange Service':3,'Fast Moving Consumer Goods (FMCG)':1,'Financial Consultant':3,'Fish Trading':1,'Food Supplier':1,'Foreign Exchange':3,'Fruits & Vegetable Trading':1,'Fuel Station':1,'Furniture Trading':2,'General Trading Company':2,'Gift Shop':1,'Gold Smith':3,'Grocery':1,'Healthcare':1,'Heavy Equipment Leasing':2,'HMO/Travel Clinic':1,'Hotel':1,'HR / Recruitment Consultant':2,'Import/Export':2,'Insurance':2,'Interior Decoration':1,'Internet Service Provider':1,'Investment Company':3,'Jewellery Trading':3,'Laundry Services':1,'Law Firm':2,'Leather Items Trading':1,'Leasing Company':2,'Logistics':3,'Machinery Supplier':1,'Marble Trading':1,'Marketing':1,'Medicals/Pharmaceutical':1,'Metals & Minerals Trading':2,'Mining':1,'Mobile Phone Sales':1,'Mobile Phone Service & Repairs':1,'Money Lender':3,'Motor Vehicle Agent':2,'Motor Vehicle Spare Parts':1,'Nursery/School':1,'Nut Trading':1,'Oil Gas Supply':3,'Packaging Material Supplier':1,'Paper Trading':1,'Perfumes Trading':1,'Petrol Station':1,'Pharmaceutical':1,'Photography':1,'Plastic Item Trading':1,'Printing Press':1,'Professional Services':2,'Property Leasing':1,'Property Rental':1,'Publishing':1,'Real Estate':2,'Restaurant':1,'Retail':1,'Retail Gems & Jewellery':3,'Retail Supermarket':1,'Sales Agent/Distributor':2,'Sanitary Ware Trading':1,'Security Services':1,'Shell Company':3,'Shoe Shop':1,'Showroom':1,'Spare Parts Trading':1,'Sports':1,'Stationery Shop':1,'Steel Trading':2,'Stone Supplier':2,'Transportation':1,'Travel & Tourism':1,'Travel Agency':2,'Trading Company':2,'Tube Well Contractor':1,'Wholesale Jewellery':3,'Wholesale Trade':2,'Window Material Trading':1},
    'products': {'Hotels & Resorts':3,'Luxury Villas':3,'Penthouses':3,'Residential Apartments':2,'Retail Shops':2,'Townhouses':2,'Vacant Land':3,'Warehouses':1,'Office Spaces':2,'Permanent Desk':2,'Virtual Office':3,'PO Box':3,'Antique & Vintage Jewellery':2,'Bullion':3,'Collectible Coins':2,'Designer & Branded Jewellery':3,'Fine Jewellery':2,'Investment-Grade Precious Metals':3,'Luxury Watches':3,'Polished Diamonds':3,'Raw Precious Metals':2,'Uncut & Rough Gemstones':3,'Wholesale Jewellery':3},
    'amount_ranges_company': {'Amount up to AED 55,000':1,'Amount upto AED 500,000':2,'Amount more than AED 500,000':3},
    'amount_ranges_individual': {'Amount up to AED 100,000':1,'Amount more than AED 100,000':3},
    'payment_methods': {'Cash':3,'Cheque':1,'Crypto':3,'Bank Transfer (Local)':2,'Bank Transfer (International)':3,'Debit/Credit Card':2,'Gold to Gold Exchange':3,'All of the Above':3},
    'third_party': {'Yes':3,'No':1},
    'yes_no': {'Yes':3,'No':1},
}

def refresh_country_scores(conn=None):
    """Merge admin-edited country scores (DB) over the built-in defaults."""
    close = False
    try:
        if conn is None:
            conn = get_db(); close = True
        rows = all_(conn, "SELECT country, score FROM risk_country_scores")
        if rows:
            for r in rows:
                RISK_LOOKUPS['countries'][r['country']] = int(r['score'])
    except Exception as e:
        logger.warning(f'Could not load country scores: {e}')
    finally:
        if close and conn:
            try: conn.close()
            except: pass

app = Flask(__name__)
# SECRET_KEY MUST come from the environment. There is deliberately no hardcoded
# fallback: a known key lets anyone forge an admin session cookie. With multiple
# workers a per-process random key would also break logins, so we fail loudly
# and clearly instead of booting into a broken/insecure state.
app.secret_key = os.getenv('SECRET_KEY')
if not app.secret_key:
    raise RuntimeError(
        'SECRET_KEY environment variable is not set. Set it in Railway '
        '(Variables tab) to a long random value before starting the app.')

# ── SESSION & SECURITY CONFIG ────────────────────────────────
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)  # auto logout after 8h idle
app.config['SESSION_COOKIE_HTTPONLY'] = True   # JS cannot read session cookie
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'  # CSRF protection
# Only send the session cookie over HTTPS in production (Railway is HTTPS).
# Disabled in local development so http://localhost testing still works.
app.config['SESSION_COOKIE_SECURE'] = os.getenv('FLASK_ENV') != 'development'
# Reject oversized uploads before they exhaust memory (files are read into RAM).
app.config['MAX_CONTENT_LENGTH'] = 15 * 1024 * 1024  # 15 MB

# ── LOGGING ──────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
logger = logging.getLogger('zewer_crm')

# ── SECURITY HEADERS ─────────────────────────────────────────
@app.after_request
def _security_headers(resp):
    """Baseline hardening applied to every response."""
    resp.headers['X-Frame-Options'] = 'DENY'            # block clickjacking
    resp.headers['X-Content-Type-Options'] = 'nosniff'  # block MIME sniffing
    resp.headers['Referrer-Policy'] = 'strict-origin-when-cross-origin'
    return resp

# ── CONNECTION CLEANUP ───────────────────────────────────────
@app.teardown_appcontext
def _close_db_conns(exc):
    """Close every DB connection opened during this request. Routes still close
    their own connections on the happy path; this guarantees closure on the
    error path too (double-close is harmless), preventing pool exhaustion."""
    for c in (g.pop('db_conns', None) or []):
        try:
            c.close()
        except Exception:
            pass

# ── SAFE ERROR RESPONSE ──────────────────────────────────────
def _fail(e, code=500):
    """Log the real exception server-side, return a generic message to the client.
    Never leak SQL/schema/paths (str(e)) to the browser."""
    try:
        logger.error(f'{request.path}: {e}')
    except Exception:
        pass
    return jsonify({'success': False, 'error': 'Server error — please try again or contact support.'}), code

DATABASE_URL = os.getenv('DATABASE_URL', '')

# Local SQLite location (used only when DATABASE_URL is not set, i.e. local installs).
# Defaults to a 'data' folder beside this file so the DB ships and backs up with the app.
# Override with the SQLITE_PATH environment variable if you want it on another drive.
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SQLITE_PATH = os.getenv('SQLITE_PATH', os.path.join(BASE_DIR, 'data', 'aml_crm.db'))

@app.context_processor
def inject_user():
    return dict(user_name=session.get('user_name',''), user_role=session.get('user_role',''),
                user_email=session.get('user_email',''), current_user_id=session.get('user_id'),
                has_perm=has_perm)

# ── DATABASE ────────────────────────────────────────────────

def use_pg():
    return bool(DATABASE_URL)

def _open_db():
    if use_pg():
        try:
            import psycopg2, psycopg2.extras
            url = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
            c = psycopg2.connect(url, connect_timeout=10,
                                 cursor_factory=psycopg2.extras.RealDictCursor)
            c.autocommit = False
            return c
        except ImportError:
            print("psycopg2 not installed, using SQLite")
        except Exception as e:
            print(f"PG failed: {e}, using SQLite")
    os.makedirs(os.path.dirname(SQLITE_PATH) or '.', exist_ok=True)
    c = sqlite3.connect(SQLITE_PATH)
    c.row_factory = sqlite3.Row
    c.execute('PRAGMA foreign_keys = ON')
    return c

def get_db():
    """Open a DB connection and register it so the request teardown always
    closes it — even if a route raises before its own conn.close(). This is the
    safety net against connection-pool exhaustion under load."""
    c = _open_db()
    try:
        if 'db_conns' not in g:
            g.db_conns = []
        g.db_conns.append(c)
    except RuntimeError:
        pass  # called outside an app context (e.g. startup) — nothing to track
    return c

def is_pg(conn):
    return not isinstance(conn, sqlite3.Connection)

def P():
    return '%s' if use_pg() else '?'

def x(conn, sql, p=None):
    """Execute, auto-converting ? to %s for postgres"""
    if is_pg(conn):
        sql = sql.replace('?', '%s')
        cur = conn.cursor()
        cur.execute(sql, p or [])
        return cur
    return conn.execute(sql, p or [])

def one(conn, sql, p=None):
    r = x(conn, sql, p).fetchone()
    return dict(r) if r else None

def all_(conn, sql, p=None):
    rows = x(conn, sql, p).fetchall()
    return [dict(r) for r in rows]

def cnt(conn, sql, p=None):
    try:
        r = x(conn, sql, p).fetchone()
        if r is None: return 0
        if isinstance(r, dict): return list(r.values())[0]
        try: return r[0]
        except: return 0
    except:
        # On PostgreSQL a failed query aborts the transaction; roll back so the
        # connection isn't poisoned for every following query in the request.
        try: conn.rollback()
        except: pass
        return 0

def lastid(conn):
    if is_pg(conn):
        return dict(x(conn, 'SELECT lastval() as id').fetchone())['id']
    return conn.execute('SELECT last_insert_rowid()').fetchone()[0]

def commit(conn):
    conn.commit()

def _pg_ensure_columns():
    """Add new columns/tables on PostgreSQL using a FRESH autocommit connection.

    A brand-new connection cannot be in an aborted-transaction state, so every
    DDL statement here runs and commits independently — immune to the
    'current transaction is aborted' cascade that can otherwise skip ALTERs
    during the main migration. This is what guarantees the new schema exists.
    Runs at import time, independent of setup_db(), so a failure elsewhere in
    startup cannot prevent it.
    """
    if not use_pg():
        return
    ddl = [
        "ALTER TABLE companies ADD COLUMN IF NOT EXISTS disabled BOOLEAN DEFAULT FALSE",
        "ALTER TABLE clients ADD COLUMN IF NOT EXISTS disabled BOOLEAN DEFAULT FALSE",
        "ALTER TABLE clients ADD COLUMN IF NOT EXISTS pep TEXT",
        "ALTER TABLE ubos ADD COLUMN IF NOT EXISTS pep_status TEXT",
        "ALTER TABLE aml_tracker ADD COLUMN IF NOT EXISTS exchange_rate NUMERIC",
        "ALTER TABLE internal_documents ADD COLUMN IF NOT EXISTS file_url TEXT",
        "ALTER TABLE internal_documents ADD COLUMN IF NOT EXISTS file_name TEXT",
        "ALTER TABLE internal_documents ADD COLUMN IF NOT EXISTS public_id TEXT",
        "CREATE TABLE IF NOT EXISTS risk_country_scores (country TEXT PRIMARY KEY, score INTEGER NOT NULL DEFAULT 2, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "CREATE TABLE IF NOT EXISTS login_history (id SERIAL PRIMARY KEY, user_id INTEGER, username TEXT, success BOOLEAN DEFAULT FALSE, ip_address TEXT, user_agent TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "ALTER TABLE login_history ADD COLUMN IF NOT EXISTS logout_at TIMESTAMP",
        "CREATE TABLE IF NOT EXISTS risk_questions (id SERIAL PRIMARY KEY, applies_to TEXT DEFAULT 'both', question TEXT NOT NULL, sort_order INTEGER DEFAULT 0, is_active INTEGER DEFAULT 1, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "CREATE TABLE IF NOT EXISTS risk_answer_options (id SERIAL PRIMARY KEY, question_id INTEGER NOT NULL, label TEXT NOT NULL, score INTEGER NOT NULL DEFAULT 1, sort_order INTEGER DEFAULT 0)",
        "CREATE TABLE IF NOT EXISTS risk_responses (id SERIAL PRIMARY KEY, assessment_type TEXT, assessment_id INTEGER, question_id INTEGER, question_text TEXT, answer_label TEXT, score INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
    ]
    # (risk_questions seeding for PG happens in _run_migrations via _seed_risk_questions)
    try:
        import psycopg2
        url = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
        c = psycopg2.connect(url, connect_timeout=10)
        c.autocommit = True
        cur = c.cursor()
        for stmt in ddl:
            try:
                cur.execute(stmt)
            except Exception as e:
                logger.warning(f'ensure_columns ({stmt[:45]}...): {e}')
        cur.close(); c.close()
        logger.info('ensure_columns: new columns/tables verified on PostgreSQL')
    except Exception as e:
        logger.error(f'_pg_ensure_columns failed: {e}')

def setup_db():
    # Guarantee new columns exist FIRST, on a fresh autocommit connection,
    # so the rest of startup (and every request) can rely on them.
    _pg_ensure_columns()
    conn = get_db()
    if is_pg(conn):
        _pg_schema(conn)
        commit(conn)
    else:
        _sqlite_schema(conn)
    _run_migrations(conn)
    commit(conn)
    conn.close()
    # Belt-and-suspenders: ensure again after base tables are guaranteed to exist
    # (covers a brand-new database where the tables were just created above).
    _pg_ensure_columns()

def _run_migrations(conn):
    """Universal migrations that run for BOTH PostgreSQL and SQLite on every startup."""
    pg = is_pg(conn)

    # New tables (use x() which handles both DB types correctly)
    new_tables = [
        "company_groups (id {pk}, group_name TEXT UNIQUE NOT NULL, description TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "clients (id {pk}, name TEXT NOT NULL, phone TEXT, whatsapp_number TEXT, email TEXT, date_of_birth DATE, profession TEXT, address TEXT, notes TEXT, nationality TEXT, passport_no TEXT, passport_expiry DATE, emirates_id TEXT, emirates_id_expiry DATE, address_proof TEXT, created_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "app_settings (key TEXT PRIMARY KEY, value TEXT, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "regular_task_templates (id {pk}, title TEXT NOT NULL, description TEXT, frequency TEXT DEFAULT 'daily', assigned_role TEXT DEFAULT 'all', assigned_user_id INTEGER, created_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "regular_task_logs (id {pk}, template_id INTEGER NOT NULL, user_id INTEGER NOT NULL, notes TEXT, status TEXT DEFAULT 'done', logged_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "internal_documents (id {pk}, doc_name TEXT NOT NULL, doc_category TEXT DEFAULT 'Staff', person_name TEXT, issuing_authority TEXT, issue_date DATE, expiry_date DATE, notes TEXT, added_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "client_documents (id {pk}, client_id INTEGER NOT NULL, doc_type TEXT NOT NULL, file_name TEXT NOT NULL, file_url TEXT NOT NULL, public_id TEXT, uploaded_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "additional_tasks (id {pk}, title TEXT NOT NULL, task_details TEXT, remarks TEXT, from_datetime TIMESTAMP NOT NULL, to_datetime TIMESTAMP NOT NULL, created_by INTEGER NOT NULL, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "aml_tracker (id {pk}, company_id INTEGER, individual_id INTEGER, transaction_date DATE NOT NULL, period TEXT, due_date DATE, vc_no TEXT, payment_mode TEXT, ac_type TEXT, client_name TEXT, transaction_currency TEXT, usd_amount NUMERIC, aed_amount NUMERIC, payment_remarks TEXT, invoice_no TEXT, invoice_amount NUMERIC, invoice_currency TEXT, goaml_submission_date DATE, goaml_status TEXT DEFAULT 'pending', goaml_ref_no TEXT, submitted_by INTEGER NOT NULL, checked_by INTEGER, comment TEXT, verified_ledger BOOLEAN DEFAULT FALSE, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "company_risk_assessments (id {pk}, company_id INTEGER NOT NULL, jurisdiction_score NUMERIC, ownership_score NUMERIC, delivery_channel_score NUMERIC, payment_method_score NUMERIC, transaction_volume_score NUMERIC, product_score NUMERIC, pep_status_score NUMERIC, nationality_score NUMERIC, years_relationship_score NUMERIC, years_operation_score NUMERIC, third_party_score NUMERIC, sanctions_score NUMERIC, final_score NUMERIC, risk_rating TEXT, assessment_date DATE, notes TEXT, assessed_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "individual_risk_assessments (id {pk}, individual_id INTEGER NOT NULL, nationality_score NUMERIC, residence_status_score NUMERIC, pep_status_score NUMERIC, profession_score NUMERIC, product_score NUMERIC, delivery_channel_score NUMERIC, payment_method_score NUMERIC, transaction_amount_score NUMERIC, years_relationship_score NUMERIC, place_of_birth_score NUMERIC, third_party_score NUMERIC, sanctions_score NUMERIC, final_score NUMERIC, risk_rating TEXT, assessment_date DATE, notes TEXT, assessed_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "risk_country_scores (country TEXT PRIMARY KEY, score INTEGER NOT NULL DEFAULT 2, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "risk_questions (id {pk}, applies_to TEXT DEFAULT 'both', question TEXT NOT NULL, sort_order INTEGER DEFAULT 0, is_active INTEGER DEFAULT 1, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "risk_answer_options (id {pk}, question_id INTEGER NOT NULL, label TEXT NOT NULL, score INTEGER NOT NULL DEFAULT 1, sort_order INTEGER DEFAULT 0)",
        "risk_responses (id {pk}, assessment_type TEXT, assessment_id INTEGER, question_id INTEGER, question_text TEXT, answer_label TEXT, score INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
        "login_history (id {pk}, user_id INTEGER, username TEXT, success BOOLEAN DEFAULT FALSE, ip_address TEXT, user_agent TEXT, logout_at TIMESTAMP, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)",
    ]
    pk = "SERIAL PRIMARY KEY" if pg else "INTEGER PRIMARY KEY AUTOINCREMENT"
    for t in new_tables:
        sql = "CREATE TABLE IF NOT EXISTS " + t.format(pk=pk)
        try:
            x(conn, sql)
            commit(conn)
        except Exception as e:
            print(f"Migration table error ({sql[:50]}...): {e}")
            conn.rollback() if pg else None

    # ALTER TABLE additions — each wrapped individually so one failure doesn't block the rest
    def safe_alter(table, col, coltype='TEXT'):
        try:
            if pg:
                x(conn, f'ALTER TABLE {table} ADD COLUMN IF NOT EXISTS {col} {coltype}')
            else:
                x(conn, f'ALTER TABLE {table} ADD COLUMN {col} {coltype}')
            commit(conn)
        except Exception:
            if pg: conn.rollback()

    for col in ['contact_number','mobile','username','permissions']:
        safe_alter('users', col)
    # Ensure username column exists with explicit PG check
    if use_pg():
        try:
            x(conn, "ALTER TABLE users ADD COLUMN IF NOT EXISTS username TEXT")
            x(conn, "ALTER TABLE users ADD COLUMN IF NOT EXISTS permissions TEXT DEFAULT ''")
            commit(conn)
        except: conn.rollback()
    safe_alter('companies', 'group_name')
    safe_alter('companies', 'kyc_expiry_date', 'DATE')
    safe_alter('companies', 'id_type')
    safe_alter('companies', 'health_note')
    for col in ['contact_person_name','contact_person_number','account_manager',
                'whatsapp_number','deal_after_vat','registration_screening_tool']:
        safe_alter('companies', col)
    safe_alter('dropdowns', 'description')
    for col in ['nationality','passport_no','passport_expiry','emirates_id','emirates_id_expiry','address_proof','emirate','location','account_number','mode_of_ac','ac_status','id_type','pep_status','kyc_status','kyc_expiry_date','risk_status','screening_status','screening_date','is_resident']:
        coltype = 'DATE' if ('expiry' in col or col == 'screening_date') else ('BOOLEAN DEFAULT FALSE' if col == 'is_resident' else 'TEXT')
        safe_alter('clients', col, coltype)
    # Disable (out-of-scope) flag for companies & individuals
    safe_alter('companies', 'disabled', 'BOOLEAN DEFAULT FALSE')
    safe_alter('clients', 'disabled', 'BOOLEAN DEFAULT FALSE')
    # New PEP determination field (Yes/No) on individuals — separate from pep_status declaration
    safe_alter('clients', 'pep', 'TEXT')
    # PEP status on UBOs / authorized persons
    safe_alter('ubos', 'pep_status', 'TEXT')
    # Exchange rate (to AED) for AML tracker transactions
    safe_alter('aml_tracker', 'exchange_rate', 'NUMERIC')
    # File attachment for internal (Zewer) documents
    safe_alter('internal_documents', 'file_url')
    safe_alter('internal_documents', 'file_name')
    safe_alter('internal_documents', 'public_id')
    # Logout timestamp on login history
    safe_alter('login_history', 'logout_at', 'TIMESTAMP')
    # Force-insert new dropdown values on existing DBs
    new_dd = [
        ('ID TYPE','National ID'),('ID TYPE','Passport'),('ID TYPE','Emirates ID'),('ID TYPE','Visa No'),('ID TYPE','Other'),
        ('SCREENING REGISTRATION STATUS','Yes'),('SCREENING REGISTRATION STATUS','No'),
        ('SCREENING REGISTRATION STATUS','Not Required'),('SCREENING REGISTRATION STATUS','Pending'),
    ]
    for field, val in new_dd:
        try:
            if use_pg():
                x(conn, "INSERT INTO dropdowns (field_name,value,is_active) VALUES (%s,%s,1) ON CONFLICT DO NOTHING", (field, val))
                commit(conn)
            else:
                conn.execute("INSERT OR IGNORE INTO dropdowns (field_name,value,is_active) VALUES (?,?,1)", (field, val))
        except Exception:
            # Without this rollback, a failed insert here leaves the PG transaction
            # aborted, silently breaking every later migration step on this same
            # connection (including the risk-questionnaire seeding below).
            if pg: conn.rollback()

    # Seed editable country risk scores from the built-in defaults (first run only)
    for country, score in RISK_LOOKUPS.get('countries', {}).items():
        try:
            if use_pg():
                x(conn, "INSERT INTO risk_country_scores (country,score) VALUES (%s,%s) ON CONFLICT (country) DO NOTHING", (country, score))
                commit(conn)
            else:
                conn.execute("INSERT OR IGNORE INTO risk_country_scores (country,score) VALUES (?,?)", (country, score))
        except Exception:
            if pg: conn.rollback()
    # Load any admin-edited scores over the in-memory defaults
    refresh_country_scores(conn)

    # Seed the editable risk questionnaire (first run only)
    _seed_risk_questions(conn)

# Default questionnaire content, shared by the startup seeder and the admin
# "Load Default Questions" recovery action (/api/risk-questions/load-defaults).
_LMH = [('Low risk', 1), ('Medium risk', 2), ('High risk', 3)]
_YN = [('No', 1), ('Yes', 3)]
_PAYMENT = [('Cheque',1),('Bank Transfer (Local)',2),('Debit/Credit Card',2),('Cash',3),('Crypto',3),('Bank Transfer (International)',3)]
DEFAULT_RISK_QUESTIONS = {
    'company': [
        ('Jurisdiction / Country of Incorporation', _LMH), ('Ownership Structure', _LMH),
        ('Delivery Channel', _LMH), ('Payment Method', _PAYMENT),
        ('Transaction Volume', _LMH), ('Product / Service', _LMH),
        ('PEP Status', _YN), ('Nationality (of owners)', _LMH),
        ('Years of Relationship', [('More than 3 years',1),('1 to 3 years',2),('Less than 1 year',3)]),
        ('Years of Operation', [('More than 3 years',1),('1 to 3 years',2),('Less than 1 year',3)]),
        ('Third-Party Involvement', _YN), ('Sanctions Exposure', _YN),
    ],
    'individual': [
        ('Nationality', _LMH), ('Residence Status', [('Resident',1),('Non-Resident',3)]),
        ('PEP Status', _YN), ('Profession', _LMH),
        ('Product / Service', _LMH), ('Delivery Channel', _LMH),
        ('Payment Method', _PAYMENT),
        ('Transaction Amount', [('Amount up to AED 100,000',1),('Amount more than AED 100,000',3)]),
        ('Years of Relationship', [('More than 3 years',1),('1 to 3 years',2),('Less than 1 year',3)]),
        ('Place of Birth', _LMH), ('Third-Party Involvement', _YN), ('Sanctions Exposure', _YN),
    ],
}

def _insert_question_set(conn, applies_to, factors):
    """Insert one applies_to's default questions + answer options. Each question
    is independently try/excepted so one bad row can't block the rest."""
    inserted = 0
    for order, (q, answers) in enumerate(factors):
        try:
            if use_pg():
                x(conn, "INSERT INTO risk_questions (applies_to,question,sort_order,is_active) VALUES (%s,%s,%s,1)", (applies_to, q, order))
            else:
                conn.execute("INSERT INTO risk_questions (applies_to,question,sort_order,is_active) VALUES (?,?,?,1)", (applies_to, q, order))
            qid = lastid(conn)
            for aorder, (label, score) in enumerate(answers):
                if use_pg():
                    x(conn, "INSERT INTO risk_answer_options (question_id,label,score,sort_order) VALUES (%s,%s,%s,%s)", (qid, label, score, aorder))
                else:
                    conn.execute("INSERT INTO risk_answer_options (question_id,label,score,sort_order) VALUES (?,?,?,?)", (qid, label, score, aorder))
            commit(conn)
            inserted += 1
        except Exception as e:
            logger.warning(f'seed risk question skipped ({q}): {e}')
            if is_pg(conn): conn.rollback()
    return inserted

def _seed_risk_questions(conn):
    """Populate the configurable questionnaire with the default factors + graded
    answers, but only if it's empty (so admin edits are never clobbered)."""
    # Defensive: guarantee a clean transaction before checking/seeding, so an
    # earlier migration step that failed to roll back can never silently skip
    # this seeding (this is what caused production to end up with 0 questions).
    if is_pg(conn):
        try: conn.rollback()
        except Exception: pass
    try:
        existing = cnt(conn, 'SELECT COUNT(*) FROM risk_questions')
        if existing and int(existing) > 0:
            return
    except Exception:
        return
    _insert_question_set(conn, 'company', DEFAULT_RISK_QUESTIONS['company'])
    _insert_question_set(conn, 'individual', DEFAULT_RISK_QUESTIONS['individual'])

def _pg_schema(conn):
    stmts = [
        """CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY, email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL, name TEXT NOT NULL,
            role TEXT DEFAULT 'staff', contact_number TEXT, mobile TEXT,
            is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS companies (
            id SERIAL PRIMARY KEY, ac_code TEXT UNIQUE NOT NULL,
            client_name TEXT NOT NULL, ac_opening_date DATE,
            ac_status TEXT DEFAULT 'Active', active_till_year TEXT,
            nature TEXT, type_of_client TEXT, name_of_freezone TEXT,
            mode_of_ac TEXT, country_of_incorporation TEXT, region TEXT,
            address TEXT, telephone TEXT, mobile TEXT, whatsapp_number TEXT,
            email_id TEXT, contact_person_name TEXT, contact_person_number TEXT,
            account_manager TEXT, address_proof_type TEXT, address_proof_expiry DATE,
            kyc_status TEXT, trade_license_no TEXT, issuing_authority TEXT,
            legal_type TEXT, incorporation_date DATE, trade_license_expiry DATE,
            tax_no_trn TEXT, vat_cert TEXT, vat_declaration TEXT, deal_after_vat TEXT,
            num_beneficial_owners INTEGER DEFAULT 0, moa TEXT, pep TEXT,
            undertaking TEXT, source_of_fund TEXT, software_updation TEXT,
            doc_status TEXT DEFAULT 'Incompleted', screening_date DATE,
            registration_screening_tool TEXT, risk_status TEXT DEFAULT 'Unspecified',
            verified_by TEXT, verified_date DATE, followup_details TEXT,
            crowe_feedback TEXT, zewer_comments TEXT, group_name TEXT, kyc_expiry_date DATE, id_type TEXT, created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS ubos (
            id SERIAL PRIMARY KEY, company_id INTEGER NOT NULL,
            position TEXT, share_percentage REAL, person_name TEXT NOT NULL,
            nationality TEXT, residential_status TEXT, passport_no TEXT,
            passport_expiry DATE, emirates_id TEXT, emirates_id_expiry DATE,
            doc_status TEXT DEFAULT 'Incompleted', verified_by TEXT,
            verified_date DATE, followup_details TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS dropdowns (
            id SERIAL PRIMARY KEY, field_name TEXT NOT NULL, value TEXT NOT NULL,
            description TEXT, is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(field_name, value))""",
        """CREATE TABLE IF NOT EXISTS tasks (
            id SERIAL PRIMARY KEY, title TEXT NOT NULL, description TEXT,
            assigned_to INTEGER, created_by INTEGER, company_id INTEGER,
            priority TEXT DEFAULT 'normal', status TEXT DEFAULT 'todo',
            due_date DATE, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS documents (
            id SERIAL PRIMARY KEY, company_id INTEGER NOT NULL,
            doc_type TEXT NOT NULL, file_name TEXT NOT NULL,
            file_url TEXT NOT NULL, public_id TEXT, uploaded_by INTEGER,
            notes TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS regular_task_templates (
            id SERIAL PRIMARY KEY, title TEXT NOT NULL, description TEXT,
            frequency TEXT DEFAULT 'daily', assigned_role TEXT DEFAULT 'all',
            assigned_user_id INTEGER, created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS regular_task_logs (
            id SERIAL PRIMARY KEY, template_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL, notes TEXT, status TEXT DEFAULT 'done',
            logged_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS internal_documents (
            id SERIAL PRIMARY KEY, doc_name TEXT NOT NULL,
            doc_category TEXT DEFAULT 'Staff', person_name TEXT,
            issuing_authority TEXT, issue_date DATE, expiry_date DATE,
            notes TEXT, added_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS company_groups (
            id SERIAL PRIMARY KEY, group_name TEXT UNIQUE NOT NULL,
            description TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS clients (
            id SERIAL PRIMARY KEY, name TEXT NOT NULL, phone TEXT,
            whatsapp_number TEXT, email TEXT, date_of_birth DATE,
            profession TEXT, address TEXT, notes TEXT,
            passport_file TEXT, emirates_id_file TEXT, address_proof_file TEXT,
            created_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY, value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
        """CREATE TABLE IF NOT EXISTS additional_tasks (
            id SERIAL PRIMARY KEY, title TEXT NOT NULL,
            task_details TEXT, remarks TEXT,
            from_datetime TIMESTAMP NOT NULL, to_datetime TIMESTAMP NOT NULL,
            created_by INTEGER NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""",
    ]
    for s in stmts:
        x(conn, s)
    # Seed admin
    if not one(conn, 'SELECT id FROM users WHERE email=%s', ('admin@zewer.ae',)):
        x(conn, 'INSERT INTO users (email,password_hash,name,role) VALUES (%s,%s,%s,%s)',
          ('admin@zewer.ae', generate_password_hash('Admin@123'), 'Administrator', 'admin'))
        x(conn, 'INSERT INTO users (email,password_hash,name,role) VALUES (%s,%s,%s,%s)',
          ('compliance@zewer.ae', generate_password_hash('Compliance@123'), 'Compliance Officer', 'compliance'))
    if cnt(conn, 'SELECT COUNT(*) FROM dropdowns') == 0:
        _seed(conn)

def _sqlite_schema(conn):
    conn.executescript('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT, email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL, name TEXT NOT NULL,
            role TEXT DEFAULT 'staff', contact_number TEXT, mobile TEXT,
            is_active INTEGER DEFAULT 1, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS companies (
            id INTEGER PRIMARY KEY AUTOINCREMENT, ac_code TEXT UNIQUE NOT NULL,
            client_name TEXT NOT NULL, ac_opening_date DATE,
            ac_status TEXT DEFAULT 'Active', active_till_year TEXT, nature TEXT,
            type_of_client TEXT, name_of_freezone TEXT, mode_of_ac TEXT,
            country_of_incorporation TEXT, region TEXT, address TEXT,
            telephone TEXT, mobile TEXT, whatsapp_number TEXT, email_id TEXT,
            contact_person_name TEXT, contact_person_number TEXT, account_manager TEXT,
            address_proof_type TEXT, address_proof_expiry DATE, kyc_status TEXT,
            trade_license_no TEXT, issuing_authority TEXT, legal_type TEXT,
            incorporation_date DATE, trade_license_expiry DATE, tax_no_trn TEXT,
            vat_cert TEXT, vat_declaration TEXT, deal_after_vat TEXT,
            num_beneficial_owners INTEGER DEFAULT 0, moa TEXT, pep TEXT,
            undertaking TEXT, source_of_fund TEXT, software_updation TEXT,
            doc_status TEXT DEFAULT 'Incompleted', screening_date DATE,
            registration_screening_tool TEXT, risk_status TEXT DEFAULT 'Unspecified',
            verified_by TEXT, verified_date DATE, followup_details TEXT,
            crowe_feedback TEXT, zewer_comments TEXT, group_name TEXT, kyc_expiry_date DATE, id_type TEXT, created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS ubos (
            id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL,
            position TEXT, share_percentage REAL, person_name TEXT NOT NULL,
            nationality TEXT, residential_status TEXT, passport_no TEXT,
            passport_expiry DATE, emirates_id TEXT, emirates_id_expiry DATE,
            doc_status TEXT DEFAULT 'Incompleted', verified_by TEXT,
            verified_date DATE, followup_details TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS dropdowns (
            id INTEGER PRIMARY KEY AUTOINCREMENT, field_name TEXT NOT NULL,
            value TEXT NOT NULL, description TEXT, is_active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, UNIQUE(field_name, value));
        CREATE TABLE IF NOT EXISTS tasks (
            id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL,
            description TEXT, assigned_to INTEGER, created_by INTEGER,
            company_id INTEGER, priority TEXT DEFAULT 'normal',
            status TEXT DEFAULT 'todo', due_date DATE,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER NOT NULL,
            doc_type TEXT NOT NULL, file_name TEXT NOT NULL, file_url TEXT NOT NULL,
            public_id TEXT, uploaded_by INTEGER, notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS regular_task_templates (
            id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL, description TEXT,
            frequency TEXT DEFAULT 'daily', assigned_role TEXT DEFAULT 'all',
            assigned_user_id INTEGER, created_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS regular_task_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT, template_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL, notes TEXT, status TEXT DEFAULT 'done',
            logged_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS internal_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT, doc_name TEXT NOT NULL,
            doc_category TEXT DEFAULT 'Staff', person_name TEXT,
            issuing_authority TEXT, issue_date DATE, expiry_date DATE,
            notes TEXT, added_by INTEGER,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS company_groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT, group_name TEXT UNIQUE NOT NULL,
            description TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL, phone TEXT,
            whatsapp_number TEXT, email TEXT, date_of_birth DATE,
            profession TEXT, address TEXT, notes TEXT,
            passport_file TEXT, emirates_id_file TEXT, address_proof_file TEXT,
            created_by INTEGER, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
        CREATE TABLE IF NOT EXISTS app_settings (
            key TEXT PRIMARY KEY, value TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP);
    ''')

    if not conn.execute('SELECT id FROM users WHERE email=?',('admin@zewer.ae',)).fetchone():
        conn.execute('INSERT INTO users (email,password_hash,name,role) VALUES (?,?,?,?)',
            ('admin@zewer.ae', generate_password_hash('Admin@123'), 'Administrator', 'admin'))
        conn.execute('INSERT INTO users (email,password_hash,name,role) VALUES (?,?,?,?)',
            ('compliance@zewer.ae', generate_password_hash('Compliance@123'), 'Compliance Officer', 'compliance'))
    if not conn.execute('SELECT id FROM dropdowns LIMIT 1').fetchone():
        _seed(conn)
    conn.commit()

def _seed(conn):
    data = {
        'AC STATUS':['Active','Inactive'],
        'NATURE':['Individual','Legal entity'],
        'TYPE OF CLIENT':['MainLand','Free Zone','Abroad','International Corporate'],
        'MODE OF AC':['Supplier','Customer','Bullion','Refinery','Logistics Co','Exchange','Bank','Insurance','Investor','Technical Services'],
        'RISK STATUS':['High','Medium','Low','Unspecified'],
        'DOC STATUS':['Completed','Incompleted'],
        'KYC STATUS':['New Kyc Updated','Kyc 2025 Updated','Kyc 2024 Updated','Kyc 2023 Updated','Kyc 2022 Updated','Kyc 2021 Updated','Kyc 2020 Updated','Kyc 2019 Updated','Kyc 2018 Updated','Kyc 2017 Updated','Kyc 2016 Updated','Not Updated'],
        'REGION':['Dubai','Abu Dhabi','Sharjah','Ajman','Ras Al Khaimah','Fujairah','Umm Al Quwain','Al Ain','West Bengal','United Kingdom','Canada','Pakistan','Malaysia','Singapore','Bahrain','Italy','REPUBLIC OF CONGO','HONG KONG','Saudi Arabia','India'],
        'FREEZONE':['Jabel Ali FZ','DMCC','Sharjah Saif Zone','Ajman FZ','Fujairah','Dubai Production City','N/A','Ras Al Khaimah Fz','Dubai Free Zone','Dubai Gold & Diamond Park'],
        'LEGAL TYPE':['Limited Liability Company(LLC)','Limited Liability Company (WLL)','Civil Company Professional','Foreign Company','Public Company','Privet Company','DMCC','Free Zone Limited Liability Company (FZ-LLC)','Sole Establishment','Partnership Company','Services Agency','Limited (LTD)','Individual Institution','Establishment','FZCO','FZE','FZC'],
        'ISSUING AUTHORITY':['Dubai Economy & Tourism','Department of Economic Development Ajman','Abu-Dhabi Department of Economic Development','DMCC','Saif Zone','Ajman FZ','Jebel Ali FZ','Dubai Development Authority','Government Of Sharjah Economic Development Department','Government of Ras Al-Khaimah Department of Economic Development','Department of Economic Development Dubai','Dubai Integrated Economic Zones Authority','Fujairah Municipality','Trade Development Authority Of Pakistan','UK HMRC','The Registrar Of Companies For England And Wales','Canada Revenue Agency','Kolkata Municipal Corporation','Abroad'],
        'ADDRESS PROOF TYPE':['Ejari','Tenancy','Electricity Bill','Gst Registration Certificate','Certification Of Incorporation','Certificate Of Registration For Value Added Tax','Vat Certificate','Telephone Bill','Title Deed','Certificate Of Enlistment','Association Of Article Details','Warehouse Lease Agreement','Not Required'],
        'VAT CERT':['Yes','No','Not Required'],
        'VAT DECLARATION':['Yes','No','Not Required'],
        'MOA':['Yes','No'], 'PEP':['Yes','No'],
        'UNDERTAKING':['Yes','No'], 'SOURCE OF FUND':['Yes','No'],
        'POSITION':['UBO','Authorized Person','Director','Manager','Partner'],
        'RESIDENTIAL STATUS':['Resident','Non Resident'],
        'COUNTRY':['United Arab Emirates','Saudi Arabia','Kuwait','Qatar','Bahrain','Oman','India','Pakistan','Bangladesh','Sri Lanka','Philippines','Malaysia','Singapore','China','Hong Kong','Jordan','Lebanon','Syria','Iraq','Yemen','Egypt','Libya','Nigeria','Ethiopia','Republic Of Congo','Turkey','Iran','Afghanistan','Algeria','Canada','United Kingdom','United States of America','France','Ireland','Italy','Germany','Armenia','Belize'],
        'TASK TEMPLATE':['Collect Updated Trade License','KYC Update Required','Address Proof Renewal','Passport Renewal Follow-up','Emirates ID Update','VAT Certificate Collection','Screening Review','MOA Collection','Undertaking Form','Source of Funds Verification','Risk Assessment Review','Annual KYC Review'],
        'ID TYPE':['National ID','Passport','Emirates ID','Visa No'],
        'SCREENING REGISTRATION STATUS':['Yes','No','Not Required','Pending'],
    }
    pg = is_pg(conn)
    for field, vals in data.items():
        for v in vals:
            try:
                if pg:
                    x(conn, 'INSERT INTO dropdowns (field_name,value,is_active) VALUES (%s,%s,1) ON CONFLICT DO NOTHING', (field,v))
                else:
                    conn.execute('INSERT OR IGNORE INTO dropdowns (field_name,value,is_active) VALUES (?,?,1)', (field,v))
            except: pass

try:
    setup_db()
    print("DB ready")
except Exception as e:
    print(f"DB setup warning: {e}")
# Final guarantee: even if setup_db() raised, make sure the new columns exist.
try:
    _pg_ensure_columns()
except Exception as e:
    print(f"ensure_columns warning: {e}")

# ── HELPERS ─────────────────────────────────────────────────

def days_left(d):
    if not d: return None
    try:
        if hasattr(d, 'year'):  # already a date object
            return (d - dubai_today()).days
        return (datetime.strptime(str(d)[:10],'%Y-%m-%d').date()-dubai_today()).days
    except: return None

def exp_status(d):
    if d is None: return 'unknown'
    if d<0: return 'expired'
    if d<=30: return 'critical'
    if d<=90: return 'warning'
    return 'ok'

def dropdowns():
    conn=get_db()
    rows=all_(conn,"SELECT field_name,value FROM dropdowns WHERE is_active=1 ORDER BY field_name,value")
    conn.close()
    dd={}
    for r in rows: dd.setdefault(r['field_name'],[]).append(r['value'])
    return dd

def login_required(f):
    @wraps(f)
    def d(*a,**k):
        if 'user_id' not in session: return redirect(url_for('login'))
        return f(*a,**k)
    return d

def admin_required(f):
    @wraps(f)
    def d(*a,**k):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('user_role')!='admin':
            flash("That page is for administrators only.")
            return redirect(url_for('dashboard'))
        return f(*a,**k)
    return d

def compliance_required(f):
    @wraps(f)
    def d(*a,**k):
        if 'user_id' not in session: return redirect(url_for('login'))
        if session.get('user_role') not in ('admin','compliance'):
            flash("You don't have access to that section.")
            return redirect(url_for('tasks'))
        return f(*a,**k)
    return d

# ── FINE-GRAINED ROLE PERMISSIONS ────────────────────────────
# Admin-editable per-role permissions (Settings → Role Permissions), stored as JSON
# under app_settings 'role_permissions'. Admin ALWAYS has every permission and can
# never be locked out. Server defaults below mirror the UI defaults, so behaviour is
# unchanged until an admin edits the checkboxes.
ROLE_PERM_DEFAULTS = {
    'compliance': {'dashboard','companies_view','companies_add','companies_edit','companies_docs',
                   'companies_export','companies_import','alerts','clients','reports',
                   'aml_export','aml_import',
                   'tasks_view','tasks_create','tasks_edit','tasks_delete',
                   'regular_tasks_view','regular_tasks_log','regular_tasks_manage',
                   'zewer_docs_view','zewer_docs_edit'},
    'staff': {'dashboard','tasks_view','tasks_create','tasks_edit','regular_tasks_view','regular_tasks_log'},
}

def _role_perms_config():
    """Admin-saved role→perms map (cached per request); None if never configured."""
    try:
        if hasattr(g, '_rp_cache'):
            return g._rp_cache
    except RuntimeError:
        pass
    val = None
    try:
        conn = get_db()
        row = one(conn, "SELECT value FROM app_settings WHERE key='role_permissions'")
        conn.close()
        if row and row.get('value'):
            import json as _json
            val = _json.loads(row['value'])
    except Exception:
        val = None
    try:
        g._rp_cache = val
    except RuntimeError:
        pass
    return val

def has_perm(perm):
    """True if the current user's role has `perm`. Admin is always allowed."""
    role = session.get('user_role')
    if role == 'admin':
        return True
    if not role:
        return False
    cfg = _role_perms_config()
    if cfg and role in cfg:
        return perm in (cfg.get(role) or [])
    return perm in ROLE_PERM_DEFAULTS.get(role, set())

def require_perm(perm):
    """Route guard: 403 (API) or flash+redirect (page) if the user lacks `perm`."""
    def deco(f):
        @wraps(f)
        def d(*a, **k):
            if 'user_id' not in session:
                if request.path.startswith('/api/') or request.is_json:
                    return jsonify({'success': False, 'error': 'Not signed in'}), 401
                return redirect(url_for('login'))
            if not has_perm(perm):
                if request.path.startswith('/api/') or request.is_json:
                    return jsonify({'success': False, 'error': 'You do not have permission for this action.'}), 403
                flash("You don't have permission to do that.")
                return redirect(url_for('dashboard'))
            return f(*a, **k)
        return d
    return deco

def _action_pw_matches(pw, stored):
    """Verify pw against a stored action-password hash. Accepts both the new
    salted werkzeug hashes and the legacy unsalted sha256 hex, so an already-set
    password keeps working after the upgrade."""
    if not stored:
        return True  # no action password configured
    pw = pw or ''
    if stored.startswith(('pbkdf2:', 'scrypt:')):
        try:
            return check_password_hash(stored, pw)
        except Exception:
            return False
    import hashlib  # legacy unsalted sha256
    return hashlib.sha256(pw.encode()).hexdigest() == stored

def _check_action_pw(pw):
    """True if the supplied action password matches the stored one (or none is set)."""
    try:
        conn = get_db()
        s = one(conn, "SELECT value FROM app_settings WHERE key='action_password'")
        conn.close()
        return _action_pw_matches(pw, s.get('value') if s else None)
    except Exception:
        return False

def _require_admin_pw():
    """Guard for edit/delete: caller must be admin AND supply a valid action password.
    Returns an error (response, status) tuple to return, or None if allowed."""
    if session.get('user_role') != 'admin':
        return jsonify({'success': False, 'error': 'Only an admin can edit or delete records.'}), 403
    if request.is_json:
        pw = (request.get_json(silent=True) or {}).get('action_password')
    else:
        pw = request.form.get('action_password')
    if not _check_action_pw(pw):
        return jsonify({'success': False, 'error': 'Action password is required or incorrect.'}), 403
    return None

def admin_pw_required(f):
    """Decorator form of _require_admin_pw for JSON edit/delete endpoints."""
    @wraps(f)
    def d(*a, **k):
        if 'user_id' not in session:
            return jsonify({'success': False, 'error': 'Not signed in'}), 401
        err = _require_admin_pw()
        if err: return err
        return f(*a, **k)
    return d

# ── ROUTES ──────────────────────────────────────────────────

@app.route('/')
def index(): return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))

@app.route('/healthz')
def healthz():
    """No-auth health + schema check, used to verify a deploy from outside.
    Returns 200 only if every new column/table exists; 503 if any is missing."""
    checks = {}
    try:
        conn = get_db()
        for tbl, col in [('companies', 'disabled'), ('clients', 'disabled'), ('clients', 'pep'),
                         ('ubos', 'pep_status'), ('aml_tracker', 'exchange_rate'),
                         ('internal_documents', 'file_url')]:
            try:
                x(conn, f'SELECT {col} FROM {tbl} LIMIT 1').fetchone()
                checks[f'{tbl}.{col}'] = 'ok'
            except Exception:
                try: conn.rollback()
                except: pass
                checks[f'{tbl}.{col}'] = 'MISSING'
        checks['login_history.logout_at'] = 'ok'
        try:
            x(conn, 'SELECT logout_at FROM login_history LIMIT 1').fetchone()
        except Exception:
            try: conn.rollback()
            except: pass
            checks['login_history.logout_at'] = 'MISSING'
        for t in ['risk_country_scores', 'login_history', 'risk_questions', 'risk_answer_options', 'risk_responses']:
            try:
                x(conn, f'SELECT 1 FROM {t} LIMIT 1').fetchone()
                checks[t] = 'ok'
            except Exception:
                try: conn.rollback()
                except: pass
                checks[t] = 'MISSING'
        conn.close()
        ok = all(v == 'ok' for v in checks.values())
        return jsonify({'status': 'ok' if ok else 'degraded', 'schema': checks}), (200 if ok else 503)
    except Exception as e:
        logger.error(f'/healthz: {e}')
        return jsonify({'status': 'error'}), 500

def _client_ip():
    """Best-effort client IP, honouring the first X-Forwarded-For hop (Railway proxy)."""
    ip = request.headers.get('X-Forwarded-For', request.remote_addr or '')
    if ip and ',' in ip:
        ip = ip.split(',')[0].strip()
    return ip

# Login throttle: block after this many failures from one IP within the window.
LOGIN_MAX_FAILS = 10
LOGIN_WINDOW_MIN = 15

def _too_many_attempts(conn, ip):
    """True if this IP has exceeded the failed-login limit inside the time window."""
    if not ip:
        return False
    try:
        cutoff = (datetime.utcnow() - timedelta(minutes=LOGIN_WINDOW_MIN)).strftime('%Y-%m-%d %H:%M:%S')
        n = cnt(conn, "SELECT COUNT(*) FROM login_history WHERE ip_address=? AND success IS NOT TRUE AND created_at >= ?",
                (ip, cutoff))
        return int(n or 0) >= LOGIN_MAX_FAILS
    except Exception:
        return False  # never lock people out because the check itself failed

def _log_login(conn, user_id, username, success):
    """Record a login attempt for the suspicious-login audit trail. Returns the row id (or None)."""
    try:
        ip = _client_ip()
        ua = (request.headers.get('User-Agent', '') or '')[:300]
        x(conn, '''INSERT INTO login_history (user_id, username, success, ip_address, user_agent)
                   VALUES (?,?,?,?,?)''', (user_id, username, success, ip, ua))
        commit(conn)
        try:
            return lastid(conn)
        except Exception:
            return None
    except Exception as e:
        logger.warning(f'Could not record login attempt: {e}')
        return None

@app.route('/login', methods=['GET','POST'])
def login():
    if request.method=='POST':
        d=request.get_json(silent=True) or {}
        conn=get_db()
        # Brute-force throttle: too many recent failures from this IP → refuse early.
        if _too_many_attempts(conn, _client_ip()):
            conn.close()
            logger.warning(f'Login throttled for IP {_client_ip()}')
            return jsonify({'success':False,'error':'Too many failed attempts. Please wait 15 minutes and try again.'}),429
        login_id = d.get('username') or d.get('email','')
        u=one(conn,'SELECT * FROM users WHERE username=? OR email=?',(login_id,login_id))
        if u and check_password_hash(u['password_hash'],d.get('password','')) and u['is_active']:
            session.permanent = True  # enables PERMANENT_SESSION_LIFETIME timeout
            session.update(user_id=u['id'],user_email=u['email'],user_name=u['name'],user_role=u['role'],user_permissions=(u.get('permissions') or ''))
            logger.info(f"Login: {u['email']} (role={u['role']})")
            log_id = _log_login(conn, u['id'], u.get('username') or u['email'], True)
            if log_id:
                session['login_log_id'] = log_id
            conn.close()
            return jsonify({'success':True})
        logger.warning(f"Failed login attempt for: {login_id}")
        _log_login(conn, u['id'] if u else None, login_id, False)
        conn.close()
        return jsonify({'success':False,'error':'Invalid credentials'}),401
    return render_template('login.html')

@app.route('/logout')
def logout():
    # Stamp the logout time on this session's login-history row
    log_id = session.get('login_log_id')
    if log_id:
        try:
            conn = get_db()
            x(conn, 'UPDATE login_history SET logout_at=CURRENT_TIMESTAMP WHERE id=?', (log_id,))
            commit(conn); conn.close()
        except Exception as e:
            logger.warning(f'Could not record logout time: {e}')
    session.clear()
    return redirect(url_for('login'))

@app.route('/dashboard')
@login_required
def dashboard():
    # Staff get their own task dashboard
    if session.get('user_role') == 'staff':
        uid = session.get('user_id')
        conn = get_db()
        today = dubai_today()
        my_tasks = all_(conn, '''SELECT t.*,c.client_name as company_name,c.ac_code
            FROM tasks t LEFT JOIN companies c ON t.company_id=c.id
            WHERE t.assigned_to=? AND t.status NOT IN ('done')
            ORDER BY CASE t.priority WHEN 'urgent' THEN 1 WHEN 'high' THEN 2 WHEN 'normal' THEN 3 ELSE 4 END,t.due_date''', (uid,))
        tl = []
        for t in my_tasks:
            d = days_left(t['due_date'])
            tl.append({**t,'priority':t['priority'] or 'normal','status':t['status'] or 'todo',
                       'due_date':str(t['due_date']) if t['due_date'] else None,
                       'days_until_due':d,'is_overdue':(d is not None and d<0)})
        todo_c = sum(1 for t in tl if t['status']=='todo')
        inprog_c = sum(1 for t in tl if t['status']=='inprogress')
        pending_c = sum(1 for t in tl if t['status']=='pending_close')
        overdue_c = sum(1 for t in tl if t['is_overdue'])
        conn.close()
        return render_template('staff_dashboard.html', my_tasks=tl,
            todo_count=todo_c, inprogress_count=inprog_c,
            pending_count=pending_c, overdue_count=overdue_c, today=str(today))

    conn=get_db(); today=dubai_today()
    def c(sql,p=None): return cnt(conn,sql,p or [])
    # Disabled (out-of-scope) records are excluded from all dashboard figures.
    ND = ' AND disabled IS NOT TRUE'           # for companies / clients
    NDU = ' AND company_id NOT IN (SELECT id FROM companies WHERE disabled IS TRUE)'  # for ubos
    total=c('SELECT COUNT(*) FROM companies WHERE 1=1'+ND)
    active=c('SELECT COUNT(*) FROM companies WHERE ac_status=?'+ND,('Active',))
    total_clients=c('SELECT COUNT(*) FROM clients WHERE 1=1'+ND)
    pep_count=(c("SELECT COUNT(*) FROM companies WHERE pep='Yes'"+ND) +
               c("SELECT COUNT(*) FROM clients WHERE pep_status='Yes'"+ND))
    etl=c('SELECT COUNT(*) FROM companies WHERE trade_license_expiry<?'+ND,(today,))
    e30tl=c('SELECT COUNT(*) FROM companies WHERE trade_license_expiry BETWEEN ? AND ?'+ND,(today,today+timedelta(days=30)))
    eap=c('SELECT COUNT(*) FROM companies WHERE address_proof_expiry<?'+ND,(today,))
    e30ap=c('SELECT COUNT(*) FROM companies WHERE address_proof_expiry BETWEEN ? AND ?'+ND,(today,today+timedelta(days=30)))
    epass=c('SELECT COUNT(*) FROM ubos WHERE passport_expiry<?'+NDU,(today,))
    e30p=c('SELECT COUNT(*) FROM ubos WHERE passport_expiry BETWEEN ? AND ?'+NDU,(today,today+timedelta(days=30)))
    eid_exp=c('SELECT COUNT(*) FROM ubos WHERE emirates_id_expiry<?'+NDU,(today,))
    risk_rows=all_(conn,'SELECT risk_status,COUNT(*) as c FROM companies WHERE 1=1'+ND+' GROUP BY risk_status')
    risk_bd={r['risk_status']:r['c'] for r in risk_rows}
    doc_rows=all_(conn,'SELECT doc_status,COUNT(*) as c FROM companies WHERE 1=1'+ND+' GROUP BY doc_status')
    doc_bd={r['doc_status']:r['c'] for r in doc_rows}
    # Count ALL expiring documents (TL + AP + Passport + EID) by window
    def doc_count(days):
        t1 = cnt(conn,'SELECT COUNT(*) FROM companies WHERE trade_license_expiry BETWEEN ? AND ?'+ND,(today,today+timedelta(days=days)))
        t2 = cnt(conn,'SELECT COUNT(*) FROM companies WHERE address_proof_expiry BETWEEN ? AND ?'+ND,(today,today+timedelta(days=days)))
        t3 = cnt(conn,'SELECT COUNT(*) FROM ubos WHERE passport_expiry BETWEEN ? AND ?'+NDU,(today,today+timedelta(days=days)))
        t4 = cnt(conn,'SELECT COUNT(*) FROM ubos WHERE emirates_id_expiry BETWEEN ? AND ?'+NDU,(today,today+timedelta(days=days)))
        return t1 + t2 + t3 + t4
    exp_30 = doc_count(30)
    exp_60 = doc_count(60)
    exp_90 = doc_count(90)
    urgent = []  # no longer used in template
    otasks=c("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('done')")
    overtasks=c("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('done','pending_close') AND due_date<?", (today,))
    dtasks=c("SELECT COUNT(*) FROM tasks WHERE status NOT IN ('done') AND due_date=?", (today,))
    ptasks=c("SELECT COUNT(*) FROM tasks WHERE status='pending_close'")
    utasks_raw=all_(conn,"""SELECT t.id,t.title,t.priority,t.status,t.due_date,u.name as assigned_name
        FROM tasks t LEFT JOIN users u ON t.assigned_to=u.id
        WHERE t.status NOT IN ('done') AND (t.priority IN ('urgent','high') OR t.due_date<=?)
        ORDER BY CASE t.priority WHEN 'urgent' THEN 1 WHEN 'high' THEN 2 ELSE 3 END,t.due_date LIMIT 8""",
        (today+timedelta(days=2),))
    # Convert date objects to strings
    utasks = []
    for t in utasks_raw:
        row = dict(t)
        row['due_date'] = str(row['due_date'])[:10] if row.get('due_date') else None
        utasks.append(row)
    try:
        staff_task_counts=all_(conn,"SELECT u.name,u.id,COUNT(t.id) as pending FROM users u LEFT JOIN tasks t ON t.assigned_to=u.id AND t.status NOT IN ('done') WHERE u.is_active=1 GROUP BY u.id,u.name ORDER BY pending DESC")
    except: staff_task_counts=[]

    # Upcoming client birthdays (next 30 days)
    upcoming_birthdays = []
    try:
        all_clients = all_(conn, "SELECT id,name,whatsapp_number,phone,date_of_birth FROM clients WHERE date_of_birth IS NOT NULL AND disabled IS NOT TRUE")
        for cl in all_clients:
            dob = cl.get('date_of_birth')
            if not dob: continue
            try:
                d = datetime.strptime(str(dob)[:10], '%Y-%m-%d').date()
                this_year = d.replace(year=today.year)
                if this_year < today:
                    this_year = d.replace(year=today.year+1)
                days_away = (this_year - today).days
                if days_away <= 30:
                    upcoming_birthdays.append({
                        'id': cl['id'], 'name': cl['name'],
                        'whatsapp_number': cl.get('whatsapp_number') or cl.get('phone'),
                        'days_away': days_away,
                        'is_today': days_away == 0,
                        'date_display': d.strftime('%b %d')
                    })
            except: pass
        upcoming_birthdays.sort(key=lambda x: x['days_away'])
    except: pass

    # Client KYC expiry alerts (expired or expiring within 90 days)
    kyc_alerts = []
    try:
        kyc_clients = all_(conn, "SELECT id,name,kyc_expiry_date,kyc_status FROM clients WHERE kyc_expiry_date IS NOT NULL AND disabled IS NOT TRUE")
        # Also add companies with expiring KYC
        kyc_cos = all_(conn, "SELECT id,client_name as name,kyc_expiry_date,kyc_status FROM companies WHERE kyc_expiry_date IS NOT NULL AND disabled IS NOT TRUE")
        for cl in kyc_clients:
            ke = cl.get('kyc_expiry_date')
            if not ke: continue
            dl = days_left(ke)
            if dl is not None and dl <= 90:
                kyc_alerts.append({
                    'id': cl['id'], 'name': cl['name'],
                    'kyc_expiry_date': str(ke)[:10],
                    'days_left': dl, 'is_expired': dl < 0,
                    'kyc_status': cl.get('kyc_status')
                })
        for co in kyc_cos:
            ke = co.get('kyc_expiry_date')
            if not ke: continue
            dl = days_left(ke)
            if dl is not None and dl <= 90:
                kyc_alerts.append({'id': co['id'], 'name': co['name'] + ' (Co)',
                    'kyc_expiry_date': str(ke)[:10], 'days_left': dl,
                    'is_expired': dl < 0, 'kyc_status': co.get('kyc_status'), 'type': 'company'})
        kyc_alerts.sort(key=lambda x: x['days_left'])
    except: pass

    conn.close()
    return render_template('dashboard.html',total_companies=total,active_companies=active,total_clients=total_clients,pep_count=pep_count,eid_exp=eid_exp,
        expired_tl=etl,expiring_30_tl=e30tl,
        expired_ap=eap,expiring_30_ap=e30ap,expired_pass=epass,expiring_30_pass=e30p,
        exp_30=exp_30,exp_60=exp_60,exp_90=exp_90,
        risk_breakdown=risk_bd,doc_breakdown=doc_bd,urgent_companies=urgent,
        open_tasks=otasks,overdue_tasks=overtasks,due_today=dtasks,pending_close=ptasks,
        urgent_tasks=utasks,today=str(today),days_left=days_left,staff_task_counts=staff_task_counts,
        upcoming_birthdays=upcoming_birthdays,kyc_alerts=kyc_alerts)

@app.route('/companies')
@login_required
def companies():
    conn=get_db()
    s=request.args.get('search',''); sf=request.args.get('status','')
    rgf=request.args.get('region',''); modf=request.args.get('mode',''); grpf=request.args.get('group','')
    # Advanced filter: extra fields + multi-term search (each space-separated term
    # must match, so terms act like progressively-refining chips: "UAE Dubai" etc.)
    naf=request.args.get('nature',''); tcf=request.args.get('type_of_client','')
    cof=request.args.get('country',''); rif=request.args.get('risk_status','')
    kyf=request.args.get('kyc_status',''); dsf=request.args.get('doc_status','')
    q='SELECT * FROM companies WHERE 1=1'; p=[]
    for term in [t for t in s.split() if t]:
        q+=' AND (client_name LIKE ? OR ac_code LIKE ? OR mobile LIKE ? OR trade_license_no LIKE ? OR region LIKE ? OR country_of_incorporation LIKE ?)'
        p+=[f'%{term}%']*6
    # Status filter: Active/Inactive use ac_status; Disabled uses the out-of-scope flag.
    # Disabled records are hidden by default unless explicitly requested (or status='all').
    if sf == 'Disabled':
        q+=' AND disabled IS TRUE'
    elif sf == 'all':
        pass  # show everything incl. disabled
    elif sf in ('Active','Inactive'):
        q+=' AND ac_status=? AND disabled IS NOT TRUE'; p.append(sf)
    else:
        q+=' AND disabled IS NOT TRUE'
    if rgf: q+=' AND region=?'; p.append(rgf)
    if modf: q+=' AND mode_of_ac=?'; p.append(modf)
    if grpf: q+=' AND group_name=?'; p.append(grpf)
    if naf: q+=' AND nature=?'; p.append(naf)
    if tcf: q+=' AND type_of_client=?'; p.append(tcf)
    if cof: q+=' AND country_of_incorporation=?'; p.append(cof)
    if rif: q+=' AND risk_status=?'; p.append(rif)
    if kyf: q+=' AND kyc_status=?'; p.append(kyf)
    if dsf: q+=' AND doc_status=?'; p.append(dsf)

    page = max(1, int(request.args.get('page', 1)))
    per_page = 100
    total_count = cnt(conn, f"SELECT COUNT(*) FROM companies WHERE {q.split('WHERE',1)[1]}", p or None)
    rows=all_(conn,q+f' ORDER BY created_at DESC LIMIT {per_page} OFFSET {(page-1)*per_page}',p or None)
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    dd=dropdowns()
    try: groups=all_(conn,'SELECT group_name FROM company_groups ORDER BY group_name')
    except: groups=[]
    conn.close()
    cl=[]
    for c in rows:
        tl=days_left(c['trade_license_expiry']); ap=days_left(c['address_proof_expiry'])
        cl.append({**c,'tl_days':tl,'tl_status':exp_status(tl),'ap_days':ap,'ap_status':exp_status(ap),
                   'trade_license_expiry':str(c['trade_license_expiry']) if c['trade_license_expiry'] else None,
                   'address_proof_expiry':str(c['address_proof_expiry']) if c['address_proof_expiry'] else None})
    return render_template('companies.html',companies=cl,search=s,
        status_filter=sf,region_filter=rgf,mode_filter=modf,group_filter=grpf,
        nature_filter=naf,type_filter=tcf,country_filter=cof,risk_filter=rif,kyc_filter=kyf,doc_filter=dsf,
        regions=dd.get('REGION',[]),modes=dd.get('MODE OF AC',[]),
        natures=dd.get('NATURE',[]),types=dd.get('TYPE OF CLIENT',[]),countries=dd.get('COUNTRY',[]),
        risk_statuses=dd.get('RISK STATUS',['High','Medium','Low','Unspecified']),
        kyc_statuses=dd.get('KYC STATUS',[]),doc_statuses=dd.get('DOC STATUS',['Completed','Incompleted']),
        groups=[g['group_name'] for g in groups],
        page=page, total_pages=total_pages, total_count=total_count, per_page=per_page)

@app.route('/company/new')
@require_perm('companies_add')
def company_new():
    conn=get_db()
    staff=all_(conn,'SELECT id,name FROM users WHERE is_active=1 ORDER BY name')
    conn.close()
    max_kyc = (dubai_today().replace(year=dubai_today().year+2)).isoformat()
    return render_template('company_form.html',dropdown_data=dropdowns(),company=None,ubos=[],edit=False,staff_users=staff,max_kyc_date=max_kyc)

@app.route('/company/<int:id>')
@login_required
def company_detail(id):
    conn=get_db()
    co=one(conn,'SELECT * FROM companies WHERE id=?',(id,))
    if not co: conn.close(); return redirect(url_for('companies'))
    ubos=all_(conn,'SELECT * FROM ubos WHERE company_id=? ORDER BY share_percentage DESC',(id,))
    conn.close()
    tl=days_left(co['trade_license_expiry']); ap=days_left(co['address_proof_expiry'])
    ul=[]
    for u in ubos:
        pd=days_left(u['passport_expiry']); ed=days_left(u['emirates_id_expiry'])
        ul.append({**u,'p_days':pd,'p_status':exp_status(pd),'e_days':ed,'e_status':exp_status(ed),
                   'passport_expiry':str(u['passport_expiry']) if u['passport_expiry'] else None,
                   'emirates_id_expiry':str(u['emirates_id_expiry']) if u['emirates_id_expiry'] else None})
    return render_template('company_detail.html',company=co,ubos=ul,
        tl_days=tl,tl_status=exp_status(tl),ap_days=ap,ap_status=exp_status(ap),
        today=str(dubai_today()))

@app.route('/company/<int:id>/edit')
@admin_required
def company_edit(id):
    conn=get_db()
    co=one(conn,'SELECT * FROM companies WHERE id=?',(id,))
    if not co: conn.close(); return redirect(url_for('companies'))
    ubos=all_(conn,'SELECT * FROM ubos WHERE company_id=? ORDER BY share_percentage DESC',(id,))
    staff=all_(conn,'SELECT id,name FROM users WHERE is_active=1 ORDER BY name')
    conn.close()
    max_kyc = (dubai_today().replace(year=dubai_today().year+2)).isoformat()
    return render_template('company_form.html',dropdown_data=dropdowns(),company=co,ubos=ubos,edit=True,staff_users=staff,max_kyc_date=max_kyc)

def _cv(d):
    return (d.get('ac_opening_date') or None,d.get('ac_status','Active'),d.get('active_till_year'),
        d.get('nature'),d.get('type_of_client'),d.get('name_of_freezone'),d.get('mode_of_ac'),
        d.get('country_of_incorporation'),d.get('region'),d.get('address'),d.get('telephone'),
        d.get('mobile'),d.get('whatsapp_number'),d.get('email_id'),d.get('contact_person_name'),
        d.get('contact_person_number'),d.get('account_manager'),d.get('address_proof_type'),
        d.get('address_proof_expiry') or None,d.get('kyc_status'),d.get('trade_license_no'),
        d.get('issuing_authority'),d.get('legal_type'),d.get('incorporation_date') or None,
        d.get('trade_license_expiry') or None,d.get('tax_no_trn'),d.get('vat_cert'),
        d.get('vat_declaration'),d.get('deal_after_vat'),int(d.get('num_beneficial_owners') or 0),
        d.get('moa'),d.get('pep'),d.get('undertaking'),d.get('source_of_fund'),
        d.get('software_updation'),d.get('doc_status','Incompleted'),d.get('screening_date') or None,
        d.get('registration_screening_tool'),d.get('risk_status','Unspecified'),d.get('verified_by'),
        d.get('verified_date') or None,d.get('followup_details'),d.get('crowe_feedback'),d.get('zewer_comments'),
           d.get('kyc_expiry_date') or None, d.get('id_type'))

def _save_ubos(conn,cid,ubos):
    x(conn,'DELETE FROM ubos WHERE company_id=?',(cid,))
    for u in ubos:
        if not u.get('person_name'): continue
        x(conn,'''INSERT INTO ubos (company_id,position,share_percentage,person_name,nationality,
            residential_status,pep_status,passport_no,passport_expiry,emirates_id,emirates_id_expiry,
            doc_status,verified_by,verified_date,followup_details) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (cid,u.get('position'),u.get('share_percentage') or None,u.get('person_name'),
             u.get('nationality'),u.get('residential_status'),u.get('pep_status'),u.get('passport_no'),
             u.get('passport_expiry') or None,u.get('emirates_id'),u.get('emirates_id_expiry') or None,
             u.get('doc_status','Incompleted'),u.get('verified_by'),u.get('verified_date') or None,u.get('followup_details')))

@app.route('/api/company/add', methods=['POST'])
@require_perm('companies_add')
def api_add_company():
    d=request.get_json()
    kyc_err = _validate_kyc_expiry(d.get('kyc_expiry_date'))
    if kyc_err: return jsonify({'success':False,'error':kyc_err}),400
    try:
        conn=get_db()
        x(conn,'''INSERT INTO companies (ac_code,client_name,ac_opening_date,ac_status,active_till_year,
            nature,type_of_client,name_of_freezone,mode_of_ac,country_of_incorporation,region,address,
            telephone,mobile,whatsapp_number,email_id,contact_person_name,contact_person_number,account_manager,
            address_proof_type,address_proof_expiry,kyc_status,trade_license_no,issuing_authority,legal_type,
            incorporation_date,trade_license_expiry,tax_no_trn,vat_cert,vat_declaration,deal_after_vat,
            num_beneficial_owners,moa,pep,undertaking,source_of_fund,software_updation,doc_status,
            screening_date,registration_screening_tool,risk_status,verified_by,verified_date,
            followup_details,crowe_feedback,zewer_comments,kyc_expiry_date,id_type,created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (d.get('ac_code'),d.get('client_name'))+_cv(d)+(session.get('user_id'),))
        cid=lastid(conn)
        _save_ubos(conn,cid,d.get('ubos',[]))
        commit(conn); conn.close()
        return jsonify({'success':True,'id':cid})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/company/<int:id>/edit', methods=['POST'])
@admin_pw_required
def api_edit_company(id):
    d=request.get_json()
    kyc_err = _validate_kyc_expiry(d.get('kyc_expiry_date'))
    if kyc_err: return jsonify({'success':False,'error':kyc_err}),400
    try:
        conn=get_db()
        x(conn,'''UPDATE companies SET client_name=?,ac_opening_date=?,ac_status=?,active_till_year=?,
            nature=?,type_of_client=?,name_of_freezone=?,mode_of_ac=?,country_of_incorporation=?,region=?,
            address=?,telephone=?,mobile=?,whatsapp_number=?,email_id=?,contact_person_name=?,
            contact_person_number=?,account_manager=?,address_proof_type=?,address_proof_expiry=?,
            kyc_status=?,trade_license_no=?,issuing_authority=?,legal_type=?,incorporation_date=?,
            trade_license_expiry=?,tax_no_trn=?,vat_cert=?,vat_declaration=?,deal_after_vat=?,
            num_beneficial_owners=?,moa=?,pep=?,undertaking=?,source_of_fund=?,software_updation=?,
            doc_status=?,screening_date=?,registration_screening_tool=?,risk_status=?,verified_by=?,
            verified_date=?,followup_details=?,crowe_feedback=?,zewer_comments=?,kyc_expiry_date=?,id_type=?,
            updated_at=CURRENT_TIMESTAMP WHERE id=?''',
            (d.get('client_name'),)+_cv(d)+(id,))
        _save_ubos(conn,id,d.get('ubos',[]))
        commit(conn); conn.close()
        return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/company/<int:id>/delete', methods=['POST'])
@admin_pw_required
def api_delete_company(id):
    try:
        conn=get_db(); x(conn,'DELETE FROM companies WHERE id=?',(id,))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/company/<int:id>/toggle-disable', methods=['POST'])
@admin_pw_required
def api_toggle_disable_company(id):
    """Mark a company as disabled (out of scope) or re-enable it."""
    try:
        conn=get_db()
        co=one(conn,'SELECT disabled FROM companies WHERE id=?',(id,))
        new_val = not bool(co and co.get('disabled'))
        x(conn,'UPDATE companies SET disabled=? WHERE id=?',(new_val,id))
        commit(conn); conn.close()
        return jsonify({'success':True,'disabled':new_val})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/client/<int:id>/toggle-disable', methods=['POST'])
@admin_pw_required
def api_toggle_disable_client(id):
    """Mark an individual as disabled (out of scope) or re-enable it."""
    try:
        conn=get_db()
        cl=one(conn,'SELECT disabled FROM clients WHERE id=?',(id,))
        new_val = not bool(cl and cl.get('disabled'))
        x(conn,'UPDATE clients SET disabled=? WHERE id=?',(new_val,id))
        commit(conn); conn.close()
        return jsonify({'success':True,'disabled':new_val})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/alerts')
@login_required
def alerts():
    conn=get_db(); today=dubai_today()
    # Build unified alert list from all document types
    all_alerts = []

    def _co_status(r):
        return 'Disabled' if r.get('disabled') else (r.get('ac_status') or 'Active')

    # Trade licenses
    tl_rows = all_(conn,'''SELECT id,ac_code,client_name,mobile,whatsapp_number,
        trade_license_expiry,risk_status,region,account_manager,ac_status,disabled,
        contact_person_name,contact_person_number
        FROM companies WHERE trade_license_expiry IS NOT NULL ORDER BY trade_license_expiry''')
    for r in tl_rows:
        d = days_left(r['trade_license_expiry'])
        if d is None: continue
        all_alerts.append({'company_id':r['id'],'ac_code':r['ac_code'],
            'client_name':r['client_name'],'mobile':r['mobile'],
            'whatsapp_number':r['whatsapp_number'],'account_manager':r.get('account_manager'),
            'status':_co_status(r),'region':r.get('region'),'risk_status':r.get('risk_status'),
            'contact_person_name':r.get('contact_person_name'),
            'contact_person_number':r.get('contact_person_number'),
            'doc_type':'Trade License','doc_subtype':None,
            'expiry_date':str(r['trade_license_expiry'])[:10],'days':d})

    # Address proofs
    ap_rows = all_(conn,'''SELECT id,ac_code,client_name,mobile,whatsapp_number,
        address_proof_expiry,address_proof_type,account_manager,ac_status,disabled,region,risk_status,
        contact_person_name,contact_person_number
        FROM companies WHERE address_proof_expiry IS NOT NULL ORDER BY address_proof_expiry''')
    for r in ap_rows:
        d = days_left(r['address_proof_expiry'])
        if d is None: continue
        all_alerts.append({'company_id':r['id'],'ac_code':r['ac_code'],
            'client_name':r['client_name'],'mobile':r['mobile'],
            'whatsapp_number':r['whatsapp_number'],'account_manager':r.get('account_manager'),
            'status':_co_status(r),'region':r.get('region'),'risk_status':r.get('risk_status'),
            'contact_person_name':r.get('contact_person_name'),
            'contact_person_number':r.get('contact_person_number'),
            'doc_type':'Address Proof','doc_subtype':r.get('address_proof_type'),
            'expiry_date':str(r['address_proof_expiry'])[:10],'days':d})

    # Passports
    ubo_rows = all_(conn,'''SELECT u.person_name,u.passport_no,u.passport_expiry,
        u.emirates_id,u.emirates_id_expiry,
        c.id as company_id,c.client_name,c.ac_code,c.mobile,c.whatsapp_number,
        c.account_manager,c.ac_status,c.disabled,c.region,c.risk_status,
        c.contact_person_name,c.contact_person_number
        FROM ubos u JOIN companies c ON u.company_id=c.id
        WHERE u.passport_expiry IS NOT NULL OR u.emirates_id_expiry IS NOT NULL''')
    for r in ubo_rows:
        if r.get('passport_expiry'):
            d = days_left(r['passport_expiry'])
            if d is not None:
                all_alerts.append({'company_id':r['company_id'],'ac_code':r['ac_code'],
                    'client_name':r['client_name'],'mobile':r['mobile'],
                    'whatsapp_number':r['whatsapp_number'],'account_manager':r.get('account_manager'),
                    'status':_co_status(r),'region':r.get('region'),'risk_status':r.get('risk_status'),
                    'contact_person_name':r['person_name'],'contact_person_number':None,
                    'doc_type':'Passport','doc_subtype':r.get('passport_no'),
                    'expiry_date':str(r['passport_expiry'])[:10],'days':d})
        if r.get('emirates_id_expiry'):
            d = days_left(r['emirates_id_expiry'])
            if d is not None:
                all_alerts.append({'company_id':r['company_id'],'ac_code':r['ac_code'],
                    'client_name':r['client_name'],'mobile':r['mobile'],
                    'whatsapp_number':r['whatsapp_number'],'account_manager':r.get('account_manager'),
                    'status':_co_status(r),'region':r.get('region'),'risk_status':r.get('risk_status'),
                    'contact_person_name':r['person_name'],'contact_person_number':None,
                    'doc_type':'Emirates ID','doc_subtype':r.get('emirates_id'),
                    'expiry_date':str(r['emirates_id_expiry'])[:10],'days':d})

    # Clients (Individuals) documents
    client_rows = all_(conn,'''SELECT id,name,phone,whatsapp_number,
        passport_expiry,emirates_id_expiry,account_number,ac_status,disabled,risk_status
        FROM clients WHERE passport_expiry IS NOT NULL OR emirates_id_expiry IS NOT NULL''')
    for r in client_rows:
        st = 'Disabled' if r.get('disabled') else (r.get('ac_status') or 'Active')
        if r.get('passport_expiry'):
            d = days_left(r['passport_expiry'])
            if d is not None:
                all_alerts.append({'company_id':None,'client_id':r['id'],'ac_code':r.get('account_number','N/A'),
                    'client_name':r['name'],'mobile':r.get('phone'),'whatsapp_number':r.get('whatsapp_number'),
                    'account_manager':None,'status':st,'region':None,'risk_status':r.get('risk_status'),
                    'contact_person_name':r['name'],'contact_person_number':r.get('phone'),
                    'doc_type':'Passport (Individual)','doc_subtype':None,
                    'expiry_date':str(r['passport_expiry'])[:10],'days':d})
        if r.get('emirates_id_expiry'):
            d = days_left(r['emirates_id_expiry'])
            if d is not None:
                all_alerts.append({'company_id':None,'client_id':r['id'],'ac_code':r.get('account_number','N/A'),
                    'client_name':r['name'],'mobile':r.get('phone'),'whatsapp_number':r.get('whatsapp_number'),
                    'account_manager':None,'status':st,'region':None,'risk_status':r.get('risk_status'),
                    'contact_person_name':r['name'],'contact_person_number':r.get('phone'),
                    'doc_type':'Emirates ID (Individual)','doc_subtype':None,
                    'expiry_date':str(r['emirates_id_expiry'])[:10],'days':d})

    # Sort by days (most urgent first)
    all_alerts.sort(key=lambda x: x['days'])

    # Distinct values for the advanced filter
    doc_types = sorted(set(a['doc_type'] for a in all_alerts))
    regions = sorted(set(a['region'] for a in all_alerts if a.get('region')))

    # Get users for task assignment
    all_users = all_(conn,'SELECT id,name,role FROM users WHERE is_active=1 ORDER BY name')

    conn.close()
    return render_template('alerts.html', all_alerts=all_alerts,
        doc_types=doc_types, regions=regions,
        all_users=all_users, today=str(today))

@app.route('/health-check')
@compliance_required
def health_check():
    conn = get_db()
    companies = all_(conn, 'SELECT id, ac_code, client_name FROM companies WHERE disabled IS NOT TRUE ORDER BY client_name')
    conn.close()
    return render_template('health_check.html', companies=companies)

@app.route('/health-check/<int:id>')
@compliance_required
def health_check_detail(id):
    conn = get_db()
    co = one(conn, 'SELECT * FROM companies WHERE id=?', (id,))
    if not co: conn.close(); return redirect(url_for('health_check'))
    if 'health_note' not in co: co['health_note'] = ''
    ubos = all_(conn, 'SELECT * FROM ubos WHERE company_id=? ORDER BY share_percentage DESC', (id,))
    
    # Get latest risk assessment
    risk_assessment = one(conn, """
        SELECT final_score, risk_rating, assessment_date FROM company_risk_assessments 
        WHERE company_id = ? ORDER BY assessment_date DESC LIMIT 1
    """, (id,))
    
    # Convert date object to string for Jinja
    if risk_assessment and risk_assessment.get('assessment_date'):
        risk_assessment = dict(risk_assessment)
        risk_assessment['assessment_date'] = str(risk_assessment['assessment_date'])
    
    conn.close()
    today = dubai_today()

    def doc_status(expiry):
        dl = days_left(expiry)
        if expiry is None: return 'missing', None
        if dl is None: return 'missing', None
        if dl < 0: return 'expired', dl
        if dl <= 30: return 'critical', dl
        if dl <= 90: return 'warning', dl
        return 'ok', dl

    # Build document list with icons
    docs = []

    # Trade License
    tl_st, tl_dl = doc_status(co.get('trade_license_expiry'))
    docs.append({'name': 'Trade License', 'icon': '📜', 'ref': co.get('trade_license_no'), 'expiry': str(co.get('trade_license_expiry',''))[:10] if co.get('trade_license_expiry') else None, 'days': tl_dl, 'status': tl_st, 'category': 'Company'})

    # Address Proof
    ap_st, ap_dl = doc_status(co.get('address_proof_expiry'))
    docs.append({'name': 'Address Proof', 'icon': '🏠', 'ref': co.get('address_proof_type'), 'expiry': str(co.get('address_proof_expiry',''))[:10] if co.get('address_proof_expiry') else None, 'days': ap_dl, 'status': ap_st, 'category': 'Company'})

    # VAT / TRN
    vat_present = bool(co.get('tax_no_trn'))
    docs.append({'name': 'VAT / TRN', 'icon': '🧾', 'ref': co.get('tax_no_trn'), 'expiry': None, 'days': None, 'status': 'ok' if vat_present else 'missing', 'category': 'Tax'})

    # KYC Review
    kyc_st, kyc_dl = doc_status(co.get('kyc_expiry_date'))
    docs.append({'name': 'KYC Review', 'icon': '🛡️', 'ref': co.get('kyc_status'), 'expiry': str(co.get('kyc_expiry_date',''))[:10] if co.get('kyc_expiry_date') else None, 'days': kyc_dl, 'status': kyc_st, 'category': 'Compliance'})

    # UBO documents
    for u in ubos:
        pp_st, pp_dl = doc_status(u.get('passport_expiry'))
        docs.append({'name': f"Passport", 'icon': '🛂', 'ref': u['person_name'], 'expiry': str(u.get('passport_expiry',''))[:10] if u.get('passport_expiry') else None, 'days': pp_dl, 'status': pp_st, 'category': u.get('position') or 'UBO'})
        eid_st, eid_dl = doc_status(u.get('emirates_id_expiry'))
        docs.append({'name': f"Emirates ID", 'icon': '🪪', 'ref': u['person_name'], 'expiry': str(u.get('emirates_id_expiry',''))[:10] if u.get('emirates_id_expiry') else None, 'days': eid_dl, 'status': eid_st, 'category': u.get('position') or 'UBO'})

    # ── SUMMARY COUNTS ──
    summary = {
        'total': len(docs),
        'valid': sum(1 for d in docs if d['status'] == 'ok'),
        'expiring': sum(1 for d in docs if d['status'] in ('warning','critical')),
        'expired': sum(1 for d in docs if d['status'] == 'expired'),
        'missing': sum(1 for d in docs if d['status'] == 'missing'),
    }

    # ── HEALTH SCORE CALCULATION ──
    score = 100
    deductions = []
    for d in docs:
        if d['status'] == 'expired':
            score -= 15; deductions.append(f"{d['name']} expired")
        elif d['status'] == 'critical':
            score -= 8; deductions.append(f"{d['name']} expiring in {d['days']}d")
        elif d['status'] == 'warning':
            score -= 4; deductions.append(f"{d['name']} expiring in {d['days']}d")
        elif d['status'] == 'missing':
            score -= 5; deductions.append(f"{d['name']} missing")

    risk = (co.get('risk_status') or '').lower()
    if risk == 'high': score -= 10; deductions.append('High risk client')
    elif risk == 'medium': score -= 5; deductions.append('Medium risk client')
    
    # Factor in risk assessment if available
    if risk_assessment:
        if risk_assessment.get('risk_rating') == 'High':
            score -= 10; deductions.append(f"Risk Assessment: High ({risk_assessment.get('final_score')})")
        elif risk_assessment.get('risk_rating') == 'Medium':
            score -= 5; deductions.append(f"Risk Assessment: Medium ({risk_assessment.get('final_score')})")
        # Low risk doesn't deduct

    kyc = (co.get('kyc_status') or '').lower()
    if 'not' in kyc or 'pending' in kyc: score -= 8; deductions.append('KYC not completed')
    elif 'expired' in kyc: score -= 10; deductions.append('KYC expired')

    if co.get('doc_status') == 'Incompleted': score -= 5; deductions.append('Documents incomplete')
    if (co.get('pep') or '').lower() == 'yes': score -= 5; deductions.append('PEP flagged')

    score = max(0, min(100, score))

    if score >= 90: grade = 'Excellent'; grade_color = '#22c55e'
    elif score >= 70: grade = 'Good'; grade_color = '#86efac'
    elif score >= 50: grade = 'Average'; grade_color = '#f59e0b'
    else: grade = 'Poor'; grade_color = '#ef4444'

    return render_template('health_check_detail.html',
        company=co, ubos=ubos, docs=docs, summary=summary,
        score=score, grade=grade, grade_color=grade_color,
        deductions=deductions, today=str(today), risk_assessment=risk_assessment)

@app.route('/api/company/<int:id>/health-note', methods=['POST'])
@compliance_required
def api_save_health_note(id):
    d = request.get_json()
    try:
        conn = get_db()
        # Ensure column exists (Railway PG may not have run migration)
        try:
            if is_pg(conn):
                x(conn, 'ALTER TABLE companies ADD COLUMN IF NOT EXISTS health_note TEXT')
            else:
                x(conn, 'ALTER TABLE companies ADD COLUMN health_note TEXT')
            commit(conn)
        except:
            try: conn.rollback()
            except: pass
        x(conn, 'UPDATE companies SET health_note=? WHERE id=?', (d.get('note',''), id))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f'Error saving health note: {e}')
        try: conn.close()
        except: pass
        return _fail(e)

# ──────────── RISK ASSESSMENT ────────────
def calculate_risk_score(scores):
    """Calculate final risk score and rating from individual factor scores."""
    if not scores:
        return None, 'Unspecified'
    valid = [s for s in scores if s is not None and s > 0]
    if not valid:
        return None, 'Unspecified'
    avg = sum(valid) / len(valid)
    if avg <= 1.00:
        return round(avg, 2), 'Low'
    elif avg <= 2.00:
        return round(avg, 2), 'Medium'
    else:
        return round(avg, 2), 'High'

def _risk_factors(assessment, atype):
    """Return [{label, score, level}] for the factors relevant to the assessment type,
    plus a calculation summary so the print can show how the score was derived."""
    company_factors = [
        ('Jurisdiction', 'jurisdiction_score'), ('Ownership', 'ownership_score'),
        ('Delivery Channel', 'delivery_channel_score'), ('Payment Method', 'payment_method_score'),
        ('Transaction Volume', 'transaction_volume_score'), ('Product', 'product_score'),
        ('PEP Status', 'pep_status_score'), ('Nationality', 'nationality_score'),
        ('Years of Relationship', 'years_relationship_score'), ('Years of Operation', 'years_operation_score'),
        ('Third-Party Involvement', 'third_party_score'), ('Sanctions', 'sanctions_score'),
    ]
    individual_factors = [
        ('Nationality', 'nationality_score'), ('Residence Status', 'residence_status_score'),
        ('PEP Status', 'pep_status_score'), ('Profession', 'profession_score'),
        ('Product', 'product_score'), ('Delivery Channel', 'delivery_channel_score'),
        ('Payment Method', 'payment_method_score'), ('Transaction Amount', 'transaction_amount_score'),
        ('Years of Relationship', 'years_relationship_score'), ('Place of Birth', 'place_of_birth_score'),
        ('Third-Party Involvement', 'third_party_score'), ('Sanctions', 'sanctions_score'),
    ]
    src = individual_factors if atype == 'individual' else company_factors
    factors, considered = [], []
    for label, key in src:
        sc = assessment.get(key)
        if sc is None:
            continue
        try:
            scf = float(sc)
        except (TypeError, ValueError):
            continue
        level = 'Low' if scf <= 1 else ('Medium' if scf <= 2 else 'High')
        factors.append({'label': label, 'score': scf, 'level': level})
        if scf > 0:
            considered.append(scf)
    calc = {
        'count': len(considered),
        'total': round(sum(considered), 2),
        'average': round(sum(considered) / len(considered), 2) if considered else 0,
    }
    return factors, calc

@app.route('/api/risk-lookup', methods=['POST'])
@compliance_required
def api_risk_lookup():
    """Lookup risk score for a value in a category"""
    data = request.get_json()
    category = data.get('category')
    value = data.get('value')
    
    if not category or not value:
        return jsonify({'score': None}), 400
    
    score = RISK_LOOKUPS.get(category, {}).get(value)
    return jsonify({'score': score})

@app.route('/api/send-risk-email', methods=['POST'])
@compliance_required
def api_send_risk_email():
    """Send risk assessment via email"""
    data = request.get_json()
    email = data.get('email')
    company = data.get('company')
    rating = data.get('rating')
    score = data.get('score')
    
    if not email or not company:
        return jsonify({'success': False, 'error': 'Missing email or company'}), 400
    
    try:
        # For now, just return success (email sending requires SMTP setup)
        logger.info(f'Risk assessment email requested for {company} to {email}')
        # TODO: Integrate actual email sending (SendGrid, SMTP, etc.)
        return jsonify({'success': True, 'message': f'Risk assessment for {company} ({rating} - {score}) ready to send'})
    except Exception as e:
        logger.error(f'Error sending risk email: {e}')
        return _fail(e)

@app.route('/risk-assessment-list')
@compliance_required
def risk_assessment_list():
    """View all risk assessments with filters"""
    try:
        conn = get_db()
        
        # Ensure tables exist on Railway
        if is_pg(conn):
            x(conn, '''CREATE TABLE IF NOT EXISTS company_risk_assessments (
                id SERIAL PRIMARY KEY, company_id INTEGER NOT NULL,
                jurisdiction_score FLOAT, ownership_score FLOAT, delivery_channel_score FLOAT,
                payment_method_score FLOAT, transaction_volume_score FLOAT, product_score FLOAT,
                pep_status_score FLOAT, nationality_score FLOAT, years_relationship_score FLOAT,
                years_operation_score FLOAT, third_party_score FLOAT, sanctions_score FLOAT,
                final_score FLOAT, risk_rating TEXT, assessment_date DATE, notes TEXT, assessed_by INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            x(conn, '''CREATE TABLE IF NOT EXISTS individual_risk_assessments (
                id SERIAL PRIMARY KEY, individual_id INTEGER NOT NULL,
                nationality_score FLOAT, residence_status_score FLOAT, pep_status_score FLOAT,
                profession_score FLOAT, product_score FLOAT, delivery_channel_score FLOAT,
                payment_method_score FLOAT, transaction_amount_score FLOAT, years_relationship_score FLOAT,
                place_of_birth_score FLOAT, third_party_score FLOAT, sanctions_score FLOAT,
                final_score FLOAT, risk_rating TEXT, assessment_date DATE, notes TEXT, assessed_by INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            commit(conn)
        
        # Get companies with their latest assessment status
        companies = all_(conn, """
            SELECT c.id, c.client_name as name, c.ac_code, 'company' as type,
                   COALESCE(ca.final_score, 0) as final_score,
                   COALESCE(ca.risk_rating, 'Pending') as risk_rating,
                   ca.assessment_date,
                   CASE WHEN ca.id IS NOT NULL THEN 'done' ELSE 'pending' END as status
            FROM companies c
            LEFT JOIN company_risk_assessments ca ON ca.id = (
                SELECT id FROM company_risk_assessments
                WHERE company_id = c.id
                ORDER BY assessment_date DESC, id DESC LIMIT 1
            )
            WHERE c.disabled IS NOT TRUE
            ORDER BY c.client_name
        """) or []
        
        # Get individuals with their latest assessment status
        individuals = all_(conn, """
            SELECT cl.id, cl.name, NULL as ac_code, 'individual' as type,
                   COALESCE(ia.final_score, 0) as final_score,
                   COALESCE(ia.risk_rating, 'Pending') as risk_rating,
                   ia.assessment_date,
                   CASE WHEN ia.id IS NOT NULL THEN 'done' ELSE 'pending' END as status
            FROM clients cl
            LEFT JOIN individual_risk_assessments ia ON ia.id = (
                SELECT id FROM individual_risk_assessments
                WHERE individual_id = cl.id
                ORDER BY assessment_date DESC, id DESC LIMIT 1
            )
            WHERE cl.disabled IS NOT TRUE
            ORDER BY cl.name
        """) or []
        
        # Combine
        assessments = companies + individuals
        assessments = sorted(assessments, key=lambda x: str(x.get('assessment_date') or '2000-01-01'), reverse=True)
        
        # Count by rating (only done ones)
        done_assessments = [a for a in assessments if a.get('status') == 'done']
        low = sum(1 for a in done_assessments if a.get('risk_rating') == 'Low')
        medium = sum(1 for a in done_assessments if a.get('risk_rating') == 'Medium')
        high = sum(1 for a in done_assessments if a.get('risk_rating') == 'High')
        
        conn.close()
        
        return render_template('risk_assessment_list.html',
                             assessments=assessments,
                             total_count=len(done_assessments),
                             low_count=low,
                             medium_count=medium,
                             high_count=high)
    except Exception as e:
        logger.error(f'Error loading risk assessment list: {e}', exc_info=True)
        try: conn.close()
        except: pass
        logger.error(f'{request.path}: {e}')
        return '<h1>Error</h1><p>Something went wrong. Please try again.</p>', 500


@app.route('/risk-assessment/walkin', methods=['GET','POST'])
@compliance_required
def risk_assessment_walkin():
    """Walk-in assessment - for new clients before they are in the system"""
    if request.method == 'POST':
        data = request.form
        try:
            conn = get_db()
            if is_pg(conn):
                x(conn, '''CREATE TABLE IF NOT EXISTS walkin_risk_assessments (
                    id SERIAL PRIMARY KEY,
                    entity_name TEXT NOT NULL, entity_type TEXT DEFAULT 'company',
                    jurisdiction_score FLOAT, ownership_score FLOAT, delivery_channel_score FLOAT,
                    payment_method_score FLOAT, transaction_volume_score FLOAT, product_score FLOAT,
                    pep_status_score FLOAT, nationality_score FLOAT, years_relationship_score FLOAT,
                    years_operation_score FLOAT, third_party_score FLOAT, sanctions_score FLOAT,
                    final_score FLOAT, risk_rating TEXT, assessment_date DATE, notes TEXT, assessed_by INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
                commit(conn)
            
            scores = [float(data.get(f) or 0) for f in [
                'jurisdiction_score','ownership_score','delivery_channel_score','payment_method_score',
                'transaction_volume_score','product_score','pep_status_score','nationality_score',
                'years_relationship_score','years_operation_score','third_party_score','sanctions_score'
            ]]
            final_score, risk_rating = calculate_risk_score(scores)
            
            x(conn, '''INSERT INTO walkin_risk_assessments 
               (entity_name, entity_type, jurisdiction_score, ownership_score, delivery_channel_score,
                payment_method_score, transaction_volume_score, product_score, pep_status_score,
                nationality_score, years_relationship_score, years_operation_score, third_party_score,
                sanctions_score, final_score, risk_rating, assessment_date, notes, assessed_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
               (data.get('entity_name',''), data.get('entity_type','company'),
                scores[0],scores[1],scores[2],scores[3],scores[4],scores[5],scores[6],
                scores[7],scores[8],scores[9],scores[10],scores[11],
                final_score, risk_rating, dubai_today(), data.get('notes',''), session.get('user_id')))
            commit(conn)
            walkin_id = one(conn, 'SELECT id FROM walkin_risk_assessments ORDER BY created_at DESC LIMIT 1', ())
            conn.close()
            return redirect(url_for('risk_assessment_walkin_result', id=walkin_id['id']))
        except Exception as e:
            logger.error(f'Walk-in assessment error: {e}', exc_info=True)
            try: conn.close()
            except: pass
            logger.error(f'{request.path}: {e}')
        return '<h1>Error</h1><p>Something went wrong. Please try again.</p>', 500

    refresh_country_scores()
    return render_template('risk_assessment_walkin.html', lookups=RISK_LOOKUPS)


@app.route('/risk-assessment/walkin/<int:id>/result')
@compliance_required
def risk_assessment_walkin_result(id):
    conn = get_db()
    assessment = one(conn, 'SELECT * FROM walkin_risk_assessments WHERE id=?', (id,))
    conn.close()
    if not assessment:
        return redirect(url_for('risk_assessment_list'))
    assessment = dict(assessment)
    if assessment.get('assessment_date'):
        assessment['assessment_date'] = str(assessment['assessment_date'])
    risk_colors = {'Low': '#22c55e', 'Medium': '#f59e0b', 'High': '#ef4444'}
    factors, calc = _risk_factors(assessment, 'walkin')
    return render_template('risk_assessment_result.html',
        assessment=assessment,
        assessment_type='walkin',
        company=None,
        individual=None,
        factors=factors, calc=calc,
        risk_color=risk_colors.get(assessment.get('risk_rating','Low'), '#22c55e'))

@app.route('/risk-assessment')
@compliance_required
def risk_assessment():
    """Select company or individual for risk assessment"""
    try:
        conn = get_db()
        companies = all_(conn, 'SELECT id, ac_code, client_name FROM companies WHERE disabled IS NOT TRUE ORDER BY client_name')
        # Individuals are stored in clients table
        individuals = all_(conn, 'SELECT id, name FROM clients WHERE disabled IS NOT TRUE ORDER BY name')
        conn.close()
        logger.info(f'Risk assessment: {len(companies or [])} companies, {len(individuals or [])} individuals')
        return render_template('risk_assessment.html', 
                             companies=companies or [], 
                             individuals=individuals or [])
    except Exception as e:
        logger.error(f'Error in risk_assessment: {e}', exc_info=True)
        try: conn.close()
        except: pass
        logger.error(f'{request.path}: {e}')
        return '<h1>Error</h1><p>Something went wrong. Please try again.</p>', 500

@app.route('/risk-assessment/company/<int:id>', methods=['GET','POST'])
@compliance_required
def risk_assessment_company(id):
    """Risk assessment form for company"""
    conn = get_db()
    co = one(conn, 'SELECT * FROM companies WHERE id=?', (id,))
    if not co:
        conn.close()
        return redirect(url_for('risk_assessment'))
    
    if request.method == 'POST':
        data = request.form
        logger.info(f'Risk assessment form received for company {id}')
        
        try:
            # Ensure tables exist on Railway
            if is_pg(conn):
                x(conn, '''CREATE TABLE IF NOT EXISTS company_risk_assessments (
                    id SERIAL PRIMARY KEY, company_id INTEGER NOT NULL,
                    jurisdiction_score FLOAT, ownership_score FLOAT, delivery_channel_score FLOAT,
                    payment_method_score FLOAT, transaction_volume_score FLOAT, product_score FLOAT,
                    pep_status_score FLOAT, nationality_score FLOAT, years_relationship_score FLOAT,
                    years_operation_score FLOAT, third_party_score FLOAT, sanctions_score FLOAT,
                    final_score FLOAT, risk_rating TEXT, assessment_date DATE, notes TEXT, assessed_by INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
                commit(conn)
            
            scores = [
                float(data.get('jurisdiction_score') or 0),
                float(data.get('ownership_score') or 0),
                float(data.get('delivery_channel_score') or 0),
                float(data.get('payment_method_score') or 0),
                float(data.get('transaction_volume_score') or 0),
                float(data.get('product_score') or 0),
                float(data.get('pep_status_score') or 0),
                float(data.get('nationality_score') or 0),
                float(data.get('years_relationship_score') or 0),
                float(data.get('years_operation_score') or 0),
                float(data.get('third_party_score') or 0),
                float(data.get('sanctions_score') or 0),
            ]
            logger.info(f'Parsed scores: {scores}')
            final_score, risk_rating = calculate_risk_score(scores)
            logger.info(f'Calculated risk: score={final_score}, rating={risk_rating}')
            
            
            x(conn, '''INSERT INTO company_risk_assessments 
               (company_id, jurisdiction_score, ownership_score, delivery_channel_score, payment_method_score,
                transaction_volume_score, product_score, pep_status_score, nationality_score, 
                years_relationship_score, years_operation_score, third_party_score, sanctions_score,
                final_score, risk_rating, assessment_date, notes, assessed_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
               (id, scores[0], scores[1], scores[2], scores[3], scores[4], scores[5], scores[6],
                scores[7], scores[8], scores[9], scores[10], scores[11], final_score, risk_rating,
                dubai_today(), data.get('notes',''), session.get('user_id')))
            commit(conn)
            logger.info(f'✅ Risk assessment saved for company {id}')
            conn.close()
            return redirect(url_for('risk_assessment_company_result', id=id))
        except Exception as e:
            logger.error(f'❌ Error saving company risk assessment: {str(e)}', exc_info=True)
            try: conn.close()
            except: pass
            return redirect(url_for('risk_assessment'))
    
    # Prepare pre-fill data from company
    co_country = co.get('country_of_incorporation') or co.get('region') or 'UNITED ARAB EMIRATES'
    co_pep = (co.get('pep') or '').strip()
    
    prefill = {
        'jurisdiction': co_country,
        'nationality': co_country,   # same country as incorporation
        'pep_status': '1' if co_pep.lower() == 'no' or co_pep == 'Not Applicable' or co_pep == '' else ('3' if co_pep.lower() == 'yes' else '1'),
        'third_party': 'No',         # default safe
    }
    
    # Calculate years in operation
    years_operation = 3  # default high risk
    if co.get('incorporation_date'):
        try:
            inc_date = co['incorporation_date'] if isinstance(co['incorporation_date'], str) else str(co['incorporation_date'])[:10]
            from datetime import datetime
            inc = datetime.strptime(inc_date, '%Y-%m-%d').date()
            years_diff = (dubai_today() - inc).days / 365.25
            if years_diff >= 5:
                years_operation = 1
            elif years_diff >= 1:
                years_operation = 2
        except:
            pass
    prefill['years_operation'] = years_operation
    
    # Calculate years in relationship
    years_relationship = 3  # default high risk
    if co.get('ac_opening_date'):
        try:
            opening = co['ac_opening_date'] if isinstance(co['ac_opening_date'], str) else str(co['ac_opening_date'])[:10]
            from datetime import datetime
            open_dt = datetime.strptime(opening, '%Y-%m-%d').date()
            years_diff = (dubai_today() - open_dt).days / 365.25
            if years_diff >= 5:
                years_relationship = 1
            elif years_diff >= 1:
                years_relationship = 2
        except:
            pass
    prefill['years_relationship'] = years_relationship
    
    conn.close()
    refresh_country_scores()
    return render_template('risk_assessment_company.html', company=co, prefill=prefill, lookups=RISK_LOOKUPS)

@app.route('/risk-assessment/company/<int:id>/result')
@compliance_required
def risk_assessment_company_result(id):
    """Display company risk assessment result"""
    conn = get_db()
    co = one(conn, 'SELECT * FROM companies WHERE id=?', (id,))
    assessment = one(conn, '''SELECT * FROM company_risk_assessments 
                             WHERE company_id=? ORDER BY created_at DESC LIMIT 1''', (id,))
    conn.close()
    
    if not co or not assessment:
        return redirect(url_for('risk_assessment'))
    
    # Determine color based on risk rating
    color_map = {'Low': '#22c55e', 'Medium': '#f59e0b', 'High': '#ef4444'}
    risk_color = color_map.get(assessment.get('risk_rating', 'Unspecified'), '#6b7280')
    
    factors, calc = _risk_factors(assessment, 'company')
    return render_template('risk_assessment_result.html',
        company=co, assessment=assessment, risk_color=risk_color, assessment_type='company',
        factors=factors, calc=calc)

@app.route('/risk-assessment/individual/<int:id>', methods=['GET','POST'])
@compliance_required
def risk_assessment_individual(id):
    """Risk assessment form for individual/client"""
    conn = get_db()
    ind = one(conn, 'SELECT * FROM clients WHERE id=?', (id,))
    if not ind:
        conn.close()
        return redirect(url_for('risk_assessment'))
    
    if request.method == 'POST':
        data = request.form
        try:
            # Ensure tables exist on Railway
            if is_pg(conn):
                x(conn, '''CREATE TABLE IF NOT EXISTS individual_risk_assessments (
                    id SERIAL PRIMARY KEY, individual_id INTEGER NOT NULL,
                    nationality_score FLOAT, residence_status_score FLOAT, pep_status_score FLOAT,
                    profession_score FLOAT, product_score FLOAT, delivery_channel_score FLOAT,
                    payment_method_score FLOAT, transaction_amount_score FLOAT, years_relationship_score FLOAT,
                    place_of_birth_score FLOAT, third_party_score FLOAT, sanctions_score FLOAT,
                    final_score FLOAT, risk_rating TEXT, assessment_date DATE, notes TEXT, assessed_by INTEGER,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
                commit(conn)
            
            scores = [
                float(data.get('nationality_score') or 0),
                float(data.get('residence_status_score') or 0),
                float(data.get('pep_status_score') or 0),
                float(data.get('profession_score') or 0),
                float(data.get('product_score') or 0),
                float(data.get('delivery_channel_score') or 0),
                float(data.get('payment_method_score') or 0),
                float(data.get('transaction_amount_score') or 0),
                float(data.get('years_relationship_score') or 0),
                float(data.get('place_of_birth_score') or 0),
                float(data.get('third_party_score') or 0),
                float(data.get('sanctions_score') or 0),
            ]
            final_score, risk_rating = calculate_risk_score(scores)
            
            
            x(conn, '''INSERT INTO individual_risk_assessments 
               (individual_id, nationality_score, residence_status_score, pep_status_score, profession_score,
                product_score, delivery_channel_score, payment_method_score, transaction_amount_score,
                years_relationship_score, place_of_birth_score, third_party_score, sanctions_score,
                final_score, risk_rating, assessment_date, notes, assessed_by)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
               (id, scores[0], scores[1], scores[2], scores[3], scores[4], scores[5], scores[6],
                scores[7], scores[8], scores[9], scores[10], scores[11], final_score, risk_rating,
                dubai_today(), data.get('notes',''), session.get('user_id')))
            commit(conn)
            conn.close()
            return redirect(url_for('risk_assessment_individual_result', id=id))
        except Exception as e:
            logger.error(f'Error saving individual risk assessment: {e}')
            conn.close()
            return redirect(url_for('risk_assessment'))
    
    # Prepare pre-fill data
    prefill = {
        'nationality': ind.get('nationality', ''),
        'residence_status': 'Resident' if ind.get('is_resident') else 'Non-Resident',
        'pep_status': ind.get('pep_status', 'No'),
        'profession': ind.get('profession', ''),
        'place_of_birth': ind.get('nationality', ''),
        'payment_method': '',
        'third_party': 'No',
    }
    
    # Calculate years in relationship
    years_relationship = 3  # default high risk
    if ind.get('created_at'):
        try:
            created = ind['created_at'] if isinstance(ind['created_at'], str) else str(ind['created_at'])[:10]
            from datetime import datetime
            created_dt = datetime.strptime(created, '%Y-%m-%d').date()
            years_diff = (dubai_today() - created_dt).days / 365.25
            if years_diff >= 5:
                years_relationship = 1
            elif years_diff >= 1:
                years_relationship = 2
        except:
            pass
    prefill['years_relationship'] = years_relationship
    
    conn.close()
    refresh_country_scores()
    return render_template('risk_assessment_individual.html', individual=ind, prefill=prefill, lookups=RISK_LOOKUPS)

@app.route('/risk-assessment/individual/<int:id>/result')
@compliance_required
def risk_assessment_individual_result(id):
    """Display individual risk assessment result"""
    conn = get_db()
    ind = one(conn, 'SELECT * FROM clients WHERE id=?', (id,))
    assessment = one(conn, '''SELECT * FROM individual_risk_assessments 
                             WHERE individual_id=? ORDER BY created_at DESC LIMIT 1''', (id,))
    conn.close()
    
    if not ind or not assessment:
        return redirect(url_for('risk_assessment'))
    
    color_map = {'Low': '#22c55e', 'Medium': '#f59e0b', 'High': '#ef4444'}
    risk_color = color_map.get(assessment.get('risk_rating', 'Unspecified'), '#6b7280')
    
    factors, calc = _risk_factors(assessment, 'individual')
    return render_template('risk_assessment_result.html',
        individual=ind, assessment=assessment, risk_color=risk_color, assessment_type='individual',
        factors=factors, calc=calc)

@app.route('/reports')
@login_required
def reports():
    conn=get_db(); today=dubai_today()
    df=request.args.get('from',''); dt=request.args.get('to','')
    w='disabled IS NOT TRUE'; p=[]
    if df: w+=' AND created_at >= ?'; p.append(df)
    if dt: w+=' AND created_at <= ?'; p.append(dt+' 23:59:59')
    total=cnt(conn,f'SELECT COUNT(*) FROM companies WHERE {w}',p or None)
    def q(sql): return all_(conn,sql,p or None)
    res=render_template('reports.html',
        risk_data=q(f'SELECT risk_status,COUNT(*) as c FROM companies WHERE {w} GROUP BY risk_status'),
        doc_data=q(f'SELECT doc_status,COUNT(*) as c FROM companies WHERE {w} GROUP BY doc_status'),
        type_data=q(f'SELECT type_of_client,COUNT(*) as c FROM companies WHERE {w} GROUP BY type_of_client ORDER BY c DESC'),
        region_data=q(f'SELECT region,COUNT(*) as c FROM companies WHERE {w} GROUP BY region ORDER BY c DESC LIMIT 10'),
        kyc_data=q(f'SELECT kyc_status,COUNT(*) as c FROM companies WHERE {w} GROUP BY kyc_status ORDER BY c DESC'),
        mode_data=q(f'SELECT mode_of_ac,COUNT(*) as c FROM companies WHERE {w} GROUP BY mode_of_ac ORDER BY c DESC'),
        total=total,date_from=df,date_to=dt,
        expired_tl=cnt(conn,'SELECT COUNT(*) FROM companies WHERE trade_license_expiry<? AND disabled IS NOT TRUE',(today,)),
        expiring_30=cnt(conn,'SELECT COUNT(*) FROM companies WHERE trade_license_expiry BETWEEN ? AND ? AND disabled IS NOT TRUE',(today,today+timedelta(days=30))),
        expiring_90=cnt(conn,'SELECT COUNT(*) FROM companies WHERE trade_license_expiry BETWEEN ? AND ? AND disabled IS NOT TRUE',(today,today+timedelta(days=90))))
    conn.close(); return res

# ──────────── AML TRACKER ────────────
@app.route('/aml-tracker')
@compliance_required
def aml_tracker():
    """List all AML Tracker records with filters"""
    try:
        conn = get_db()
        
        # Fallback: Ensure table exists (on-first-use pattern for Railway PostgreSQL)
        if is_pg(conn):
            x(conn, '''CREATE TABLE IF NOT EXISTS aml_tracker (
                id SERIAL PRIMARY KEY, company_id INTEGER, individual_id INTEGER,
                transaction_date DATE NOT NULL, period TEXT, due_date DATE, vc_no TEXT,
                payment_mode TEXT, ac_type TEXT, client_name TEXT, transaction_currency TEXT,
                usd_amount NUMERIC, aed_amount NUMERIC, payment_remarks TEXT, invoice_no TEXT,
                invoice_amount NUMERIC, invoice_currency TEXT, goaml_submission_date DATE,
                goaml_status TEXT DEFAULT 'pending', goaml_ref_no TEXT, submitted_by INTEGER NOT NULL,
                checked_by INTEGER, comment TEXT, verified_ledger BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            commit(conn)
        else:
            x(conn, '''CREATE TABLE IF NOT EXISTS aml_tracker (
                id INTEGER PRIMARY KEY AUTOINCREMENT, company_id INTEGER, individual_id INTEGER,
                transaction_date DATE NOT NULL, period TEXT, due_date DATE, vc_no TEXT,
                payment_mode TEXT, ac_type TEXT, client_name TEXT, transaction_currency TEXT,
                usd_amount NUMERIC, aed_amount NUMERIC, payment_remarks TEXT, invoice_no TEXT,
                invoice_amount NUMERIC, invoice_currency TEXT, goaml_submission_date DATE,
                goaml_status TEXT DEFAULT 'pending', goaml_ref_no TEXT, submitted_by INTEGER NOT NULL,
                checked_by INTEGER, comment TEXT, verified_ledger BOOLEAN DEFAULT FALSE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        status_filter = request.args.get('status', '')
        company_filter = request.args.get('company', '')
        vc_no_filter = request.args.get('vc_no', '')
        submitted_by_filter = request.args.get('submitted_by', '')
        period_filter = request.args.get('period', '')   # today | week | month | range
        date_from = request.args.get('from', '')
        date_to = request.args.get('to', '')

        # Build WHERE clause
        where = []
        params = []

        if status_filter:
            where.append(f'goaml_status = {P()}')
            params.append(status_filter)
        if company_filter:
            where.append(f'company_id = {P()}')
            params.append(company_filter)
        if vc_no_filter:
            where.append(f'vc_no LIKE {P()}')
            params.append(f'%{vc_no_filter}%')
        if submitted_by_filter:
            where.append(f'submitted_by = {P()}')
            params.append(submitted_by_filter)

        # Period search on transaction_date
        _td = dubai_today()
        if period_filter == 'today':
            where.append(f'transaction_date = {P()}'); params.append(str(_td))
        elif period_filter == 'week':
            _mon = _td - timedelta(days=_td.weekday())
            where.append(f'transaction_date BETWEEN {P()} AND {P()}'); params += [str(_mon), str(_mon + timedelta(days=6))]
        elif period_filter == 'month':
            _first = _td.replace(day=1)
            where.append(f'transaction_date BETWEEN {P()} AND {P()}'); params += [str(_first), str(_td)]
        elif period_filter == 'range':
            if date_from:
                where.append(f'transaction_date >= {P()}'); params.append(date_from)
            if date_to:
                where.append(f'transaction_date <= {P()}'); params.append(date_to)

        where_clause = ' AND '.join(where) if where else '1=1'
        
        # Get AML records with company/individual/user names
        sql = f'''SELECT a.*, c.client_name as company_name, cli.name as individual_name, u.email as submitted_by_email, 
                  u2.email as checked_by_email
                  FROM aml_tracker a
                  LEFT JOIN companies c ON a.company_id = c.id
                  LEFT JOIN clients cli ON a.individual_id = cli.id
                  LEFT JOIN users u ON a.submitted_by = u.id
                  LEFT JOIN users u2 ON a.checked_by = u2.id
                  WHERE {where_clause}
                  ORDER BY a.transaction_date DESC'''
        
        records = all_(conn, sql, params or None)

        # Dashboard figures + per-record days-left-to-submit (pending records only)
        today = dubai_today()
        total_pending = 0
        overdue = 0
        due_soon = 0  # pending & due within 7 days (not yet overdue)
        for r in records:
            is_pending = (r.get('goaml_status') == 'pending')
            dl = days_left(r.get('due_date')) if r.get('due_date') else None
            r['days_to_submit'] = dl if is_pending else None
            if is_pending:
                total_pending += 1
                if dl is not None:
                    if dl < 0:
                        overdue += 1
                    elif dl <= 7:
                        due_soon += 1

        # Get dropdown options for filters
        companies = all_(conn, 'SELECT id, client_name as company_name FROM companies WHERE disabled IS NOT TRUE ORDER BY client_name')
        individuals = all_(conn, 'SELECT id, name FROM clients WHERE disabled IS NOT TRUE ORDER BY name')
        users = all_(conn, 'SELECT id, email FROM users WHERE is_active=1 ORDER BY email')
        statuses = ['pending', 'submitted', 'approved', 'rejected']

        conn.close()
        return render_template('aml_tracker.html',
                             records=records,
                             companies=companies,
                             users=users,
                             statuses=statuses,
                             total_pending=total_pending,
                             overdue_count=overdue,
                             due_soon_count=due_soon,
                             status_filter=status_filter,
                             company_filter=company_filter,
                             vc_no_filter=vc_no_filter,
                             submitted_by_filter=submitted_by_filter,
                             period_filter=period_filter,
                             date_from=date_from,
                             date_to=date_to)
    except Exception as e:
        logger.error(f'Error in aml_tracker: {e}')
        return render_template('aml_tracker.html', records=[], companies=[], 
                             users=[], statuses=[], error=str(e))

def _aml_resolve_client(conn, data):
    """Derive client_name + ac_type from the selected company/individual."""
    cid = data.get('company_id')
    iid = data.get('individual_id')
    if cid:
        co = one(conn, 'SELECT client_name FROM companies WHERE id=?', (cid,))
        return (co['client_name'] if co else data.get('client_name')), 'Company'
    if iid:
        ind = one(conn, 'SELECT name FROM clients WHERE id=?', (iid,))
        return (ind['name'] if ind else data.get('client_name')), 'Individual'
    return data.get('client_name'), data.get('ac_type')

def _aml_aed(data, exchange_rate):
    """Compute AED amount from foreign amount × rate; fall back to supplied aed_amount."""
    try:
        amt = float(data.get('usd_amount')) if data.get('usd_amount') not in (None, '') else None
        rate = float(exchange_rate) if exchange_rate not in (None, '') else None
        cur = (data.get('transaction_currency') or '').upper()
        if cur == 'AED' and amt is not None:
            return amt
        if amt is not None and rate is not None:
            return round(amt * rate, 2)
    except (TypeError, ValueError):
        pass
    return data.get('aed_amount') or None

@app.route('/aml-tracker/add', methods=['GET', 'POST'])
@compliance_required
def aml_tracker_add():
    """Add new AML Tracker record"""
    try:
        conn = get_db()
        companies = all_(conn, 'SELECT id, client_name as company_name FROM companies WHERE disabled IS NOT TRUE ORDER BY client_name')
        individuals = all_(conn, 'SELECT id, name FROM clients WHERE disabled IS NOT TRUE ORDER BY name')
        users = all_(conn, 'SELECT id, email FROM users WHERE is_active=1 ORDER BY email')
        
        if request.method == 'POST':
            data = request.get_json() if request.is_json else request.form
            
            transaction_date = data.get('transaction_date')
            if not transaction_date:
                conn.close()
                return jsonify({'success': False, 'error': 'Transaction date is required'}), 400
            
            # Auto-calculate PERIOD (reporting quarter) and DUE DATE (transaction + 15 days)
            tdate = datetime.strptime(transaction_date, '%Y-%m-%d').date()
            quarter_num = (tdate.month - 1) // 3 + 1
            period = data.get('period') or f'Q{quarter_num} {tdate.year}'
            due_date = tdate + timedelta(days=15)

            # Resolve client name + account type from the selected company / individual
            client_name, ac_type = _aml_resolve_client(conn, data)
            # Exchange rate → auto AED amount when not explicitly supplied
            exchange_rate = data.get('exchange_rate') or None
            aed_amount = _aml_aed(data, exchange_rate)

            # Insert AML record
            x(conn, '''INSERT INTO aml_tracker
               (company_id, individual_id, transaction_date, period, due_date, vc_no, payment_mode,
                ac_type, client_name, transaction_currency, usd_amount, aed_amount, exchange_rate, payment_remarks,
                invoice_no, invoice_amount, invoice_currency, goaml_submission_date, goaml_status,
                goaml_ref_no, submitted_by, checked_by, comment, verified_ledger, created_at, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)''',
              (data.get('company_id') or None,
               data.get('individual_id') or None,
               transaction_date,
               period,
               str(due_date),
               data.get('vc_no'),
               data.get('payment_mode'),
               ac_type,
               client_name,
               data.get('transaction_currency'),
               data.get('usd_amount') or None,
               aed_amount,
               exchange_rate,
               data.get('payment_remarks'),
               data.get('invoice_no'),
               data.get('invoice_amount') or None,
               data.get('invoice_currency'),
               data.get('goaml_submission_date') or None,
               'pending',
               None,
               session.get('user_id'),
               data.get('checked_by') or None,
               data.get('comment'),
               'true' if data.get('verified_ledger') else 'false'))

            aml_id = lastid(conn)
            
            # Create task for checked_by user if specified
            checked_by_id = data.get('checked_by')
            if checked_by_id:
                x(conn, '''INSERT INTO tasks 
                   (title, description, assigned_to, created_by, status, due_date)
                   VALUES (?, ?, ?, ?, ?, ?)''',
                  (f'AML Verification - {client_name}',
                   f'Review and verify AML Tracker record #{aml_id}. VC No: {data.get("vc_no")}',
                   int(checked_by_id),
                   session.get('user_id'),
                   'todo',
                   str(due_date)))

            commit(conn)
            conn.close()
            if request.is_json:
                return jsonify({'success': True, 'message': 'AML record created', 'id': aml_id})
            return redirect(url_for('aml_tracker'))
        
        conn.close()
        return render_template('aml_tracker_form.html', 
                             companies=companies, 
                             individuals=individuals,
                             users=users,
                             record=None)
    except Exception as e:
        logger.error(f'Error in aml_tracker_add: {e}')
        return _fail(e)

@app.route('/aml-tracker/<int:id>/edit', methods=['GET', 'POST'])
@compliance_required
def aml_tracker_edit(id):
    """Edit AML Tracker record"""
    try:
        conn = get_db()
        record = one(conn, 'SELECT * FROM aml_tracker WHERE id=?', (id,))
        
        if not record:
            conn.close()
            return redirect(url_for('aml_tracker'))
        
        companies = all_(conn, 'SELECT id, client_name as company_name FROM companies WHERE disabled IS NOT TRUE ORDER BY client_name')
        individuals = all_(conn, 'SELECT id, name FROM clients WHERE disabled IS NOT TRUE ORDER BY name')
        users = all_(conn, 'SELECT id, email FROM users WHERE is_active=1 ORDER BY email')
        
        # Once submitted, only an admin may save edits (View/GET stays allowed)
        locked = record.get('goaml_status') and record.get('goaml_status') != 'pending' and session.get('user_role') != 'admin'

        if request.method == 'POST':
            # Edit is admin-only and requires the action password
            if session.get('user_role') != 'admin':
                conn.close()
                return ("Only an admin can edit records.", 403)
            _pw = request.form.get('action_password') if request.form else (request.get_json(silent=True) or {}).get('action_password')
            if not _check_action_pw(_pw):
                conn.close()
                return ("Action password is required or incorrect.", 403)
            if locked:
                conn.close()
                if request.is_json:
                    return jsonify({'success': False, 'error': 'This report is submitted — only an admin can edit it.'}), 403
                return ("This report is submitted — only an admin can edit it.", 403)
            data = request.get_json() if request.is_json else request.form

            client_name, ac_type = _aml_resolve_client(conn, data)
            exchange_rate = data.get('exchange_rate') or None
            aed_amount = _aml_aed(data, exchange_rate)

            # Update record
            x(conn, '''UPDATE aml_tracker
               SET company_id=?, individual_id=?, vc_no=?, payment_mode=?, ac_type=?, client_name=?, transaction_currency=?,
                   usd_amount=?, aed_amount=?, exchange_rate=?, payment_remarks=?, invoice_no=?, invoice_amount=?,
                   invoice_currency=?, goaml_submission_date=?, goaml_status=?, goaml_ref_no=?,
                   checked_by=?, comment=?, verified_ledger=?, updated_at=CURRENT_TIMESTAMP
               WHERE id=?''',
              (data.get('company_id') or None,
               data.get('individual_id') or None,
               data.get('vc_no'),
               data.get('payment_mode'),
               ac_type,
               client_name,
               data.get('transaction_currency'),
               data.get('usd_amount') or None,
               aed_amount,
               exchange_rate,
               data.get('payment_remarks'),
               data.get('invoice_no'),
               data.get('invoice_amount') or None,
               data.get('invoice_currency'),
               data.get('goaml_submission_date') or None,
               data.get('goaml_status'),
               data.get('goaml_ref_no'),
               data.get('checked_by') or None,
               data.get('comment'),
               'true' if data.get('verified_ledger') else 'false',
               id))

            commit(conn)
            conn.close()
            if request.is_json:
                return jsonify({'success': True, 'message': 'AML record updated'})
            return redirect(url_for('aml_tracker'))
        
        # Format dates for template
        for field in ['transaction_date', 'due_date', 'goaml_submission_date']:
            if record.get(field):
                record[field] = str(record[field])[:10] if record[field] else None
        
        conn.close()
        return render_template('aml_tracker_form.html',
                             companies=companies,
                             individuals=individuals,
                             users=users,
                             record=record)
    except Exception as e:
        logger.error(f'Error in aml_tracker_edit: {e}')
        return _fail(e)

@app.route('/api/aml-tracker/<int:id>/delete', methods=['POST'])
@admin_pw_required
def api_aml_tracker_delete(id):
    """Delete AML Tracker record"""
    try:
        conn = get_db()
        rec = one(conn, 'SELECT goaml_status FROM aml_tracker WHERE id=?', (id,))
        # Once submitted, only an admin may delete
        if rec and rec.get('goaml_status') and rec.get('goaml_status') != 'pending' and session.get('user_role') != 'admin':
            conn.close()
            return jsonify({'success': False, 'error': 'This report is submitted — only an admin can delete it.'}), 403
        x(conn, 'DELETE FROM aml_tracker WHERE id=?', (id,))
        commit(conn)
        conn.close()
        return jsonify({'success': True, 'message': 'AML record deleted'})
    except Exception as e:
        logger.error(f'Error deleting AML record: {e}')
        return _fail(e)

# AML Tracker import column order (used by export, template, and import)
AML_IMPORT_COLS = ['transaction_date','company_name','individual_name','vc_no','payment_mode',
                   'transaction_currency','amount','exchange_rate','aed_amount','payment_remarks',
                   'invoice_no','invoice_amount','invoice_currency','goaml_status','goaml_ref_no','comment']

@app.route('/aml-tracker/export')
@require_perm('aml_export')
def aml_tracker_export():
    """Export all AML Tracker records to Excel."""
    if not HAS_XL:
        return "openpyxl not installed", 500
    conn = get_db()
    rows = all_(conn, '''SELECT a.transaction_date, c.client_name as company_name, cli.name as individual_name,
                  a.vc_no, a.payment_mode, a.transaction_currency, a.usd_amount as amount, a.exchange_rate,
                  a.aed_amount, a.payment_remarks, a.invoice_no, a.invoice_amount, a.invoice_currency,
                  a.goaml_status, a.goaml_ref_no, a.due_date, a.period, a.comment
                  FROM aml_tracker a
                  LEFT JOIN companies c ON a.company_id = c.id
                  LEFT JOIN clients cli ON a.individual_id = cli.id
                  ORDER BY a.transaction_date DESC''')
    conn.close()
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'AML Tracker'
    headers = ['Transaction Date','Company','Individual','VC No','Payment Mode','Currency','Amount',
               'Exchange Rate','Amount AED','Payment Remarks','Invoice No','Invoice Amount','Invoice Currency',
               'GoAML Status','GoAML Ref No','Due Date','Period','Comment']
    ws.append(headers)
    for r in rows:
        ws.append([str(r.get(k) if r.get(k) is not None else '') for k in
                   ['transaction_date','company_name','individual_name','vc_no','payment_mode','transaction_currency',
                    'amount','exchange_rate','aed_amount','payment_remarks','invoice_no','invoice_amount',
                    'invoice_currency','goaml_status','goaml_ref_no','due_date','period','comment']])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname = f'aml_tracker_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

@app.route('/aml-tracker/import-template')
@require_perm('aml_import')
def aml_tracker_import_template():
    """Download a blank Excel template for importing AML records."""
    if not HAS_XL:
        return "openpyxl not installed", 500
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = 'AML Import'
    headers = ['transaction_date (YYYY-MM-DD)','company_name','individual_name','vc_no','payment_mode',
               'transaction_currency','amount','exchange_rate','aed_amount (optional)','payment_remarks',
               'invoice_no','invoice_amount','invoice_currency','goaml_status','goaml_ref_no','comment']
    ws.append(headers)
    bold = Font(bold=True, color='FFFFFF'); fill = PatternFill('solid', fgColor='D97706')
    for cell in ws[1]:
        cell.font = bold; cell.fill = fill
    ws.append(['2026-06-25','ABC Trading LLC','','VC-1001','Cash','USD','10000','3.6725','','Sample row',
               'INV-001','10000','USD','pending','',''])
    out = io.BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='aml_import_template.xlsx')

@app.route('/aml-tracker/import', methods=['POST'])
@require_perm('aml_import')
def aml_tracker_import():
    """Import AML records from an uploaded Excel/CSV file."""
    f = request.files.get('file')
    if not f or not f.filename:
        return redirect(url_for('aml_tracker'))
    try:
        conn = get_db()
        # Build name → id lookups for matching company / individual
        co_map = {(r['client_name'] or '').strip().lower(): r['id']
                  for r in all_(conn, 'SELECT id, client_name FROM companies')}
        ind_map = {(r['name'] or '').strip().lower(): r['id']
                   for r in all_(conn, 'SELECT id, name FROM clients')}

        # Read rows as list of dicts keyed by AML_IMPORT_COLS
        records = []
        if f.filename.lower().endswith('.csv'):
            import csv as _csv
            text = f.read().decode('utf-8-sig').splitlines()
            reader = _csv.reader(text)
            next(reader, None)  # skip header
            for row in reader:
                if any(c.strip() for c in row):
                    records.append(dict(zip(AML_IMPORT_COLS, row + [''] * (len(AML_IMPORT_COLS) - len(row)))))
        else:
            if not HAS_XL:
                return "openpyxl not installed", 500
            wb = openpyxl.load_workbook(f, data_only=True)
            ws = wb.active
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:  # header
                    continue
                if row and any(v not in (None, '') for v in row):
                    vals = [('' if v is None else str(v)) for v in row]
                    records.append(dict(zip(AML_IMPORT_COLS, vals + [''] * (len(AML_IMPORT_COLS) - len(vals)))))

        imported = 0
        for rec in records:
            tdate = (rec.get('transaction_date') or '').strip()[:10]
            if not tdate:
                continue
            try:
                td = datetime.strptime(tdate, '%Y-%m-%d').date()
            except ValueError:
                continue
            period = f'Q{(td.month - 1)//3 + 1} {td.year}'
            due = td + timedelta(days=15)
            co_id = co_map.get((rec.get('company_name') or '').strip().lower())
            ind_id = ind_map.get((rec.get('individual_name') or '').strip().lower())
            ac_type = 'Company' if co_id else ('Individual' if ind_id else None)
            client_name = (rec.get('company_name') or rec.get('individual_name') or '').strip() or None
            exch = rec.get('exchange_rate') or None
            aed = _aml_aed({'usd_amount': rec.get('amount'), 'transaction_currency': rec.get('transaction_currency'),
                            'aed_amount': rec.get('aed_amount')}, exch)
            x(conn, '''INSERT INTO aml_tracker
               (company_id, individual_id, transaction_date, period, due_date, vc_no, payment_mode,
                ac_type, client_name, transaction_currency, usd_amount, aed_amount, exchange_rate, payment_remarks,
                invoice_no, invoice_amount, invoice_currency, goaml_status, goaml_ref_no, submitted_by, comment,
                created_at, updated_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,CURRENT_TIMESTAMP,CURRENT_TIMESTAMP)''',
              (co_id, ind_id, tdate, period, str(due), rec.get('vc_no') or None, rec.get('payment_mode') or None,
               ac_type, client_name, rec.get('transaction_currency') or None, rec.get('amount') or None, aed, exch,
               rec.get('payment_remarks') or None, rec.get('invoice_no') or None, rec.get('invoice_amount') or None,
               rec.get('invoice_currency') or None, (rec.get('goaml_status') or 'pending').lower(),
               rec.get('goaml_ref_no') or None, session.get('user_id'), rec.get('comment') or None))
            imported += 1
        commit(conn); conn.close()
        logger.info(f'AML import: {imported} records imported by user {session.get("user_id")}')
        return redirect(url_for('aml_tracker'))
    except Exception as e:
        logger.error(f'Error importing AML records: {e}')
        return render_template('aml_tracker.html', records=[], companies=[], users=[], statuses=[],
                               error=f'Import failed: {e}')

@app.route('/login-history')
@admin_required
def login_history():
    """Admin view of login attempts (retained for 90 days) for suspicious-login monitoring."""
    conn = get_db()
    cutoff = dubai_today() - timedelta(days=90)
    # Retention: purge anything older than 90 days on each visit
    try:
        x(conn, 'DELETE FROM login_history WHERE created_at < ?', (cutoff,))
        commit(conn)
    except Exception as e:
        logger.warning(f'login_history purge skipped: {e}')
    show = request.args.get('show', '')  # '', 'failed', 'success'
    conds = ['lh.created_at >= ?']; params = [str(cutoff)]
    if show == 'failed': conds.append('lh.success IS NOT TRUE')
    elif show == 'success': conds.append('lh.success IS TRUE')
    where = 'WHERE ' + ' AND '.join(conds)
    rows = all_(conn, f'''SELECT lh.*, u.name as user_name, u.role
                  FROM login_history lh LEFT JOIN users u ON lh.user_id = u.id
                  {where}
                  ORDER BY lh.created_at DESC LIMIT 1000''', params)
    failed_24h = cnt(conn, "SELECT COUNT(*) FROM login_history WHERE success IS NOT TRUE AND created_at >= ?",
                     (dubai_today() - timedelta(days=1),))
    conn.close()
    return render_template('login_history.html', rows=rows, show=show, failed_24h=failed_24h)

@app.route('/settings/country-scores')
@admin_required
def country_scores():
    """Admin editor for risk-assessment country scores (change periodically)."""
    conn = get_db()
    rows = all_(conn, 'SELECT country, score FROM risk_country_scores ORDER BY country')
    if not rows:
        # Fallback to in-memory defaults if table somehow empty
        rows = [{'country': k, 'score': v} for k, v in sorted(RISK_LOOKUPS['countries'].items())]
    conn.close()
    return render_template('country_scores.html', countries=rows)

@app.route('/api/risk-country-scores', methods=['POST'])
@admin_required
def api_save_country_scores():
    """Save edited country scores and refresh the in-memory lookup."""
    d = request.get_json() or {}
    scores = d.get('scores', {})  # { 'COUNTRY NAME': 1|2|3 }
    if not scores:
        return jsonify({'success': False, 'error': 'No scores provided'}), 400
    try:
        conn = get_db()
        for country, score in scores.items():
            try:
                sc = int(score)
            except (TypeError, ValueError):
                continue
            if sc not in (1, 2, 3):
                continue
            if use_pg():
                x(conn, '''INSERT INTO risk_country_scores (country, score, updated_at)
                           VALUES (%s,%s,CURRENT_TIMESTAMP)
                           ON CONFLICT (country) DO UPDATE SET score=EXCLUDED.score, updated_at=CURRENT_TIMESTAMP''',
                  (country, sc))
            else:
                conn.execute('''INSERT INTO risk_country_scores (country, score, updated_at)
                                VALUES (?,?,CURRENT_TIMESTAMP)
                                ON CONFLICT(country) DO UPDATE SET score=excluded.score, updated_at=CURRENT_TIMESTAMP''',
                             (country, sc))
        commit(conn)
        refresh_country_scores(conn)
        conn.close()
        return jsonify({'success': True, 'updated': len(scores)})
    except Exception as e:
        logger.error(f'Error saving country scores: {e}')
        return _fail(e)

def _load_risk_questions(conn, applies_to=None):
    """Return questions (optionally for a given entity type) with nested answer options."""
    q_sql = "SELECT * FROM risk_questions WHERE is_active=1"
    params = []
    if applies_to:
        q_sql += " AND (applies_to=? OR applies_to='both')"; params.append(applies_to)
    q_sql += " ORDER BY sort_order, id"
    questions = all_(conn, q_sql, params or None)
    if not questions:
        return []
    qids = [q['id'] for q in questions]
    ph = ','.join([P()]*len(qids))
    opts = all_(conn, f"SELECT * FROM risk_answer_options WHERE question_id IN ({ph}) ORDER BY sort_order, id", qids)
    by_q = {}
    for o in opts:
        by_q.setdefault(o['question_id'], []).append({'id': o['id'], 'label': o['label'], 'score': o['score']})
    for q in questions:
        q['options'] = by_q.get(q['id'], [])
    return questions

@app.route('/settings/risk-questions')
@admin_required
def risk_questions_admin():
    """Admin builder for the risk-assessment questionnaire."""
    conn = get_db()
    company_qs = _load_risk_questions(conn, 'company')
    individual_qs = _load_risk_questions(conn, 'individual')
    conn.close()
    return render_template('risk_questions.html', company_qs=company_qs, individual_qs=individual_qs)

@app.route('/api/risk-questions/load-defaults', methods=['POST'])
@admin_required
def api_load_default_risk_questions():
    """Recovery action: insert the built-in default questions for one entity type,
    but ONLY if that type currently has none — never duplicates and never
    touches questions an admin has already added/edited."""
    d = request.get_json() or {}
    applies_to = d.get('applies_to')
    if applies_to not in DEFAULT_RISK_QUESTIONS:
        return jsonify({'success': False, 'error': 'Invalid entity type'}), 400
    try:
        conn = get_db()
        existing = cnt(conn, "SELECT COUNT(*) FROM risk_questions WHERE applies_to=?", (applies_to,))
        if existing and int(existing) > 0:
            conn.close()
            return jsonify({'success': False, 'error': 'Questions already exist for this type.'}), 400
        inserted = _insert_question_set(conn, applies_to, DEFAULT_RISK_QUESTIONS[applies_to])
        conn.close()
        return jsonify({'success': True, 'inserted': inserted})
    except Exception as e:
        return _fail(e)

@app.route('/api/risk-questions/save', methods=['POST'])
@admin_required
def api_save_risk_questions():
    """Replace the whole questionnaire for one entity type from posted JSON.
    Payload: {applies_to:'company'|'individual', questions:[{question, options:[{label,score}]}]}"""
    d = request.get_json() or {}
    applies_to = d.get('applies_to')
    questions = d.get('questions', [])
    if applies_to not in ('company', 'individual'):
        return jsonify({'success': False, 'error': 'Invalid entity type'}), 400
    try:
        conn = get_db()
        # Wipe existing questions (+ their options) for this entity type, then re-insert
        old = all_(conn, "SELECT id FROM risk_questions WHERE applies_to=?", (applies_to,))
        for o in old:
            x(conn, "DELETE FROM risk_answer_options WHERE question_id=?", (o['id'],))
        x(conn, "DELETE FROM risk_questions WHERE applies_to=?", (applies_to,))
        for order, q in enumerate(questions):
            qtext = (q.get('question') or '').strip()
            if not qtext:
                continue
            x(conn, "INSERT INTO risk_questions (applies_to,question,sort_order,is_active) VALUES (?,?,?,1)",
              (applies_to, qtext, order))
            qid = lastid(conn)
            for aorder, opt in enumerate(q.get('options', [])):
                label = (opt.get('label') or '').strip()
                if not label:
                    continue
                try: score = int(opt.get('score', 1))
                except (TypeError, ValueError): score = 1
                x(conn, "INSERT INTO risk_answer_options (question_id,label,score,sort_order) VALUES (?,?,?,?)",
                  (qid, label, score, aorder))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f'Error saving risk questions: {e}')
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/settings')
@admin_required
def settings():
    conn=get_db()
    users=all_(conn,'SELECT id,email,name,role,is_active,contact_number,username,permissions FROM users ORDER BY role,name')
    dds=all_(conn,'SELECT * FROM dropdowns WHERE is_active=1 ORDER BY field_name,value')
    docs_path_row = one(conn, "SELECT value FROM app_settings WHERE key='local_docs_path'")
    conn.close()
    groups={}
    for d in dds: groups.setdefault(d['field_name'],[]).append(d)
    return render_template('settings.html',users=users,dropdown_groups=groups,
        local_docs_path=(docs_path_row['value'] if docs_path_row else ''))

def get_local_docs_path():
    """Admin-configured folder to also save a local copy of documents (on-prem installs only)."""
    try:
        conn = get_db()
        row = one(conn, "SELECT value FROM app_settings WHERE key='local_docs_path'")
        conn.close()
        return (row['value'] if row else '').strip()
    except Exception:
        return ''

def save_local_copy(file_bytes, subfolder, filename):
    """Best-effort local copy of an uploaded document, alongside the Cloudinary copy.
    Only runs if an admin has configured a local path. Never raises — a missing/
    unwritable path (e.g. on a cloud host with no persistent disk) is just skipped."""
    base = get_local_docs_path()
    if not base:
        return
    try:
        safe_sub = re.sub(r'[^A-Za-z0-9_\-]', '_', str(subfolder))
        safe_name = re.sub(r'[^A-Za-z0-9_\-.]', '_', str(filename))
        folder = os.path.join(base, safe_sub)
        os.makedirs(folder, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        with open(os.path.join(folder, f'{ts}_{safe_name}'), 'wb') as f:
            f.write(file_bytes)
    except Exception as e:
        logger.warning(f'Local document copy skipped ({subfolder}/{filename}): {e}')

@app.route('/api/settings/local-docs-path', methods=['POST'])
@admin_required
def api_set_local_docs_path():
    d = request.get_json() or {}
    path = (d.get('path') or '').strip()
    try:
        conn = get_db()
        if use_pg():
            x(conn, "INSERT INTO app_settings (key,value) VALUES ('local_docs_path',%s) ON CONFLICT(key) DO UPDATE SET value=%s,updated_at=CURRENT_TIMESTAMP", (path, path))
        else:
            conn.execute("INSERT INTO app_settings (key,value) VALUES ('local_docs_path',?) ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=CURRENT_TIMESTAMP", (path,))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

@app.route('/api/settings/browse-folder', methods=['POST'])
@admin_required
def api_browse_folder():
    """Open a native folder picker on the SERVER machine and return the chosen path.
    Works on the local/on-prem Windows install (where server = the user's PC).
    On a headless/cloud host there is no desktop, so it fails gracefully and the
    admin types the path instead."""
    try:
        import subprocess
        ps = ("Add-Type -AssemblyName System.Windows.Forms | Out-Null;"
              "$f = New-Object System.Windows.Forms.FolderBrowserDialog;"
              "$f.Description = 'Select folder for Zewer CRM storage';"
              "$f.ShowNewFolderButton = $true;"
              "if ($f.ShowDialog() -eq [System.Windows.Forms.DialogResult]::OK) { [Console]::Out.Write($f.SelectedPath) }")
        result = subprocess.run(['powershell', '-NoProfile', '-Sta', '-Command', ps],
                                capture_output=True, text=True, timeout=180)
        path = (result.stdout or '').strip()
        if path:
            return jsonify({'success': True, 'path': path})
        return jsonify({'success': False, 'error': 'No folder selected.'})
    except Exception as e:
        return jsonify({'success': False, 'error': f'Folder picker not available on this server — type the path manually. ({e})'})

@app.route('/api/user/add',methods=['POST'])
@admin_required
def api_add_user():
    d=request.get_json()
    if len((d.get('password') or '')) < 8:
        return jsonify({'success':False,'error':'Password must be at least 8 characters'}),400
    try:
        conn=get_db()
        username = d.get('username','').strip().lower()
        if username:
            existing = one(conn,'SELECT id FROM users WHERE username=?',(username,))
            if existing: return jsonify({'success':False,'error':'Username already taken'}),400
        x(conn,'INSERT INTO users (email,password_hash,name,role,contact_number,username,permissions,is_active) VALUES (?,?,?,?,?,?,?,1)',
          (d.get('email',''),generate_password_hash(d.get('password','')),d.get('name'),d.get('role','staff'),d.get('contact_number'),username or None,d.get('permissions','')))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/user/<int:id>/edit',methods=['POST'])
@admin_required
def api_edit_user(id):
    d=request.get_json()
    try:
        conn=get_db()
        username = d.get('username','').strip().lower() or None
        if username:
            existing = one(conn,'SELECT id FROM users WHERE username=? AND id!=?',(username,id))
            if existing: return jsonify({'success':False,'error':'Username already taken'}),400
        if d.get('password'):
            if len(d.get('password')) < 8:
                return jsonify({'success':False,'error':'Password must be at least 8 characters'}),400
            x(conn,'UPDATE users SET name=?,email=?,role=?,contact_number=?,username=?,permissions=?,password_hash=? WHERE id=?',
              (d.get('name'),d.get('email'),d.get('role'),d.get('contact_number'),username,d.get('permissions',''),generate_password_hash(d.get('password')),id))
        else:
            x(conn,'UPDATE users SET name=?,email=?,role=?,contact_number=?,username=?,permissions=? WHERE id=?',
              (d.get('name'),d.get('email'),d.get('role'),d.get('contact_number'),username,d.get('permissions',''),id))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/user/<int:id>/toggle',methods=['POST'])
@admin_required
def api_toggle_user(id):
    try:
        conn=get_db(); u=one(conn,'SELECT is_active FROM users WHERE id=?',(id,))
        if u: x(conn,'UPDATE users SET is_active=? WHERE id=?',(0 if u['is_active'] else 1,id))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/user/<int:id>/delete',methods=['POST'])
@admin_required
def api_delete_user(id):
    try:
        conn=get_db(); x(conn,'DELETE FROM users WHERE id=?',(id,))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/dropdown/add',methods=['POST'])
@admin_required
def api_add_dropdown():
    d=request.get_json()
    try:
        conn=get_db()
        x(conn,'INSERT INTO dropdowns (field_name,value,is_active) VALUES (?,?,1)',(d.get('field_name'),d.get('value')))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/dropdown/<int:id>/delete',methods=['POST'])
@admin_required
def api_delete_dropdown(id):
    try:
        conn=get_db(); x(conn,'DELETE FROM dropdowns WHERE id=?',(id,))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/tasks')
@login_required
def tasks():
    conn=get_db(); uid=session.get('user_id'); role=session.get('user_role')
    today=dubai_today()
    active_tab=request.args.get('tab','temp')

    # ── TEMP TASKS ──
    base='''SELECT t.*,u.name as assigned_name,u.mobile as assigned_mobile,
        cu.name as created_by_name,c.client_name as company_name,c.ac_code
        FROM tasks t LEFT JOIN users u ON t.assigned_to=u.id
        LEFT JOIN users cu ON t.created_by=cu.id
        LEFT JOIN companies c ON t.company_id=c.id'''
    order=" ORDER BY CASE t.priority WHEN 'urgent' THEN 1 WHEN 'high' THEN 2 WHEN 'normal' THEN 3 ELSE 4 END,t.due_date"
    if role == 'admin':
        rows=all_(conn,base+order)
    else:
        rows=all_(conn,base+" WHERE (t.assigned_to=? OR t.created_by=?)"+order,(uid,uid))
    users=all_(conn,'SELECT id,name,role,mobile FROM users WHERE is_active=1 ORDER BY name')
    cos=all_(conn,'SELECT id,ac_code,client_name FROM companies ORDER BY client_name')
    tmpls=all_(conn,"SELECT value,description FROM dropdowns WHERE field_name='TASK TEMPLATE' AND is_active=1 ORDER BY value")
    tl=[]
    for t in rows:
        d=days_left(t['due_date'])
        tl.append({**t,'priority':t['priority'] or 'normal','status':t['status'] or 'todo',
                   'due_date':str(t['due_date']) if t['due_date'] else None,
                   'days_until_due':d,'is_overdue':(d is not None and d<0)})

    # ── REGULAR TASKS ──
    def due_dates_for_frequency(freq, since_date):
        dates=[]
        if freq=='daily':
            d=since_date
            while d<=today:
                dates.append(d); d+=timedelta(days=1)
        elif freq=='weekly':
            d=since_date
            days_ahead=(4-d.weekday())%7
            d=d+timedelta(days=days_ahead)
            while d<=today:
                dates.append(d); d+=timedelta(weeks=1)
        elif freq=='monthly':
            d=since_date.replace(day=1)
            while True:
                try: due=d.replace(day=25)
                except ValueError: due=None
                if due and due>=since_date and due<=today: dates.append(due)
                if d.month==12: d=d.replace(year=d.year+1,month=1)
                else: d=d.replace(month=d.month+1)
                if d>today: break
        return dates

    try:
        if role in ['admin','compliance']:
            rt_templates=all_(conn,"""SELECT rt.*,u.name as created_by_name,au.name as assigned_user_name
                FROM regular_task_templates rt LEFT JOIN users u ON rt.created_by=u.id
                LEFT JOIN users au ON rt.assigned_user_id=au.id
                ORDER BY rt.frequency,rt.title""")
            rt_logs=all_(conn,"""SELECT l.*,u.name as staff_name,rt.title as task_title,rt.frequency
                FROM regular_task_logs l JOIN users u ON l.user_id=u.id
                JOIN regular_task_templates rt ON l.template_id=rt.id
                ORDER BY l.logged_at DESC LIMIT 200""")
        else:
            rt_templates=all_(conn,"""SELECT rt.*,u.name as created_by_name,au.name as assigned_user_name
                FROM regular_task_templates rt LEFT JOIN users u ON rt.created_by=u.id
                LEFT JOIN users au ON rt.assigned_user_id=au.id
                WHERE rt.assigned_role='all' OR rt.assigned_user_id=? OR rt.assigned_role=?
                ORDER BY rt.frequency,rt.title""", (uid,role))
            rt_logs=all_(conn,"""SELECT l.*,u.name as staff_name,rt.title as task_title,rt.frequency
                FROM regular_task_logs l JOIN users u ON l.user_id=u.id
                JOIN regular_task_templates rt ON l.template_id=rt.id
                WHERE l.user_id=? ORDER BY l.logged_at DESC LIMIT 100""", (uid,))
    except: rt_templates=[]; rt_logs=[]

    task_status={}
    for t in rt_templates:
        try:
            user_logs=all_(conn,"""SELECT DATE(logged_at) as log_date FROM regular_task_logs
                WHERE template_id=? AND user_id=? ORDER BY logged_at ASC""", (t['id'],uid))
            logged_dates=set(r['log_date'] for r in user_logs)
            created=t.get('created_at')
            try: start=datetime.strptime(str(created)[:10],'%Y-%m-%d').date()
            except: start=today
            all_due=due_dates_for_frequency(t['frequency'],start)
            missed=[d for d in all_due if str(d) not in logged_dates]
            freq=t['frequency']
            if freq=='daily': next_due=today+timedelta(days=1)
            elif freq=='weekly':
                days_ahead=(4-today.weekday())%7
                next_due=today+timedelta(days=(days_ahead if days_ahead>0 else 7))
            elif freq=='monthly':
                if today.day<25: next_due=today.replace(day=25)
                elif today.month==12: next_due=today.replace(year=today.year+1,month=1,day=25)
                else: next_due=today.replace(month=today.month+1,day=25)
            else: next_due=today
            is_due_today=str(today) in [str(d) for d in all_due] and str(today) not in logged_dates
            task_status[t['id']]={
                'overdue_count':len(missed),'is_due_today':is_due_today,
                'next_due':str(next_due),'last_logged':user_logs[-1]['log_date'] if user_logs else None,
                'missed_dates':[str(d) for d in missed[-3:]]
            }
        except: task_status[t['id']]={'overdue_count':0,'is_due_today':False,'next_due':str(today),'last_logged':None,'missed_dates':[]}

    # ── ADDITIONAL TASKS ──
    try:
        # Ensure table exists on Railway PG
        if is_pg(conn):
            x(conn, '''CREATE TABLE IF NOT EXISTS additional_tasks (
                id SERIAL PRIMARY KEY, title TEXT NOT NULL,
                task_details TEXT, remarks TEXT,
                from_datetime TIMESTAMP NOT NULL, to_datetime TIMESTAMP NOT NULL,
                created_by INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        else:
            x(conn, '''CREATE TABLE IF NOT EXISTS additional_tasks (
                id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL,
                task_details TEXT, remarks TEXT,
                from_datetime TIMESTAMP NOT NULL, to_datetime TIMESTAMP NOT NULL,
                created_by INTEGER NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        commit(conn)
    except: pass
    try:
        if role == 'admin':
            add_tasks = all_(conn, '''SELECT at.*,u.name as staff_name
                FROM additional_tasks at LEFT JOIN users u ON at.created_by=u.id
                ORDER BY at.from_datetime DESC''')
        else:
            add_tasks = all_(conn, '''SELECT at.*,u.name as staff_name
                FROM additional_tasks at LEFT JOIN users u ON at.created_by=u.id
                WHERE at.created_by=?
                ORDER BY at.from_datetime DESC''', (uid,))
        add_tasks = [{**t,
            'from_datetime': str(t['from_datetime'])[:16] if t['from_datetime'] else '',
            'to_datetime': str(t['to_datetime'])[:16] if t['to_datetime'] else '',
            'created_at': str(t['created_at'])[:10] if t['created_at'] else ''
        } for t in add_tasks]
    except: add_tasks = []

    conn.close()
    return render_template('tasks.html',tasks=tl,all_users=users,all_companies=cos,task_templates=tmpls,
                           rt_templates=rt_templates,rt_logs=rt_logs,task_status=task_status,
                           active_tab=active_tab,today=str(today),add_tasks=add_tasks,
                           current_user_id=uid)

@app.route('/api/task/add',methods=['POST'])
@login_required
def api_add_task():
    d=request.get_json()
    try:
        conn=get_db()
        x(conn,'INSERT INTO tasks (title,description,assigned_to,created_by,company_id,priority,due_date,status) VALUES (?,?,?,?,?,?,?,?)',
          (d.get('title'),d.get('description'),d.get('assigned_to') or None,session.get('user_id'),
           d.get('company_id') or None,d.get('priority','normal'),d.get('due_date') or None,'todo'))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/task/<int:id>/edit',methods=['POST'])
@login_required
def api_edit_task(id):
    d=request.get_json()
    try:
        conn=get_db()
        x(conn,'UPDATE tasks SET title=?,description=?,assigned_to=?,company_id=?,priority=?,due_date=?,updated_at=CURRENT_TIMESTAMP WHERE id=?',
          (d.get('title'),d.get('description'),d.get('assigned_to') or None,d.get('company_id') or None,
           d.get('priority','normal'),d.get('due_date') or None,id))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/task/<int:id>/status',methods=['POST'])
@login_required
def api_task_status(id):
    d=request.get_json()
    try:
        conn=get_db()
        x(conn,'UPDATE tasks SET status=?,updated_at=CURRENT_TIMESTAMP WHERE id=?',(d.get('status'),id))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/task/<int:id>/delete',methods=['POST'])
@login_required
def api_delete_task(id):
    if session.get('user_role') == 'staff':
        return jsonify({'success':False,'error':'Staff cannot delete tasks'}),403
    try:
        conn=get_db(); x(conn,'DELETE FROM tasks WHERE id=?',(id,))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/company/<int:cid>/documents')
@compliance_required
def api_get_documents(cid):
    conn=get_db()
    docs=all_(conn,'''SELECT d.*,u.name as uploader_name FROM documents d
        LEFT JOIN users u ON d.uploaded_by=u.id WHERE d.company_id=? ORDER BY d.created_at DESC''',(cid,))
    conn.close(); return jsonify(docs)

@app.route('/api/company/<int:cid>/upload',methods=['POST'])
@compliance_required
def api_upload_document(cid):
    if 'file' not in request.files: return jsonify({'success':False,'error':'No file'}),400
    file=request.files['file']
    if not file.filename: return jsonify({'success':False,'error':'No filename'}),400
    if not HAS_CLD or not os.getenv('CLOUDINARY_CLOUD_NAME'):
        return jsonify({'success':False,'error':'Cloudinary not configured'}),500
    try:
        file_bytes = file.read(); file.seek(0)
        r=cloudinary.uploader.upload(file,folder=f'zewer_crm/co_{cid}',resource_type='auto',use_filename=True,unique_filename=True)
        conn=get_db()
        x(conn,'INSERT INTO documents (company_id,doc_type,file_name,file_url,public_id,uploaded_by,notes) VALUES (?,?,?,?,?,?,?)',
          (cid,request.form.get('doc_type','General'),file.filename,r['secure_url'],r['public_id'],session.get('user_id'),request.form.get('notes','')))
        commit(conn); conn.close()
        save_local_copy(file_bytes, f'companies/co_{cid}', file.filename)
        return jsonify({'success':True,'url':r['secure_url'],'name':file.filename})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/document/<int:did>/delete',methods=['POST'])
@compliance_required
def api_delete_document(did):
    try:
        conn=get_db(); doc=one(conn,'SELECT public_id FROM documents WHERE id=?',(did,))
        if doc and doc.get('public_id') and HAS_CLD and os.getenv('CLOUDINARY_CLOUD_NAME'):
            try: cloudinary.uploader.destroy(doc['public_id'],resource_type='raw')
            except: pass
        x(conn,'DELETE FROM documents WHERE id=?',(did,))
        commit(conn); conn.close(); return jsonify({'success':True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/export/template')
@require_perm('companies_export')
def export_template():
    if not HAS_XL: return "openpyxl not installed",500
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.Workbook()

    # ── SHEET 1: DATA ENTRY ──────────────────────────────────
    ws = wb.active
    ws.title = "Companies"

    headers = ['ac_code','client_name','ac_opening_date','ac_status','nature','type_of_client',
               'name_of_freezone','mode_of_ac','country_of_incorporation','region','address',
               'telephone','mobile','whatsapp_number','email_id','contact_person_name',
               'contact_person_number','account_manager','address_proof_type','address_proof_expiry',
               'trade_license_no','issuing_authority','legal_type','incorporation_date',
               'trade_license_expiry','tax_no_trn','vat_cert','vat_declaration','num_beneficial_owners',
               'moa','pep','undertaking','source_of_fund','doc_status','risk_status','kyc_status',
               'verified_by','followup_details','zewer_comments']

    # Style header row
    header_fill = PatternFill(start_color="1C1917", end_color="1C1917", fill_type="solid")
    header_font = Font(color="D97706", bold=True, size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h.replace('_',' ').title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.row_dimensions[1].height = 30

    # Sample row
    sample = ['TJ5003','SAMPLE COMPANY LLC','2020-01-01','Active','Legal entity','MainLand','N/A',
              'Supplier','United Arab Emirates','Dubai','Unit 101, Gold Souq, Dubai','+97142258019',
              '+971506594165','+971506594165','info@sample.com','Ahmed Ali','+971501234567','Jaseel',
              'Ejari','2025-12-31','534230','Dubai Economy & Tourism','Limited Liability Company(LLC)',
              '2019-06-01','2025-12-31','100003063300003','Yes','Yes',2,'Yes','Yes','Yes','Yes',
              'Completed','Medium','Kyc 2025 Updated','Jaseel','','']
    ws.append(sample)

    # Style sample row
    sample_fill = PatternFill(start_color="292524", end_color="292524", fill_type="solid")
    sample_font = Font(color="A8A29E", size=10, italic=True)
    for col in range(1, len(headers)+1):
        cell = ws.cell(row=2, column=col)
        cell.fill = sample_fill
        cell.font = sample_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = {'ac_code':12,'client_name':30,'ac_opening_date':16,'ac_status':12,'nature':16,
                  'type_of_client':18,'name_of_freezone':20,'mode_of_ac':16,'country_of_incorporation':22,
                  'region':16,'address':30,'telephone':18,'mobile':18,'whatsapp_number':18,
                  'email_id':25,'contact_person_name':22,'contact_person_number':20,'account_manager':18,
                  'address_proof_type':22,'address_proof_expiry':18,'trade_license_no':18,
                  'issuing_authority':28,'legal_type':30,'incorporation_date':18,'trade_license_expiry':18,
                  'tax_no_trn':16,'vat_cert':12,'vat_declaration':16,'num_beneficial_owners':12,
                  'moa':8,'pep':8,'undertaking':14,'source_of_fund':16,'doc_status':14,
                  'risk_status':12,'kyc_status':22,'verified_by':16,'followup_details':30,'zewer_comments':30}
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 16)

    # Freeze top 2 rows (header + sample)
    ws.freeze_panes = 'A3'

    # ── CELL DROPDOWN VALIDATION ──────────────────────────────
    conn = get_db()
    dd_rows = all_(conn, "SELECT field_name, value FROM dropdowns WHERE is_active=1 ORDER BY field_name, value")
    conn.close()

    dd = {}
    for r in dd_rows:
        dd.setdefault(r['field_name'], []).append(r['value'])

    # Map column field names to DB dropdown field_names
    dropdown_map = {
        'ac_status': 'AC STATUS',
        'nature': 'NATURE',
        'type_of_client': 'TYPE OF CLIENT',
        'name_of_freezone': 'NAME OF FREEZONE',
        'mode_of_ac': 'MODE OF AC',
        'country_of_incorporation': 'COUNTRY',
        'region': 'REGION',
        'address_proof_type': 'ADDRESS PROOF TYPE',
        'account_manager': 'ACCOUNT MANAGER',
        'issuing_authority': 'ISSUING AUTHORITY',
        'legal_type': 'LEGAL TYPE',
        'vat_cert': 'VAT CERT',
        'vat_declaration': 'VAT DECLARATION',
        'moa': 'MOA',
        'pep': 'PEP',
        'undertaking': 'UNDERTAKING',
        'source_of_fund': 'SOURCE OF FUND',
        'doc_status': 'DOC STATUS',
        'risk_status': 'RISK STATUS',
        'kyc_status': 'KYC STATUS',
    }

    for col_idx, h in enumerate(headers, 1):
        db_field = dropdown_map.get(h)
        if not db_field:
            continue
        values = dd.get(db_field, [])
        if not values:
            continue
        joined = ','.join(values)
        if len(joined) > 250:
            joined = joined[:250].rsplit(',', 1)[0]  # trim to fit Excel limit
        formula = '\"' + joined.replace(',', '\",\"') + '\"'
        dv = DataValidation(
            type="list",
            formula1='"' + joined + '"',
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="Invalid Value",
            error="Please select a value from the dropdown list."
        )
        col_letter = get_column_letter(col_idx)
        dv.sqref = f"{col_letter}3:{col_letter}1000"
        ws.add_data_validation(dv)

    # ── SHEET 2: INSTRUCTIONS ───────────────────────────────
    ws3 = wb.create_sheet("Instructions")
    instructions = [
        ["ZEWER AML CRM — Company Import Template"],
        [""],
        ["HOW TO USE:"],
        ["1. Fill in company data in the 'Companies' sheet starting from Row 3"],
        ["2. Row 2 is a sample — you can delete it before importing"],
        ["3. AC Code and Client Name are REQUIRED — all others are optional"],
        ["4. For dropdown fields, click the cell — a dropdown arrow will appear to select valid values"],
        ["5. Dates must be in YYYY-MM-DD format (e.g. 2025-12-31)"],
        ["6. Phone numbers should include country code (e.g. +97142258019)"],
        ["7. Duplicate AC Codes will be skipped on import"],
        [""],
        ["REQUIRED FIELDS:"],
        ["  • ac_code — Unique account code (e.g. TJ5003)"],
        ["  • client_name — Full legal company name"],
        [""],
        ["DATE FORMAT:"],
        ["  • ac_opening_date, address_proof_expiry, incorporation_date,"],
        ["    trade_license_expiry — all use YYYY-MM-DD"],
    ]
    for r, row_data in enumerate(instructions, 1):
        cell = ws3.cell(row=r, column=1, value=row_data[0] if row_data else '')
        if r == 1:
            cell.font = Font(bold=True, size=14, color="D97706")
        elif row_data and row_data[0].startswith(('HOW','REQUIRED','DATE')):
            cell.font = Font(bold=True, size=11, color="F5F5F4")
        else:
            cell.font = Font(size=10, color="A8A29E")
    ws3.column_dimensions['A'].width = 70
    ws3.sheet_view.showGridLines = False

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='zewer_company_template.xlsx')

@app.route('/export/companies')
@require_perm('companies_export')
def export_companies():
    scope=request.args.get('scope','all'); fmt=request.args.get('format','csv')
    conn=get_db()
    q='SELECT * FROM companies WHERE 1=1'
    if scope=='active': q+=' AND ac_status=\'Active\''
    elif scope=='inactive': q+=' AND ac_status=\'Inactive\''
    elif scope=='high': q+=' AND risk_status=\'High\''
    elif scope=='medium': q+=' AND risk_status=\'Medium\''
    elif scope=='low': q+=' AND risk_status=\'Low\''
    elif scope=='incomplete': q+=' AND doc_status=\'Incompleted\''
    rows=all_(conn,q)

    # Pull UBO/Authorized Person records for these companies too
    company_ids = [r['id'] for r in rows]
    ubo_rows = []
    ubos_by_company = {}
    if company_ids:
        placeholders = ','.join([P()]*len(company_ids))
        ubo_rows = all_(conn, f'SELECT * FROM ubos WHERE company_id IN ({placeholders}) ORDER BY company_id, share_percentage DESC', company_ids)
        for u in ubo_rows:
            ubos_by_company.setdefault(u['company_id'], []).append(u)
    conn.close()

    fname=f'companies_{scope}_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
    if fmt=='xlsx' and HAS_XL:
        wb=openpyxl.Workbook()
        ws=wb.active; ws.title='Companies'
        if rows: ws.append(list(rows[0].keys())); [ws.append([str(v) if v else '' for v in r.values()]) for r in rows]
        # Second sheet: every UBO / Authorized Person, tagged with their company
        ws2 = wb.create_sheet('UBOs')
        ubo_headers = ['company_id','ac_code','client_name','position','share_percentage','person_name',
                       'nationality','residential_status','pep_status','passport_no','passport_expiry',
                       'emirates_id','emirates_id_expiry','doc_status','verified_by','followup_details']
        ws2.append(ubo_headers)
        co_lookup = {r['id']: r for r in rows}
        for u in ubo_rows:
            co = co_lookup.get(u['company_id'], {})
            ws2.append([str(u.get('company_id') or ''), co.get('ac_code',''), co.get('client_name',''),
                       u.get('position') or '', u.get('share_percentage') or '', u.get('person_name') or '',
                       u.get('nationality') or '', u.get('residential_status') or '', u.get('pep_status') or '',
                       u.get('passport_no') or '', str(u.get('passport_expiry') or ''),
                       u.get('emirates_id') or '', str(u.get('emirates_id_expiry') or ''),
                       u.get('doc_status') or '', u.get('verified_by') or '', u.get('followup_details') or ''])
        out=io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,download_name=fname+'.xlsx')

    # CSV can't hold two sheets — append a compact UBO summary column instead
    out=io.StringIO(); w=csv.writer(out)
    if rows:
        headers = list(rows[0].keys()) + ['ubos_summary']
        w.writerow(headers)
        for r in rows:
            ubo_list = ubos_by_company.get(r['id'], [])
            summary = ' | '.join(
                f"{u.get('person_name','')} ({u.get('position','')}, {u.get('share_percentage','') or 0}%, "
                f"PEP:{u.get('pep_status') or '-'}, Passport:{u.get('passport_no') or '-'}, EID:{u.get('emirates_id') or '-'})"
                for u in ubo_list)
            w.writerow(list(r.values()) + [summary])
    out.seek(0)
    return send_file(io.BytesIO(out.getvalue().encode()),mimetype='text/csv',as_attachment=True,download_name=fname+'.csv')

@app.route('/export/report')
@require_perm('companies_export')
def export_report():
    rt=request.args.get('type','all'); fmt=request.args.get('format','xlsx')
    conn=get_db(); today=dubai_today()
    
    if rt=='expiry':
        # Use the same logic as alerts page
        all_alerts = []
        
        # Trade licenses
        tl_rows = all_(conn,'''SELECT id,ac_code,client_name,mobile,whatsapp_number,
            trade_license_expiry,risk_status,region,account_manager,
            contact_person_name,contact_person_number
            FROM companies WHERE trade_license_expiry IS NOT NULL ORDER BY trade_license_expiry''')
        for r in tl_rows:
            d = days_left(r['trade_license_expiry'])
            if d is None: continue
            all_alerts.append({'ac_code':r['ac_code'],'client_name':r['client_name'],
                'doc_type':'Trade License','expiry_date':str(r['trade_license_expiry'])[:10],'days':d})
        
        # Address proofs
        ap_rows = all_(conn,'''SELECT id,ac_code,client_name,address_proof_type,address_proof_expiry,account_manager
            FROM companies WHERE address_proof_expiry IS NOT NULL ORDER BY address_proof_expiry''')
        for r in ap_rows:
            d = days_left(r['address_proof_expiry'])
            if d is None: continue
            all_alerts.append({'ac_code':r['ac_code'],'client_name':r['client_name'],
                'doc_type':'Address Proof','expiry_date':str(r['address_proof_expiry'])[:10],'days':d})
        
        # Passports from UBOs
        ubo_rows = all_(conn,'''SELECT u.person_name,u.passport_no,u.passport_expiry,
            u.emirates_id,u.emirates_id_expiry,
            c.id as company_id,c.client_name,c.ac_code
            FROM ubos u JOIN companies c ON u.company_id=c.id
            WHERE u.passport_expiry IS NOT NULL OR u.emirates_id_expiry IS NOT NULL''')
        for r in ubo_rows:
            if r.get('passport_expiry'):
                d = days_left(r['passport_expiry'])
                if d is not None:
                    all_alerts.append({'ac_code':r['ac_code'],'client_name':r['client_name'],
                        'doc_type':'Passport','expiry_date':str(r['passport_expiry'])[:10],'days':d})
            if r.get('emirates_id_expiry'):
                d = days_left(r['emirates_id_expiry'])
                if d is not None:
                    all_alerts.append({'ac_code':r['ac_code'],'client_name':r['client_name'],
                        'doc_type':'Emirates ID','expiry_date':str(r['emirates_id_expiry'])[:10],'days':d})
        
        # Clients documents
        client_rows = all_(conn,'''SELECT id,name,phone,account_number,
            passport_expiry,emirates_id_expiry FROM clients 
            WHERE passport_expiry IS NOT NULL OR emirates_id_expiry IS NOT NULL''')
        for r in client_rows:
            if r.get('passport_expiry'):
                d = days_left(r['passport_expiry'])
                if d is not None:
                    all_alerts.append({'ac_code':r.get('account_number','N/A'),'client_name':r['name'],
                        'doc_type':'Passport (Individual)','expiry_date':str(r['passport_expiry'])[:10],'days':d})
            if r.get('emirates_id_expiry'):
                d = days_left(r['emirates_id_expiry'])
                if d is not None:
                    all_alerts.append({'ac_code':r.get('account_number','N/A'),'client_name':r['name'],
                        'doc_type':'Emirates ID (Individual)','expiry_date':str(r['emirates_id_expiry'])[:10],'days':d})
        
        # Apply filters
        expiry_filter = request.args.get('expiry_filter', 'all')
        doc_type_filter = request.args.get('doc_type', 'all')
        search_filter = request.args.get('search', '').lower()
        
        filtered = []
        for row in all_alerts:
            # Expiry filter
            if expiry_filter == 'expired' and row['days'] >= 0:
                continue
            if expiry_filter == '30' and (row['days'] >= 30 or row['days'] < 0):
                continue
            if expiry_filter == '60' and (row['days'] >= 60 or row['days'] < 0):
                continue
            if expiry_filter == '90' and (row['days'] >= 90 or row['days'] < 0):
                continue
            
            # Doc type filter
            if doc_type_filter == 'trade' and 'Trade License' not in row.get('doc_type', ''):
                continue
            if doc_type_filter == 'address' and 'Address Proof' not in row.get('doc_type', ''):
                continue
            if doc_type_filter == 'passport' and 'Passport' not in row.get('doc_type', ''):
                continue
            if doc_type_filter == 'eid' and 'Emirates ID' not in row.get('doc_type', ''):
                continue
            
            # Search filter
            if search_filter and search_filter not in str(row.get('ac_code', '')).lower() and search_filter not in str(row.get('client_name', '')).lower():
                continue
            
            filtered.append(row)
        
        rows = filtered
    elif rt=='kyc':
        w='1=1'; p=[]
        df=request.args.get('from',''); dt=request.args.get('to','')
        if df: w+=' AND created_at >= ?'; p.append(df)
        if dt: w+=' AND created_at <= ?'; p.append(dt+' 23:59:59')
        rows=all_(conn,f'SELECT ac_code,client_name,kyc_status,kyc_expiry_date,doc_status,risk_status,verified_by,verified_date,followup_details FROM companies WHERE {w} ORDER BY kyc_status',p or None)
    elif rt=='risk':
        w='1=1'; p=[]
        df=request.args.get('from',''); dt=request.args.get('to','')
        if df: w+=' AND created_at >= ?'; p.append(df)
        if dt: w+=' AND created_at <= ?'; p.append(dt+' 23:59:59')
        rows=all_(conn,f'SELECT ac_code,client_name,risk_status,doc_status,kyc_status,pep,source_of_fund,moa,undertaking,registration_screening_tool,region,account_manager FROM companies WHERE {w} ORDER BY risk_status',p or None)
    elif rt=='ubo':
        rows=all_(conn,'''SELECT c.ac_code,c.client_name,u.person_name,u.position,u.share_percentage,
            u.nationality,u.residential_status,u.passport_no,u.passport_expiry,
            u.emirates_id,u.emirates_id_expiry,u.doc_status,u.verified_by,u.verified_date
            FROM ubos u JOIN companies c ON u.company_id=c.id
            ORDER BY c.ac_code,u.share_percentage DESC''')
    elif rt=='tasks':
        status_f=request.args.get('status','all')
        w='1=1'; p=[]
        if status_f and status_f!='all': w+=' AND t.status=?'; p.append(status_f)
        df=request.args.get('from',''); dt=request.args.get('to','')
        if df: w+=' AND t.created_at >= ?'; p.append(df)
        if dt: w+=' AND t.created_at <= ?'; p.append(dt+' 23:59:59')
        rows=all_(conn,f'''SELECT t.title,t.description,t.priority,t.status,t.due_date,
            t.created_at,u.name as assigned_to,cu.name as created_by,c.ac_code,c.client_name as company
            FROM tasks t LEFT JOIN users u ON t.assigned_to=u.id
            LEFT JOIN users cu ON t.created_by=cu.id
            LEFT JOIN companies c ON t.company_id=c.id
            WHERE {w} ORDER BY t.priority,t.due_date''', p or None)
    elif rt=='all':
        w='1=1'; p=[]
        df=request.args.get('from',''); dt=request.args.get('to','')
        if df: w+=' AND created_at >= ?'; p.append(df)
        if dt: w+=' AND created_at <= ?'; p.append(dt+' 23:59:59')
        rows=all_(conn,f'SELECT ac_code,client_name,ac_status,nature,type_of_client,mode_of_ac,region,telephone,mobile,email_id,contact_person_name,account_manager,risk_status,doc_status,kyc_status,trade_license_no,trade_license_expiry,address_proof_expiry,created_at FROM companies WHERE {w} ORDER BY client_name',p or None)
    else:
        rows=all_(conn,'SELECT * FROM companies ORDER BY client_name')
    conn.close()
    fname=f'report_{rt}_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
    if fmt=='xlsx' and HAS_XL:
        wb=openpyxl.Workbook(); ws=wb.active; ws.title=rt.upper()
        if rows: ws.append(list(rows[0].keys())); [ws.append([str(v) if v else '' for v in r.values()]) for r in rows]
        out=io.BytesIO(); wb.save(out); out.seek(0)
        return send_file(out,mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,download_name=fname+'.xlsx')
    out=io.StringIO(); w=csv.writer(out)
    if rows: w.writerow(rows[0].keys()); [w.writerow(list(r.values())) for r in rows]
    out.seek(0)
    return send_file(io.BytesIO(out.getvalue().encode()),mimetype='text/csv',as_attachment=True,download_name=fname+'.csv')

@app.route('/api/import/companies',methods=['POST'])
@require_perm('companies_import')
def api_import_companies():
    if not HAS_XL: return jsonify({'success':False,'error':'openpyxl not installed'}),500
    if 'file' not in request.files: return jsonify({'success':False,'error':'No file'}),400
    try:
        wb=openpyxl.load_workbook(request.files['file']); ws=wb.active
        headers=[str(c.value).strip().lower().replace(' ','_') if c.value else '' for c in ws[1]]
        imported=0; skipped=0; conn=get_db()
        for row in ws.iter_rows(min_row=2,values_only=True):
            if not any(row): continue
            rd={headers[i]:(str(row[i]).strip() if row[i] is not None else '') for i in range(min(len(headers),len(row)))}
            ac=rd.get('ac_code','').strip(); nm=rd.get('client_name','').strip()
            if not ac or not nm: continue
            if one(conn,'SELECT id FROM companies WHERE ac_code=?',(ac,)): skipped+=1; continue
            try:
                x(conn,'''INSERT INTO companies (ac_code,client_name,ac_status,risk_status,doc_status,
                    nature,type_of_client,region,telephone,mobile,email_id,trade_license_no,
                    trade_license_expiry,address_proof_expiry,kyc_status,account_manager,created_by)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (ac,nm,rd.get('ac_status','Active'),rd.get('risk_status','Unspecified'),
                     rd.get('doc_status','Incompleted'),rd.get('nature'),rd.get('type_of_client'),
                     rd.get('region'),rd.get('telephone'),rd.get('mobile'),rd.get('email_id'),
                     rd.get('trade_license_no'),rd.get('trade_license_expiry') or None,
                     rd.get('address_proof_expiry') or None,rd.get('kyc_status'),
                     rd.get('account_manager'),session.get('user_id')))
                imported+=1
            except: skipped+=1
        commit(conn); conn.close()
        return jsonify({'success':True,'imported':imported,'skipped':skipped})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

if __name__=='__main__':
    app.run(debug=False,host='0.0.0.0',port=int(os.getenv('PORT',8000)))


# ════════════════════════════════════════════════════════════════
# NEW FEATURES BLOCK
# ════════════════════════════════════════════════════════════════

# ── REGULAR TASKS → redirects to /tasks?tab=regular ──────────
@app.route('/regular-tasks')
@login_required
def regular_tasks():
    return redirect(url_for('tasks', tab='regular'))


@app.route('/api/regular-task/add', methods=['POST'])
@compliance_required
def api_add_regular_task():
    d = request.get_json()
    try:
        conn = get_db()
        x(conn, '''INSERT INTO regular_task_templates
            (title, description, frequency, assigned_role, assigned_user_id, created_by)
            VALUES (?,?,?,?,?,?)''',
          (d.get('title'), d.get('description'), d.get('frequency', 'daily'),
           d.get('assigned_role', 'all'), d.get('assigned_user_id') or None,
           session.get('user_id')))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        try: conn.close()
        except: pass
        return _fail(e)

@app.route('/api/regular-task/<int:id>/delete', methods=['POST'])
@compliance_required
def api_delete_regular_task(id):
    try:
        conn = get_db()
        x(conn, 'DELETE FROM regular_task_logs WHERE template_id=?', (id,))
        x(conn, 'DELETE FROM regular_task_templates WHERE id=?', (id,))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        try: conn.close()
        except: pass
        return _fail(e)

@app.route('/api/regular-task/<int:id>/log', methods=['POST'])
@login_required
def api_log_regular_task(id):
    d = request.get_json()
    try:
        conn = get_db()
        x(conn, '''INSERT INTO regular_task_logs (template_id, user_id, notes, status)
            VALUES (?,?,?,?)''',
          (id, session.get('user_id'), d.get('notes', ''), d.get('status', 'done')))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        try: conn.close()
        except: pass
        return _fail(e)

# ════════════════════════════════════════════════════════════
# ADDITIONAL TASKS
# ════════════════════════════════════════════════════════════
@app.route('/api/additional-task/add', methods=['POST'])
@login_required
def api_add_additional_task():
    d = request.get_json()
    try:
        conn = get_db()
        # Ensure table exists (Railway PG may not have run migration yet)
        try:
            if is_pg(conn):
                x(conn, '''CREATE TABLE IF NOT EXISTS additional_tasks (
                    id SERIAL PRIMARY KEY, title TEXT NOT NULL,
                    task_details TEXT, remarks TEXT,
                    from_datetime TIMESTAMP NOT NULL, to_datetime TIMESTAMP NOT NULL,
                    created_by INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            else:
                x(conn, '''CREATE TABLE IF NOT EXISTS additional_tasks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT, title TEXT NOT NULL,
                    task_details TEXT, remarks TEXT,
                    from_datetime TIMESTAMP NOT NULL, to_datetime TIMESTAMP NOT NULL,
                    created_by INTEGER NOT NULL,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
            commit(conn)
        except: pass
        x(conn, '''INSERT INTO additional_tasks
            (title, task_details, remarks, from_datetime, to_datetime, created_by)
            VALUES (?,?,?,?,?,?)''',
          (d.get('title'), d.get('task_details'), d.get('remarks'),
           d.get('from_datetime'), d.get('to_datetime'), session.get('user_id')))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

@app.route('/api/additional-task/<int:id>/delete', methods=['POST'])
@login_required
def api_delete_additional_task(id):
    try:
        conn = get_db()
        uid = session.get('user_id'); role = session.get('user_role')
        task = one(conn, 'SELECT created_by FROM additional_tasks WHERE id=?', (id,))
        if not task: return jsonify({'success': False, 'error': 'Not found'}), 404
        if role != 'admin' and task['created_by'] != uid:
            return jsonify({'success': False, 'error': 'Not allowed'}), 403
        x(conn, 'DELETE FROM additional_tasks WHERE id=?', (id,))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

# ── INTERNAL DOCUMENTS (ZEWER STAFF / COMPANY DOCS) ──────────
def _dl_url(url):
    """Cloudinary URL that forces download (fl_attachment)."""
    if url and '/upload/' in url:
        return url.replace('/upload/', '/upload/fl_attachment/', 1)
    return url

@app.route('/internal-docs')
@compliance_required
def internal_docs():
    conn = get_db()
    try:
        docs = all_(conn, '''SELECT d.*,u.name as added_by_name
            FROM internal_documents d LEFT JOIN users u ON d.added_by=u.id
            ORDER BY CASE WHEN d.expiry_date IS NULL THEN 1 ELSE 0 END, d.expiry_date ASC, d.doc_category, d.doc_name''')
    except: docs = []
    today = dubai_today()
    from collections import OrderedDict
    doc_list = []
    stats = {'total': 0, 'expired': 0, 'expiring_soon': 0, 'pending_file': 0, 'complete': 0}
    grouped = OrderedDict()   # category -> OrderedDict(person -> [docs])
    cat_counts = OrderedDict()
    for d in docs:
        dl = days_left(d['expiry_date'])
        est = exp_status(dl)
        has_file = bool(d.get('file_url'))
        row = {**d, 'days_left': dl, 'exp_status': est,
               'expiry_date': str(d['expiry_date']) if d['expiry_date'] else None,
               'has_file': has_file, 'download_url': _dl_url(d.get('file_url'))}
        doc_list.append(row)
        stats['total'] += 1
        if est == 'expired': stats['expired'] += 1
        elif est in ('critical', 'warning'): stats['expiring_soon'] += 1
        if has_file: stats['complete'] += 1
        else: stats['pending_file'] += 1
        cat = (d.get('doc_category') or 'Other').strip() or 'Other'
        person = (d.get('person_name') or '— Unassigned —').strip() or '— Unassigned —'
        grouped.setdefault(cat, OrderedDict()).setdefault(person, []).append(row)
        cat_counts[cat] = cat_counts.get(cat, 0) + 1
    users = all_(conn, 'SELECT id,name FROM users WHERE is_active=1 ORDER BY name')
    conn.close()
    return render_template('internal_docs.html', docs=doc_list, today=str(today), all_users=users,
                           stats=stats, grouped=grouped, cat_counts=cat_counts)

def _internal_doc_file(existing_public_id=None):
    """Upload an attached file to Cloudinary if present. Returns (file_url, file_name, public_id) or (None,None,None)."""
    f = request.files.get('file')
    if not f or not f.filename:
        return None, None, None
    if not HAS_CLD or not os.getenv('CLOUDINARY_CLOUD_NAME'):
        raise RuntimeError('File storage (Cloudinary) is not configured')
    file_bytes = f.read(); f.seek(0)
    r = cloudinary.uploader.upload(f, folder='zewer_crm/internal_docs', resource_type='auto',
                                   use_filename=True, unique_filename=True)
    save_local_copy(file_bytes, 'zewer_docs', f.filename)
    return r['secure_url'], f.filename, r['public_id']

@app.route('/api/internal-doc/add', methods=['POST'])
@compliance_required
def api_add_internal_doc():
    d = request.form if request.form else (request.get_json(silent=True) or {})
    try:
        file_url, file_name, public_id = _internal_doc_file()
        conn = get_db()
        x(conn, '''INSERT INTO internal_documents
            (doc_name, doc_category, person_name, issuing_authority, issue_date, expiry_date, notes,
             file_url, file_name, public_id, added_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
          (d.get('doc_name'), d.get('doc_category'), d.get('person_name'),
           d.get('issuing_authority'), d.get('issue_date') or None,
           d.get('expiry_date') or None, d.get('notes'),
           file_url, file_name, public_id, session.get('user_id')))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/internal-doc/<int:id>/edit', methods=['POST'])
@admin_pw_required
def api_edit_internal_doc(id):
    d = request.form if request.form else (request.get_json(silent=True) or {})
    try:
        file_url, file_name, public_id = _internal_doc_file()
        conn = get_db()
        if file_url:  # a new file was uploaded — replace, and remove the old one
            old = one(conn, 'SELECT public_id FROM internal_documents WHERE id=?', (id,))
            if old and old.get('public_id') and HAS_CLD and os.getenv('CLOUDINARY_CLOUD_NAME'):
                try: cloudinary.uploader.destroy(old['public_id'], resource_type='raw')
                except: pass
            x(conn, '''UPDATE internal_documents SET doc_name=?,doc_category=?,person_name=?,
                issuing_authority=?,issue_date=?,expiry_date=?,notes=?,file_url=?,file_name=?,public_id=?,
                updated_at=CURRENT_TIMESTAMP WHERE id=?''',
              (d.get('doc_name'), d.get('doc_category'), d.get('person_name'),
               d.get('issuing_authority'), d.get('issue_date') or None,
               d.get('expiry_date') or None, d.get('notes'),
               file_url, file_name, public_id, id))
        else:  # metadata only
            x(conn, '''UPDATE internal_documents SET doc_name=?,doc_category=?,person_name=?,
                issuing_authority=?,issue_date=?,expiry_date=?,notes=?,updated_at=CURRENT_TIMESTAMP
                WHERE id=?''',
              (d.get('doc_name'), d.get('doc_category'), d.get('person_name'),
               d.get('issuing_authority'), d.get('issue_date') or None,
               d.get('expiry_date') or None, d.get('notes'), id))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f'Error in %s: {e}', request.path)
        return _fail(e)

@app.route('/api/internal-doc/<int:id>/delete', methods=['POST'])
@admin_pw_required
def api_delete_internal_doc(id):
    try:
        conn = get_db()
        doc = one(conn, 'SELECT public_id FROM internal_documents WHERE id=?', (id,))
        if doc and doc.get('public_id') and HAS_CLD and os.getenv('CLOUDINARY_CLOUD_NAME'):
            try: cloudinary.uploader.destroy(doc['public_id'], resource_type='raw')
            except: pass
        x(conn, 'DELETE FROM internal_documents WHERE id=?', (id,))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

@app.route('/api/company/<int:cid>/docs-export')
@login_required
def api_export_company_docs(cid):
    if not HAS_XL:
        return "openpyxl not installed", 500
    from openpyxl.styles import Font, PatternFill, Alignment
    conn = get_db()
    co = one(conn, 'SELECT * FROM companies WHERE id=?', (cid,))
    ubos = all_(conn, 'SELECT * FROM ubos WHERE company_id=?', (cid,))
    conn.close()
    if not co:
        return "Company not found", 404
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documents"
    hf = PatternFill(start_color="1C1917", end_color="1C1917", fill_type="solid")
    hfont = Font(color="D97706", bold=True, size=11)
    headers = ['Document Type', 'Reference / Number', 'Issuing Authority', 'Expiry Date', 'Status']
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = hfont; c.fill = hf; c.alignment = Alignment(horizontal='center')
    today = dubai_today()
    def status(exp):
        if not exp: return '—'
        try:
            d = datetime.strptime(str(exp)[:10], '%Y-%m-%d').date()
            diff = (d - today).days
            if diff < 0: return 'EXPIRED'
            elif diff <= 30: return f'{diff}d - CRITICAL'
            elif diff <= 90: return f'{diff}d - WARNING'
            else: return f'{diff}d - OK'
        except: return '—'
    rows = [
        ['Trade License', co['trade_license_no'], co['issuing_authority'], co['trade_license_expiry'], status(co['trade_license_expiry'])],
        ['Address Proof', co['address_proof_type'], '—', co['address_proof_expiry'], status(co['address_proof_expiry'])],
        ['VAT Certificate', co['tax_no_trn'], '—', '—', co['vat_cert'] or '—'],
    ]
    for ubo in ubos:
        rows.append([f"Passport ({ubo['person_name']})", ubo['passport_no'], ubo['nationality'], ubo['passport_expiry'], status(ubo['passport_expiry'])])
        rows.append([f"Emirates ID ({ubo['person_name']})", ubo['emirates_id'], '—', ubo['emirates_id_expiry'], status(ubo['emirates_id_expiry'])])
    for i, row in enumerate(rows, 2):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=str(val) if val else '—')
            c.font = Font(size=10)
            c.alignment = Alignment(horizontal='left')
    for col, width in zip(['A','B','C','D','E'], [28, 22, 24, 16, 18]):
        ws.column_dimensions[col].width = width
    ws2 = wb.create_sheet("Company Info")
    info = [('AC Code', co['ac_code']), ('Company Name', co['client_name']),
            ('Status', co['ac_status']), ('Risk', co['risk_status']),
            ('Doc Status', co['doc_status']), ('Account Manager', co['account_manager']),
            ('KYC Status', co['kyc_status']), ('Region', co['region'])]
    for i, (k, v) in enumerate(info, 1):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True, color="D97706")
        ws2.cell(row=i, column=2, value=str(v) if v else '—')
    ws2.column_dimensions['A'].width = 22; ws2.column_dimensions['B'].width = 35
    out = io.BytesIO(); wb.save(out); out.seek(0)
    fname = f"{co['ac_code']}_documents.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)





# ════════════════════════════════════════════════════════════
# ROLE PERMISSIONS API
# ════════════════════════════════════════════════════════════
@app.route('/api/settings/role-permissions', methods=['GET'])
@login_required
def api_get_role_permissions():
    try:
        conn = get_db()
        row = one(conn, "SELECT value FROM app_settings WHERE key='role_permissions'")
        conn.close()
        if row:
            import json as _json
            return jsonify({'permissions': _json.loads(row['value'])})
        return jsonify({'permissions': None})
    except Exception as e:
        logger.error(f'{request.path}: {e}')
        return jsonify({'permissions': None})

@app.route('/api/settings/role-permissions', methods=['POST'])
@login_required
def api_save_role_permissions():
    if session.get('user_role') != 'admin':
        return jsonify({'success': False, 'error': 'Admin only'}), 403
    import json as _json
    d = request.get_json()
    try:
        conn = get_db()
        val = _json.dumps(d.get('permissions', {}))
        x(conn, "INSERT INTO app_settings (key,value) VALUES ('role_permissions',?) ON CONFLICT(key) DO UPDATE SET value=?,updated_at=CURRENT_TIMESTAMP",
          (val, val))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

# ════════════════════════════════════════════════════════════
# CLIENTS PAGE
# ════════════════════════════════════════════════════════════
@app.route('/clients')
@login_required
def clients():
    conn = get_db()
    today = dubai_today()

    # SQL-level filtering — only fetch what's needed
    s = request.args.get('search', '').strip()
    resident_f = request.args.get('resident', '')
    mode_f = request.args.get('mode', '')
    status_f = request.args.get('status', '')
    page = max(1, int(request.args.get('page', 1)))
    per_page = 100

    q = 'SELECT * FROM clients WHERE 1=1'; p = []
    if s:
        q += ' AND (name LIKE ? OR phone LIKE ? OR account_number LIKE ? OR email LIKE ?)'
        p += [f'%{s}%'] * 4
    # Status filter: Active/Inactive use ac_status; Disabled uses the out-of-scope flag.
    # Disabled records hidden by default unless explicitly requested (or status='all').
    if status_f == 'Disabled':
        q += ' AND disabled IS TRUE'
    elif status_f == 'all':
        pass
    elif status_f in ('Active','Inactive'):
        q += ' AND ac_status=? AND disabled IS NOT TRUE'; p.append(status_f)
    else:
        q += ' AND disabled IS NOT TRUE'
    if resident_f == 'resident':
        q += ' AND is_resident=?'; p.append(True)
    elif resident_f == 'non-resident':
        q += ' AND (is_resident IS NULL OR is_resident=?)'; p.append(False)
    if mode_f:
        q += ' AND mode_of_ac=?'; p.append(mode_f)

    total_count = cnt(conn, f"SELECT COUNT(*) FROM clients WHERE {q.split('WHERE')[1]}", p or None)
    q += ' ORDER BY name LIMIT ? OFFSET ?'
    p += [per_page, (page - 1) * per_page]

    rows = all_(conn, q, p)
    cl = []
    for c in rows:
        dob = c.get('date_of_birth')
        birthday_today = False
        days_to_birthday = None
        if dob:
            try:
                d = datetime.strptime(str(dob)[:10], '%Y-%m-%d').date()
                this_year = d.replace(year=today.year)
                if this_year < today:
                    this_year = d.replace(year=today.year+1)
                days_to_birthday = (this_year - today).days
                birthday_today = (d.month == today.month and d.day == today.day)
            except: pass
        def expiry_status(d):
            dl = days_left(d)
            if dl is None: return 'ok'
            if dl < 0: return 'expired'
            if dl <= 90: return 'warning'
            return 'ok'
        cl.append({**c,
                   'date_of_birth': str(dob)[:10] if dob else None,
                   'passport_expiry': str(c.get('passport_expiry'))[:10] if c.get('passport_expiry') else None,
                   'emirates_id_expiry': str(c.get('emirates_id_expiry'))[:10] if c.get('emirates_id_expiry') else None,
                   'kyc_expiry_date': str(c.get('kyc_expiry_date'))[:10] if c.get('kyc_expiry_date') else None,
                   'screening_date': str(c.get('screening_date'))[:10] if c.get('screening_date') else None,
                   'birthday_today': birthday_today, 'days_to_birthday': days_to_birthday,
                   'passport_expiry_status': expiry_status(c.get('passport_expiry')),
                   'eid_expiry_status': expiry_status(c.get('emirates_id_expiry')),
                   'kyc_expiry_status': expiry_status(c.get('kyc_expiry_date'))})
    conn.close()
    birthdays_today = [c for c in cl if c['birthday_today']]
    dd = dropdowns()
    total_pages = max(1, (total_count + per_page - 1) // per_page)
    return render_template('clients.html', clients=cl, birthdays_today=birthdays_today, today=str(today), days_left=days_left,
        modes=dd.get('MODE OF AC',[]), ac_statuses=dd.get('AC STATUS',[]), id_types=dd.get('ID TYPE',[]),
        kyc_statuses=dd.get('KYC STATUS',[]), risk_statuses=dd.get('RISK STATUS',[]),
        screening_statuses=dd.get('SCREENING REGISTRATION STATUS',[]),
        search=s, resident_filter=resident_f, mode_filter=mode_f, status_filter=status_f,
        page=page, total_pages=total_pages, total_count=total_count, per_page=per_page)

def _validate_kyc_expiry(kyc_expiry):
    """KYC expiry must not be more than 2 years from today."""
    if not kyc_expiry: return None
    try:
        d = datetime.strptime(str(kyc_expiry)[:10], '%Y-%m-%d').date()
        max_allowed = dubai_today().replace(year=dubai_today().year + 2)
        if d > max_allowed:
            return f"KYC expiry date cannot be more than 2 years from today (max: {max_allowed})"
    except Exception:
        return None
    return None

@app.route('/api/client/add', methods=['POST'])
@login_required
def api_add_client():
    d = request.get_json()
    err = _validate_kyc_expiry(d.get('kyc_expiry_date'))
    if err: return jsonify({'success': False, 'error': err}), 400
    try:
        conn = get_db()
        x(conn, '''INSERT INTO clients (name,phone,whatsapp_number,email,date_of_birth,
            profession,address,notes,nationality,passport_no,passport_expiry,
            emirates_id,emirates_id_expiry,address_proof,emirate,location,
            account_number,mode_of_ac,ac_status,id_type,pep_status,pep,kyc_status,kyc_expiry_date,
            risk_status,screening_status,screening_date,is_resident,created_by)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
          (d.get('name'), d.get('phone'), d.get('whatsapp_number'), d.get('email'),
           d.get('date_of_birth') or None, d.get('profession'), d.get('address'),
           d.get('notes'), d.get('nationality'), d.get('passport_no'),
           d.get('passport_expiry') or None, d.get('emirates_id'),
           d.get('emirates_id_expiry') or None, d.get('address_proof'),
           d.get('emirate'), d.get('location'),
           d.get('account_number'), d.get('mode_of_ac'), d.get('ac_status'),
           d.get('id_type'), d.get('pep_status'), d.get('pep'), d.get('kyc_status'),
           d.get('kyc_expiry_date') or None, d.get('risk_status'),
           d.get('screening_status'), d.get('screening_date') or None,
           d.get('is_resident', False),
           session.get('user_id')))
        commit(conn)
        client_id = lastid(conn)
        conn.close()
        return jsonify({'success': True, 'client_id': client_id})
    except Exception as e:
        try: conn.close()
        except: pass
        return _fail(e)

@app.route('/api/client/<int:id>/edit', methods=['POST'])
@admin_pw_required
def api_edit_client(id):
    d = request.get_json()
    err = _validate_kyc_expiry(d.get('kyc_expiry_date'))
    if err: return jsonify({'success': False, 'error': err}), 400
    try:
        conn = get_db()
        x(conn, '''UPDATE clients SET name=?,phone=?,whatsapp_number=?,email=?,date_of_birth=?,
            profession=?,address=?,notes=?,nationality=?,passport_no=?,passport_expiry=?,
            emirates_id=?,emirates_id_expiry=?,address_proof=?,emirate=?,location=?,
            account_number=?,mode_of_ac=?,ac_status=?,id_type=?,pep_status=?,pep=?,kyc_status=?,
            kyc_expiry_date=?,risk_status=?,screening_status=?,screening_date=?,is_resident=?,
            updated_at=CURRENT_TIMESTAMP WHERE id=?''',
          (d.get('name'), d.get('phone'), d.get('whatsapp_number'), d.get('email'),
           d.get('date_of_birth') or None, d.get('profession'), d.get('address'),
           d.get('notes'), d.get('nationality'), d.get('passport_no'),
           d.get('passport_expiry') or None, d.get('emirates_id'),
           d.get('emirates_id_expiry') or None, d.get('address_proof'),
           d.get('emirate'), d.get('location'),
           d.get('account_number'), d.get('mode_of_ac'), d.get('ac_status'),
           d.get('id_type'), d.get('pep_status'), d.get('pep'), d.get('kyc_status'),
           d.get('kyc_expiry_date') or None, d.get('risk_status'),
           d.get('screening_status'), d.get('screening_date') or None,
           d.get('is_resident', False), id))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        try: conn.close()
        except: pass
        return _fail(e)

@app.route('/api/client/<int:id>/delete', methods=['POST'])
@admin_pw_required
def api_delete_client(id):
    try:
        conn = get_db()
        x(conn, 'DELETE FROM clients WHERE id=?', (id,))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

# ════════════════════════════════════════════════════════════
# COMPANY GROUPS (admin panel)
# ════════════════════════════════════════════════════════════
@app.route('/api/groups', methods=['GET'])
@login_required
def api_get_groups():
    conn = get_db()
    groups = all_(conn, 'SELECT * FROM company_groups ORDER BY group_name')
    conn.close()
    return jsonify({'groups': groups})

@app.route('/api/group/add', methods=['POST'])
@compliance_required
def api_add_group():
    d = request.get_json()
    try:
        conn = get_db()
        x(conn, 'INSERT INTO company_groups (group_name, description) VALUES (?,?)',
          (d.get('group_name','').strip(), d.get('description','')))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

@app.route('/api/group/<int:id>/delete', methods=['POST'])
@compliance_required
def api_delete_group(id):
    try:
        conn = get_db()
        x(conn, 'DELETE FROM company_groups WHERE id=?', (id,))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

# ════════════════════════════════════════════════════════════
# APP SETTINGS (action password)
# ════════════════════════════════════════════════════════════
@app.route('/api/settings/action-password', methods=['POST'])
@compliance_required
def api_set_action_password():
    if session.get('user_role') != 'admin':
        return jsonify({'success': False, 'error': 'Admin only'}), 403
    d = request.get_json()
    pw = d.get('password', '').strip()
    if len(pw) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters'}), 400
    try:
        hashed = generate_password_hash(pw)  # salted, unlike the legacy sha256
        conn = get_db()
        x(conn, "INSERT INTO app_settings (key,value) VALUES ('action_password',?) ON CONFLICT(key) DO UPDATE SET value=?,updated_at=CURRENT_TIMESTAMP",
          (hashed, hashed))
        commit(conn); conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return _fail(e)

@app.route('/api/settings/verify-action-password', methods=['POST'])
@login_required
def api_verify_action_password():
    d = request.get_json()
    pw = d.get('password', '')
    try:
        conn = get_db()
        s = one(conn, "SELECT value FROM app_settings WHERE key='action_password'")
        conn.close()
        if not s:
            return jsonify({'success': True, 'note': 'No action password set'})
        return jsonify({'success': _action_pw_matches(pw, s.get('value'))})
    except Exception as e:
        return _fail(e)

# ════════════════════════════════════════════════════════════
# DATA BACKUP — full Excel export
# ════════════════════════════════════════════════════════════
@app.route('/admin/backup')
@login_required
def admin_backup():
    if session.get('user_role') != 'admin':
        return redirect(url_for('dashboard'))
    if not HAS_XL:
        return "openpyxl not installed", 500
    from openpyxl.styles import Font, PatternFill, Alignment
    conn = get_db()
    wb = openpyxl.Workbook()
    gold_fill = PatternFill(start_color="1C1917", end_color="1C1917", fill_type="solid")
    gold_font = Font(color="D97706", bold=True, size=10)

    def add_sheet(wb, name, rows, first=False):
        ws = wb.active if first else wb.create_sheet(name)
        ws.title = name
        if not rows: return ws
        headers = list(rows[0].keys())
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = gold_font; c.fill = gold_fill
        for ri, row in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                val = row.get(h)
                ws.cell(row=ri, column=ci, value=str(val) if val is not None else '')
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(40, max(12, len(str(col[0].value or ''))+4))
        ws.freeze_panes = 'A2'
        return ws

    add_sheet(wb, 'Companies',     all_(conn, 'SELECT * FROM companies ORDER BY ac_code'), first=True)
    add_sheet(wb, 'UBOs',          all_(conn, 'SELECT * FROM ubos ORDER BY company_id'))
    add_sheet(wb, 'Tasks',         all_(conn, 'SELECT * FROM tasks ORDER BY created_at DESC'))
    add_sheet(wb, 'Users',         all_(conn, 'SELECT id,email,name,role,is_active,created_at FROM users ORDER BY name'))
    add_sheet(wb, 'Clients',       all_(conn, 'SELECT * FROM clients ORDER BY name'))
    add_sheet(wb, 'InternalDocs',  all_(conn, 'SELECT * FROM internal_documents ORDER BY expiry_date'))
    add_sheet(wb, 'RegularTasks',  all_(conn, 'SELECT * FROM regular_task_templates'))
    add_sheet(wb, 'TaskLogs',      all_(conn, 'SELECT * FROM regular_task_logs ORDER BY logged_at DESC'))
    add_sheet(wb, 'Groups',        all_(conn, 'SELECT * FROM company_groups'))
    add_sheet(wb, 'Dropdowns',     all_(conn, 'SELECT * FROM dropdowns WHERE is_active=1 ORDER BY field_name,value'))
    add_sheet(wb, 'AdditionalTasks', all_(conn, 'SELECT * FROM additional_tasks ORDER BY from_datetime DESC'))
    conn.close()

    out = io.BytesIO()
    wb.save(out); out.seek(0)
    fname = f"ZewerCRM_Backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=fname)

# ════════════════════════════════════════════════════════════
# RESTORE FROM BACKUP
# ════════════════════════════════════════════════════════════
def _real_columns(conn, table):
    """Actual column names for a table — used to whitelist restore columns so an
    uploaded backup file cannot inject SQL through crafted header names."""
    try:
        if is_pg(conn):
            rows = all_(conn, "SELECT column_name AS name FROM information_schema.columns WHERE table_name=?", (table,))
        else:
            rows = all_(conn, f"PRAGMA table_info({table})")
        return {r['name'] for r in rows}
    except Exception:
        return set()

@app.route('/admin/restore', methods=['POST'])
@login_required
def admin_restore():
    if session.get('user_role') != 'admin':
        return jsonify({'success': False, 'error': 'Admin only'}), 403
    if not HAS_XL:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    try:
        f = request.files['file']
        wb = openpyxl.load_workbook(f)
        conn = get_db()
        results = {}

        def restore_sheet(sheet_name, table, pk='id', skip_cols=None):
            if sheet_name not in wb.sheetnames:
                results[sheet_name] = 'Sheet not found — skipped'
                return
            ws = wb[sheet_name]
            headers = [cell.value for cell in ws[1]]
            if not headers or not headers[0]:
                results[sheet_name] = 'Empty — skipped'
                return
            skip = set(skip_cols or [])
            # Whitelist: only real columns of this table, and only safe identifiers.
            # This is what makes the dynamic INSERT below injection-proof.
            valid = _real_columns(conn, table)
            use_headers = [h for h in headers
                           if h and h not in skip
                           and isinstance(h, str)
                           and re.match(r'^[A-Za-z_][A-Za-z0-9_]*$', h)
                           and (not valid or h in valid)]
            inserted = 0; skipped = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                rd = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
                pk_val = rd.get(pk)
                if pk_val is None: continue
                # Check if record already exists
                try:
                    existing = one(conn, f'SELECT {pk} FROM {table} WHERE {pk}=?', (pk_val,))
                    if existing: skipped += 1; continue
                except: skipped += 1; continue
                cols = [h for h in use_headers if h != pk]
                vals = [rd.get(h) or None for h in cols]
                placeholders = ','.join(['?' for _ in cols])
                try:
                    x(conn, f"INSERT INTO {table} ({','.join(cols)}) VALUES ({placeholders})", vals)
                    inserted += 1
                except: skipped += 1
            commit(conn)
            results[sheet_name] = f'{inserted} restored, {skipped} skipped'

        restore_sheet('Companies',    'companies',              pk='id')
        restore_sheet('UBOs',         'ubos',                   pk='id')
        restore_sheet('Clients',      'clients',                pk='id')
        restore_sheet('Tasks',        'tasks',                  pk='id')
        restore_sheet('InternalDocs', 'internal_documents',     pk='id')
        restore_sheet('RegularTasks', 'regular_task_templates', pk='id')
        restore_sheet('TaskLogs',     'regular_task_logs',      pk='id')
        restore_sheet('Groups',       'company_groups',         pk='id')
        restore_sheet('Dropdowns',    'dropdowns',              pk='id')
        restore_sheet('AdditionalTasks', 'additional_tasks',   pk='id')
        # Skip Users sheet — don't overwrite passwords/accounts from backup

        conn.close()
        return jsonify({'success': True, 'results': results})
    except Exception as e:
        return _fail(e)

# ════════════════════════════════════════════════════════════
# GROUPS API for companies form
# ════════════════════════════════════════════════════════════
@app.route('/api/groups/list')
@login_required
def api_groups_list():
    conn = get_db()
    try:
        groups = all_(conn, 'SELECT group_name FROM company_groups ORDER BY group_name')
    except: groups = []
    conn.close()
    return jsonify([g['group_name'] for g in groups])







# ════════════════════════════════════════════════════════════
# CLIENT DOCUMENTS (Cloudinary)
# ════════════════════════════════════════════════════════════
@app.route('/api/client/<int:cid>/documents')
@login_required
def api_client_documents(cid):
    conn = get_db()
    docs = all_(conn, 'SELECT * FROM client_documents WHERE client_id=? ORDER BY created_at DESC', (cid,))
    conn.close()
    return jsonify(docs)

@app.route('/api/client/<int:cid>/upload', methods=['POST'])
@login_required
def api_client_upload_document(cid):
    if 'file' not in request.files: return jsonify({'success':False,'error':'No file'}),400
    file = request.files['file']
    if not file.filename: return jsonify({'success':False,'error':'No filename'}),400
    if not HAS_CLD or not os.getenv('CLOUDINARY_CLOUD_NAME'):
        return jsonify({'success':False,'error':'Cloudinary not configured'}),500
    try:
        file_bytes = file.read(); file.seek(0)
        r = cloudinary.uploader.upload(file, folder=f'zewer_crm/client_{cid}', resource_type='auto',
                                        use_filename=True, unique_filename=True)
        conn = get_db()
        x(conn, 'INSERT INTO client_documents (client_id,doc_type,file_name,file_url,public_id,uploaded_by) VALUES (?,?,?,?,?,?)',
          (cid, request.form.get('doc_type','General'), file.filename, r['secure_url'], r['public_id'], session.get('user_id')))
        commit(conn); conn.close()
        save_local_copy(file_bytes, f'individuals/client_{cid}', file.filename)
        return jsonify({'success':True,'url':r['secure_url'],'name':file.filename})
    except Exception as e:
        return _fail(e)

@app.route('/api/client-document/<int:did>/delete', methods=['POST'])
@login_required
def api_delete_client_document(did):
    try:
        conn = get_db()
        doc = one(conn, 'SELECT public_id FROM client_documents WHERE id=?', (did,))
        if doc and doc.get('public_id') and HAS_CLD and os.getenv('CLOUDINARY_CLOUD_NAME'):
            try: cloudinary.uploader.destroy(doc['public_id'], resource_type='raw')
            except: pass
        x(conn, 'DELETE FROM client_documents WHERE id=?', (did,))
        commit(conn); conn.close()
        return jsonify({'success':True})
    except Exception as e:
        return _fail(e)

@app.route('/client/<int:id>')
@login_required
def client_detail(id):
    conn = get_db()
    c = one(conn, 'SELECT * FROM clients WHERE id=?', (id,))
    if not c:
        conn.close()
        return redirect(url_for('clients'))
    docs = all_(conn, 'SELECT * FROM client_documents WHERE client_id=? ORDER BY created_at DESC', (id,))
    conn.close()
    dob = c.get('date_of_birth')
    c['date_of_birth'] = str(dob)[:10] if dob else None
    pe = c.get('passport_expiry')
    c['passport_expiry'] = str(pe)[:10] if pe else None
    ee = c.get('emirates_id_expiry')
    c['emirates_id_expiry'] = str(ee)[:10] if ee else None
    ke = c.get('kyc_expiry_date')
    c['kyc_expiry_date'] = str(ke)[:10] if ke else None
    sd = c.get('screening_date')
    c['screening_date'] = str(sd)[:10] if sd else None
    c['kyc_expiry_status'] = exp_status(days_left(ke)) if ke else 'unknown'
    return render_template('client_detail.html', client=c, documents=docs)









